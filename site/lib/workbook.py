"""The Excel deliverable: a workbook of what the user is already looking at.

This is the feature closest to the licence line, so the scope rule is structural rather
than a promise. Every sheet is a query against one serving view, bounded by the same filter
scope the pages use, and there is no control anywhere that widens it — no "all seasons", no
full-corpus dump, no raw layer. A request that amounts to "pull the whole database into
Excel" has no code path here to say yes with.

Three properties carry most of the weight:

  Attribution is structural.   Every sheet writes the CFBD credit into A1 before anything
                               else, and a prediction sheet writes the model disclaimer into
                               A2. Neither is optional per sheet, because a workbook leaves
                               the site and travels on its own.

  Numbers are numbers.         Cells are written as floats with an Excel number format
                               derived from the same precision table the site renders with,
                               so a column reads identically in both places and can still be
                               summed. A workbook of numeric strings is a screenshot with
                               extra steps.

  A missing sheet is named.    A view that returns nothing is omitted and recorded on the
                               index with the reason. An empty tab looks like a bug in the
                               export; a line saying "Edges — no rows in this scope" is an
                               answer.
"""
import io
import os
import re
import zipfile
from datetime import datetime, timezone
from typing import Callable, Dict, List, NamedTuple, Optional
from urllib.parse import quote

import pandas as pd

from lib import fmt, metrics
from lib.query import query

CFBD_CREDIT = ("Data from CollegeFootballData.com. Used under their terms; attribution is "
               "optional under those terms and provided anyway.")
MODEL_DISCLAIMER = (
    "Predictions are cfdb's own, built on a commercially licensed training pack. They are "
    "NOT CollegeFootballData.com predictions and CFBD does not endorse them. Figures are "
    "held-out backtests, not realised betting results. Nothing here is betting advice.")

# THE NOTE BLOCK IS AS LONG AS IT NEEDS TO BE, THEN EXACTLY ONE BLANK ROW, THEN THE HEADER.
#
# R-181, and Marc calls it a global requirement: "there needs to be 1 empty row between
# disclaimer info and the header row of the dataset — global requirement to help play to how
# sorting works in Excel."
#
# It used to be four constants — credit 1, disclaimer 2, header 4, data 5 — which gives
# exactly one blank row on a sheet WITH the model disclaimer and TWO on a sheet without it.
# Excel treats a blank row as the end of a region, so the second blank is not cosmetic: it
# changes what Ctrl+A and a header-click select.
#
# WHAT THIS COSTS, SAID OUT LOUD. `ROW_HEADER = 4` existed so "the freeze pane, autofilter and
# header row are at the same address throughout", and a computed header gives that up. The
# replacement is stronger, which is why the trade is worth taking: with Excel Tables (R-182)
# every data range is a NAMED OBJECT that Excel resolves for itself, so nothing downstream
# needs the address at all. Anyone tempted to restore the constant should restore the Tables
# first and then discover there is nothing left for it to do.
ROW_CREDIT = 1
BLANK_ROWS_BEFORE_HEADER = 1


def header_row(note_lines: int) -> int:
    """The 1-based row the header sits on, given how many note lines precede it."""
    return note_lines + BLANK_ROWS_BEFORE_HEADER + 1


def first_data_row(note_lines: int) -> int:
    return header_row(note_lines) + 1


# Retained ONLY because `_write_index` writes a fixed two-line note block of its own and the
# rest of the module reads these for the index's own layout. Data sheets compute their rows.
ROW_DISCLAIMER = 2

# THE ROW CAP, NAMED, AND RAISED — R-196.
#
# It was `limit 400` written into three of the seven queries as a literal, which is two
# problems. Nothing named it, so nothing could report it; and 400 is below the size of the
# thing a user most obviously asks for. Measured on the serving database:
#
#     one FBS week          ~55 games   (2025 regular: avg 55.5, range 51-96)
#     one FBS season        ~900
#     one season, all divisions        3,745      <- 400 returned 11% of it
#
# 5,000 clears a full all-divisions season with room, which is the widest scope the filter
# bar can express. It is a CAP, not a target: the point is that the file cannot become
# unbounded, not that it should approach this.
ROW_CAP = 5000


def site_base_url() -> Optional[str]:
    """The public origin of the site, or None if nobody has told us what it is.

    R-183. `CFDB_SITE_HOST` is set on the droplet and in `.env`, and the tunnel's hostname
    lives in the Cloudflare dashboard — so it is reachable at build time and must never be
    guessed. **If it is unset the workbook ships with no hyperlinks and the Index says so**,
    because a file full of `http://None/...` is worse than a file with none: it is the R-151
    shape exactly, a link that resolves in dev and 404s in what the user downloaded.
    """
    host = (os.getenv("CFDB_SITE_HOST") or "").strip().rstrip("/")
    if not host:
        return None
    return host if host.startswith(("http://", "https://")) else f"https://{host}"


def _scoped_query(season, week, season_type, conference, division, **extra) -> str:
    """The query string that CARRIES THE SCOPE FORWARD, as `GameScope.link()` does.

    A link that drops the season is the defect that made choosing 2025 and clicking a team
    return a 2026 page. It is worse in a workbook, which is read weeks later.
    """
    carried = {"season": season, "week": week,
               "season_type": None if season_type == "regular" else season_type,
               "conference": conference,
               "division": None if division == "fbs" else division}
    carried.update(extra)
    pairs = [f"{k}={quote(str(v))}" for k, v in carried.items() if v is not None]
    return "&".join(pairs)


# Excel's own rules for a table's displayName, enforced rather than discovered: unique in the
# workbook, starts with a letter or underscore, no spaces, nothing Excel reads as an operator,
# and it must not look like a cell reference. A duplicate or an illegal name is a repair
# prompt, not an exception — the file writes cleanly and Excel refuses it.
CELL_REFERENCE = re.compile(r"^[A-Za-z]{1,3}[0-9]+$")


def _team_url(base, slug, **scope):
    if not base or not slug or (isinstance(slug, float) and pd.isna(slug)):
        return None
    # A SLUG HAS NO SPACES, AND THIS GUARD EXISTS BECAUSE A DISPLAY NAME IS TRUTHY (R-289).
    #
    # `Sheet.hyperlinks` used to find the slug field by string surgery on Schedule's naming
    # convention. On a sheet whose display column is plainly `team`, that surgery is a no-op
    # and hands the TEAM NAME to this function — "Western Illinois" is truthy, clears the
    # check above, and produces `/team?...&team=Western%20Illinois`: a hyperlink that is
    # written, looks right in the cell, and resolves to nothing. The convention is gone now
    # and every slug is stated, but a wrong link is worse than no link and this is cheap.
    if isinstance(slug, str) and (" " in slug or slug != slug.strip()):
        return None
    query = _scoped_query(**scope, team=slug)
    return f"{base}/team?{query}" if query else f"{base}/team"


def _matchup_url(base, game_id, **scope):
    """A link is a convenience; it must never be able to fail a build.

    `int()` on a game id that is not one raises, and this runs once per row — so a single
    malformed value would take out the whole workbook to save one hyperlink. No link is the
    right answer for a row we cannot address.
    """
    if not base or game_id is None or (isinstance(game_id, float) and pd.isna(game_id)):
        return None
    try:
        identifier = int(game_id)
    except (TypeError, ValueError):
        return None
    query = _scoped_query(**scope, game_id=identifier)
    return f"{base}/matchup?{query}" if query else f"{base}/matchup"


def _weekday(record):
    """Thu / Fri / Sat — the thing a reader plans around, which a datetime does not answer
    without a formula."""
    value = record.get("game_date")
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return pd.Timestamp(value).strftime("%a")
    except (ValueError, TypeError):
        return None


# ==========================================================================================
# R-219. THE SITE'S MARKS, IN THE SHEET.
#
# Spec §3.4 said a spreadsheet has no icon convention and no tooltip, so these get spelled-out
# words. Marc looked at the built file and overrode it. This builds what he asked and makes it
# defensible the way the site made its own icon-only exception defensible.
#
# The site's system maps onto text-presentation Unicode almost exactly, and that is the piece
# that makes it work at all: R-141 chose SHAPE PLUS FILL — circle / square / diamond, filled or
# open — precisely so every mark is self-identifying without colour, and those shapes exist as
# geometric characters.
#
# EVERY MARK SHARES ONE EAST-ASIAN-WIDTH CLASS, AND THAT IS WHY THE COLUMN LOOKS STRAIGHT.
#
# Marc: "Winner Covered values (icons) doesn't seem centered horizontally." Every cell in that
# column IS centred — checked on all 83 rows of a real file, header included. What was ragged
# was the GLYPHS: ■ □ ○ ● ▲ ▽ are all East-Asian-Width AMBIGUOUS, and the push mark was ◨
# (U+25E8), which is NEUTRAL and rare enough that most fonts have no glyph for it. A fallback
# font brings its own advance width, so those cells sat at a different offset from the rest
# and the column read as crooked rather than as centred.
#
# ▣ (U+25A3) is Ambiguous like the others, lives in the same Geometric Shapes block, and says
# "neither filled nor empty" — which is what a push is. A test pins the width class so the
# next mark cannot reintroduce the problem.
#
# GEOMETRIC SHAPES, NOT EMOJI. R-141 and R-175 both turned on this: an emoji-presentation
# character has no fixed baseline or size across platforms, and a workbook is opened on more
# platforms than a web page is. Every mark below is text-presentation.
#
# THE COST, SO IT IS A DECISION AND NOT A SURPRISE: the filter dropdown on these three columns
# now lists shapes. A reader who wants "show me every upset" picks a symbol from a list rather
# than reading the word "upset" — a real reduction in the one thing a spreadsheet is better at
# than a page. THE REVERT IS THESE THREE DICTIONARIES; nothing else knows about it.
# FILLED MEANS IT HAPPENED, OPEN MEANS IT DID NOT — R-141's shape-plus-fill system, which is
# what makes each mark self-identifying without colour.
#
# "Favorite won" was first drawn as an em dash, and that was a defect: the no-data mark is
# an EN dash, and at 11pt the two are indistinguishable. "The favorite won" and "cfdb holds
# no closing line" are opposite claims and must not look the same. An open circle against a
# filled one says it in the same visual language as the rest.
# THE MARK REPEATS WITH THE LEVEL (Marc, round 4). One circle, two, three.
#
# The site distinguishes the three levels by SHADE, which a spreadsheet cannot carry into a
# filter dropdown — "show me every blowout" would mean picking a colour. Repetition carries
# the same ordering, sorts correctly (●● sorts after ●), and each value is its own entry in
# the dropdown. It is also readable without the legend in a way that shade is not.
# PUSH IS AN EQUALS SIGN, NOT A THIRD KIND OF SQUARE.
#
# It was ▣ — a white square containing a small black one — and Marc is right that it does not
# read: at 12pt, next to ■, it is a filled square with a hairline round it. The fill states
# are already spoken for (filled = it happened, open = it did not), so a push needs to leave
# that language rather than find a third position inside it.
#
# ══ says "equal", which is exactly what a push is: the result landed on the number and
# neither side won. DOUBLED, because a single ═ is one short bar and reads as a dash at a
# glance — two of them read unmistakably as an equals sign. The same trick as the upset
# levels: repeat the character rather than find a different one, so the metrics cannot drift.
#
# Each glyph measures 0.6001 em in Arial Unicode MS, identical to every other mark, and is
# East-Asian-Width Ambiguous like the rest, so the column still lines up.
#
# It comes from Box Drawing rather than Geometric Shapes, which is a deliberate exception to
# the one-block rule below — the block rule exists to keep METRICS consistent, and this glyph
# is metrically identical. The test asserts the width, and allows this one by name.
PUSH_GLYPH = "═"
PUSH_MARK = PUSH_GLYPH * 2

UPSET_MARKS = {"none": "○", "upset": "●", "big": "●●", "blowout": "●●●"}
COVER_MARKS = {"yes": "■", "no": "□", "push": PUSH_MARK, "pending": "·"}
# Under is UNFILLED AND RED (Marc, round 6). Filled/open already carries "it happened / it
# did not"; red carries the direction a bettor cares about. ▽ is the open form of ▲, so the
# pair still reads as one system rather than two.
OVER_MARKS = {"yes": "▲", "no": "▽", "push": PUSH_MARK, "pending": "·"}
NO_DATA_MARK = "–"        # the site's own mark for "we hold nothing here"

# THE OPEN MARKS CARRY A COLOUR, AND THE COLOUR WAS MEASURED RATHER THAN PICKED.
#
# Shape alone works on the site because the marks sit inches apart in a legend the reader has
# just seen. In a spreadsheet they sit in a column of 80 rows and ○ against ● is a small
# difference to scan. Marc asked for the OPEN form to be the coloured one — "it seems to be
# more contrasting than changing the filled version to a blue or green" — and he is right:
# the filled shapes are the common case, and colouring the exception is what makes it pop.
#
# Literal burnt sienna (#E97451) MEASURES 2.97:1 AGAINST WHITE and fails WCAG AA outright.
# #B7410E is the same family and measures 5.56:1. Contrast checked, not eyeballed, because
# this project has already had to fix a glyph that was 3.6:1.


OPEN_MARK_COLOUR = "FFB7410E"

# RED, for the two marks that mean "not the good outcome" — Under, and a push.
#
# It started as UNDER_COLOUR when only one mark used it. Renamed rather than reused under the
# old name: a constant called "under" that also colours the push mark is the kind of small
# lie that makes the next reader distrust every other name in the file.
#
# Measured the same way as the burnt sienna: #C00000 is 5.89:1 against white, comfortably
# past AA, and is Excel's own "dark red" so it does not look foreign beside the rest.
RED_MARK_COLOUR = "FFC00000"

# BLUE, for the push. It is not a loss and it is not a win, so it should not borrow the
# colour of either. #0070C0 is 5.15:1 against white — past AA — and is Excel's own standard
# blue, the same reasoning that picked its standard dark red above: a reader who has ever
# used the fill palette has seen this exact colour.
BLUE_MARK_COLOUR = "FF0070C0"

MARK_FONT_SIZE = 12

# THE MARK CELLS NAME THEIR OWN FONT, AND THIS IS WHY THE COLUMN LOOKED CROOKED.
#
# Marc reported Winner covered as not horizontally centred. Every cell in it IS centred —
# checked on all 83 rows of a real file. The problem was never alignment; it was that the
# glyphs were coming from TWO DIFFERENT FONTS.
#
# Measured against the actual font files on the machine Excel runs on:
#
#     glyph   Calibri   Aptos    Arial   Arial Unicode MS
#     ●       0.604     0.749    0.604   0.6001
#     ○       0.550     0.749    0.604   0.6001
#     ■       MISSING   MISSING  0.604   0.6001
#     □       0.604     0.750    0.604   0.6001
#     ▲       MISSING   MISSING  0.990   0.6001
#     ▽       MISSING   MISSING  MISSING 0.6001
#     ▣       MISSING   MISSING  MISSING 0.6001
#
# Calibri and Aptos — Excel's old and current defaults — HAVE NO FILLED SQUARE. So □ was
# drawn by Calibri at 0.604 em and ■ by whatever font macOS substituted, at its own width,
# in the same column. Centring cannot rescue two different advance widths. Calibri's own
# ● and ○ differ by 0.054 em too, which is the same defect in miniature.
#
# Arial Unicode MS carries all seven at EXACTLY 0.6001 em. Naming it means every mark in a
# column is measured the same way.
#
# Degrading safely matters as much as being right here: if a reader does not have this font,
# Excel substitutes ONE font for the whole run, so the marks still share a width — which is
# strictly better than today, where the substitution happens per glyph.
MARK_FONT_NAME = "Arial Unicode MS"

# What the Matchup cell says. The URL is the hyperlink, not the text.
URL_CELL_LABEL = "Matchup"

# WHICH MARK IS DRAWN IN WHICH COLOUR. A mapping rather than a set, because there are two
# colours now and a set could only answer "is it coloured", not "which colour" — and the
# legend has to render each glyph exactly as the column does or it is a picture of a
# different mark.
MARK_COLOURS = {
    "○": OPEN_MARK_COLOUR,      # the favorite won
    "□": OPEN_MARK_COLOUR,      # did not cover
    "▽": RED_MARK_COLOUR,        # under
    PUSH_MARK: BLUE_MARK_COLOUR,  # push — neither side won, so neither side's colour
}

# Named for what it IS — the marks that carry a colour — rather than for "open", which the
# push mark is not.
COLOURED_MARKS = set(MARK_COLOURS)

# THE THRESHOLDS COME FROM THE DATA, so there is nothing to hold here.
#
# They were read from dbt_project.yml in this module until R-224 put them on `srv_game` as
# columns, which is the only version that is correct inside the site image — the repo root is
# outside the build context, so the file read was running on a fallback. `mark_legend(df)`
# below takes the frame the workbook is already writing.

# The Index legend. R-026's icon-only exception on the SITE is defensible because R-102's
# legend explains it once — that is the stated reason in the code. A workbook travels further
# and has no tooltip at all, so the same exception needs the same support or it is just
# undecodable symbols. This block is not optional.


def mark_legend(df=None) -> list:
    """The legend rows, phrased from the thresholds THIS workbook's data carries.

    A function rather than a constant because the numbers are columns now (R-224). The bands
    come from `metrics.upset_bands`, the same call the Schedule page makes, so the two legends
    cannot differ by a word or by a number — which they did, for months.
    """
    band_1, band_2, band_3 = metrics.upset_criteria(*metrics.from_frame(df))
    return [
        # A LEGEND MUST DEFINE, NOT RESTATE. The first version said "a level 2 upset", which tells
        # a reader who has never seen the site precisely nothing. The thresholds are dbt vars
        # and they arrive as COLUMNS on srv_game (R-224), so they are stated here as the
        # numbers a reader can check against the score in the same row — and they follow the
        # warehouse if it ever changes them.
        #
        # BY MARGIN OF VICTORY, NOT BY THE SIZE OF THE SWING, which is Marc's own call recorded in
        # srv_game.sql: an 8-point favorite losing by 5 is Level 1, not the Level 2 the 13-point
        # swing would make it.
        ("Upset level", "○", "the favorite won"),
        ("Upset level", "●",
         f"Level 1 upset — {band_1}"),
        ("Upset level", "●●",
         f"Level 2 upset — {band_2}"),
        ("Upset level", "●●●",
         f"Level 3 upset — {band_3}"),
        ("Upset level", "–",
         "no closing line, so there was no favorite to upset"),
        ("Winner covered", "■", "the winner also covered the closing spread"),
        ("Winner covered", "□", "the winner did not cover"),
        ("O/U result", "▲", "the two scores together went OVER the closing total"),
        ("O/U result", "▽", "they stayed UNDER it"),
        ("Any of the three", PUSH_MARK,
         "push — the result landed exactly on the number, so neither side won the bet"),
        ("Any of the three", "·", "not settled yet"),
        ("Any of the three", "–", "cfdb holds no closing line for this game"),
    ]


# One shared alignment object rather than one per cell: openpyxl stores styles by identity
# and a fresh Alignment for each of 80 rows inflates the styles table for no benefit.
CENTRED = None      # built lazily in `build`, where Alignment is imported

# A TEXT COLUMN WITH FEWER THAN THIS MANY DISTINCT VALUES IS A CATEGORY, NOT PROSE.
#
# Marc: "for any text fields with cardinality <5 make them center aligned". The reasoning
# holds — Yes/No, Final/Scheduled, a verdict word — those are labels, and a label centred in
# a narrow column reads as a value rather than as the start of a sentence. A team name or a
# venue is prose and stays left, where the eye finds the first letter.
#
# MEASURED FROM THE DATA IN THIS EXPORT, which has a consequence worth knowing: a column can
# be centred in a one-week file and left-aligned in a full-season one, because the season saw
# a fifth distinct value. `Condition` is the likely one — a week of clear weather has three
# values and a season has a dozen. Declaring the set instead would be stable but would need
# maintaining by hand, and would be wrong the first time a column changed. Alignment is
# cosmetic and the measurement is honest, so measurement wins; if the wobble annoys, the fix
# is a declared set and it is one dictionary.
LOW_CARDINALITY = 5

# Columns centred BY NAME, whatever their cardinality says.
#
# A won-lost record is a short token, not prose — "1-0" left-aligned in a 10-wide column sits
# hard against the left edge with the error tag right beside it, which is what Marc is
# describing. Centring gives it air on both sides. The cardinality rule cannot reach these:
# a week of play produces well over five distinct records.
# The two cover verdicts join them for a related reason (R-262). A verdict column is a
# CATEGORY whatever this week happened to contain, and the cardinality rule would decide it
# by accident: Yes / No / Push / Pending / – is five distinct values and the threshold is
# "fewer than five", so a week with a push and a game without a line would come out left
# while the week before came out centred. Worse, the closing and opening columns could
# disagree with each other — the same coin flip the away_/home_ harmonisation below exists to
# stop. Naming them settles it once.
ALWAYS_CENTRED_LABELS = {"Away record", "Home record", "Covered", "Covered open"}

# ...AND ENOUGH ROWS TO TELL A CATEGORY FROM A SMALL SAMPLE.
#
# With three games in scope, EVERY text column has fewer than five distinct values — the team
# names included. Centring them all would be the rule firing on no evidence, and a one-game
# export would come out looking nothing like a fifty-game one. Below this many rows the
# measurement is not worth making and text stays left, which is the safe default.
MIN_ROWS_TO_JUDGE_CARDINALITY = 12


def _low_cardinality_text(df, fields) -> set:
    """Text columns in this frame holding fewer than LOW_CARDINALITY distinct values.

    Nulls do not count towards the total — a column of Yes/No/blank is still two categories,
    and counting the blank would push a three-value column over the line for no reason.
    """
    out = set()
    if len(df) < MIN_ROWS_TO_JUDGE_CARDINALITY:
        return out
    for field in fields:
        if field not in df.columns:
            continue
        values = df[field].dropna()
        if values.empty:
            continue
        if not all(isinstance(v, str) for v in values.head(50)):
            continue
        if values.nunique() < LOW_CARDINALITY:
            out.add(field)

    # AN away_/home_ PAIR IS ONE DECISION, NOT TWO.
    #
    # Measured on real data, `Home record` came out centred and `Away record` left, because
    # in one week the home side happened to hold four distinct records and the away side
    # five. Two identical columns aligned differently reads as a defect, and it is a coin
    # flip that would land the other way next week. So a pair agrees, and it agrees on PROSE
    # — the safe default — whenever either half looks like prose.
    for field in list(out):
        for this, that in (("away_", "home_"), ("home_", "away_")):
            if field.startswith(this):
                sibling = that + field[len(this):]
                if sibling in fields and sibling not in out:
                    out.discard(field)
    return out


def _marked(marks):
    """Render one of the site's marks, or a dash where cfdb holds nothing.

    NaN IS NOT FALSE AND IS NOT AN EMPTY STRING. This project has been bitten three times by
    `if value:` on a pandas null — network_abbreviation, logo_url, upset_level — so the null
    check here is explicit rather than truthiness.
    """
    def render(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return NO_DATA_MARK
        return marks.get(str(value), NO_DATA_MARK)
    return render


def _marked_or_blank(marks):
    """A mark, or an EMPTY CELL where the value is absent.

    For the `_open` columns only, and the distinction is the point. Everywhere else a null
    means "cfdb holds nothing here" and earns the no-data dash. In the opening-line block a
    null means "the market never changed its mind", which is a fact about the market rather
    than a gap in our data — and every numeric column in that block already says it by being
    blank. A dash here would make the mark column disagree with the eleven columns beside it.

    Named to keep the `_marked` prefix, because `Sheet.centred` recognises a mark column by
    that prefix on the renderer's qualname. Renaming this to something tidier would silently
    un-centre the column.
    """
    def render(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        return marks.get(str(value), NO_DATA_MARK)
    return render


# THE COVER VERDICT AS A WORD (R-262).
#
# `push` and `pending` are kept as themselves rather than folded into No. A push is the bet
# refunded and a pending game has not been graded — calling either of them "No" would be a
# claim about a result that does not exist yet, and the workbook gets filtered on this column.
COVER_WORDS = {"yes": "Yes", "no": "No", "push": "Push", "pending": "Pending"}


def _cover_word(value):
    """Yes / No / Push / Pending, or the no-data dash where no line was ever recorded."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return NO_DATA_MARK
    return COVER_WORDS.get(str(value), NO_DATA_MARK)


def _cover_word_or_blank(value):
    """As `_cover_word`, but BLANK where the value is absent.

    For the opening-line column only. There a null means the spread never moved, which is a
    fact about the market rather than a gap in our data, and the eleven columns beside it
    already say it by being empty.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return COVER_WORDS.get(str(value), NO_DATA_MARK)


# Postgres renders a boolean as 't'/'f' the moment anything touches it as text — a CSV
# export, a driver without type mapping, a copy through pandas' object dtype. AND 'f' IS A
# NON-EMPTY STRING, SO `bool('f')` IS TRUE. That is the same family as the NaN-truthiness
# bug this project has hit three times, and it turns every False into "Yes" silently.
# Caught by building a real workbook and reading row 1, not by any test written first.
FALSE_TEXT = {"f", "false", "n", "no", "0", ""}
TRUE_TEXT = {"t", "true", "y", "yes", "1"}


def _yes_no(value):
    """R-218. A boolean as a word, because Marc asked and because a filter dropdown reading
    Yes / No beats one reading TRUE / FALSE.

    THE COST, NOTED ON THE INDEX: a text "Yes" is not a boolean to a formula, so
    `=SUM(--(range="Yes"))` replaces `=COUNTIF(range,TRUE)`.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str):
        text = value.strip().lower()
        if text in FALSE_TEXT:
            return "No"
        if text in TRUE_TEXT:
            return "Yes"
        # Anything else is not a boolean at all, and guessing is how 'f' became "Yes".
        return None
    return "Yes" if bool(value) else "No"


def _title_case_verdict(value):
    """`yes` / `no` / `push` / `pending` / `no_favorite` as words a reader can filter on.

    Already a string, so this is title-casing rather than a boolean conversion — and
    `no_favorite` becomes "No favorite", NOT "No_Favorite", which is what a naive
    `.title()` would leave behind.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).replace("_", " ").strip()
    return text[:1].upper() + text[1:] if text else None


def _possession_minutes(record):
    """Time of possession in minutes, from the view's seconds (Marc, R-259).

    At 2dp a game's two rows SHOULD sum to 60.00, which mm:ss cannot give the reader — every
    game carries its own arithmetic check, and a row that does not pair is visible without
    leaving the sheet. The seconds column is not shipped alongside; two spellings of one
    measurement is a column of noise.

    AND THE CHECK FIRES, WHICH IS THE POINT OF HAVING IT. Across 3,411 games where both rows
    carry possession, 3,261 total exactly 60:00, 14 exceed it (overtime) and 136 fall short —
    4.4% where CFBD's own numbers do not reconcile. The column is not claiming they will
    always add up; it is making it visible when they do not.
    """
    seconds = record.get("possession_seconds")
    if seconds is None or (isinstance(seconds, float) and pd.isna(seconds)):
        return None
    return float(seconds) / 60.0


def _status(record):
    """Scheduled / Final. `is_completed` is a boolean the reader has to translate; a word is
    what they would have written in the cell themselves."""
    value = record.get("is_completed")
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return "Final" if bool(value) else "Scheduled"


def table_name(sheet_name: str) -> str:
    """`tbl_<SheetName>`, scrubbed to something Excel will accept."""
    cleaned = re.sub(r"[^A-Za-z0-9_]", "", sheet_name.title().replace(" ", ""))
    name = f"tbl_{cleaned}"
    if CELL_REFERENCE.match(name) or not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
        raise ValueError(f"{sheet_name!r} does not yield a legal Excel table name: {name!r}")
    return name


class Sheet:
    """One tab: where it comes from, what it is called, and which columns it shows.

    `columns` are (field, label) pairs. The label is what the site's table renders in that
    column's header — kept as literal text rather than imported from a page, because the
    page's own column carries an HTML renderer (a logo, a chip) that has no meaning in a
    spreadsheet. `tests/test_site_foundation.py` asserts the labels still agree.
    """

    def __init__(self, name: str, view: str, sql: str, columns: List[tuple],
                 has_predictions: bool = False, note: str = "", scoped: bool = True,
                 derived: Optional[Dict[str, Callable]] = None,
                 link_fields: Optional[Dict[str, str]] = None,
                 display: Optional[Dict[str, Callable]] = None,
                 sheet_disclaimer: Optional[bool] = None,
                 freeze_before: Optional[str] = None,
                 field_category: Optional[Dict[str, str]] = None,
                 integer_fields: frozenset = frozenset(),
                 site_precision: frozenset = frozenset(),
                 decimals: Optional[int] = None,
                 band_field: Optional[str] = None):
        self.name, self.view = name, view
        # Columns the sheet COMPUTES rather than selects — a weekday name, a status word, a
        # URL. They are real columns to the reader and to Excel; they simply have no
        # counterpart in the view, and putting them here keeps `fields` the single list that
        # both the writer and the width measurement walk.
        self.derived: Dict[str, Callable] = derived or {}
        # How a value is RENDERED, applied after it is fetched or derived. Separate from
        # `derived` on purpose: one answers "where does this value come from", the other
        # "what does the reader see". A boolean is still a boolean in the view.
        self.display: Dict[str, Callable] = display or {}
        # WHETHER THE SHEET WRITES THE MODEL DISCLAIMER, which is NOT the same question as
        # whether it carries predictions (R-221). Marc removed row 2 from Schedule, and the
        # sheet still has six prediction columns — what makes that safe is the per-row
        # `attribution` column, not the absence of predictions. Conflating the two flags
        # would have made the model lie about the sheet.
        self.sheet_disclaimer = has_predictions if sheet_disclaimer is None \
            else sheet_disclaimer
        # THE FIRST COLUMN THAT SCROLLS, by LABEL — named for what Excel actually does.
        #
        # `freeze_panes = "M4"` splits BEFORE M, so M is the first column to move and A–L
        # stay. Calling this `freeze_after` invited an off-by-one every time it was read, and
        # produced one: Marc's macro freezes at M, and "freeze after column M" would have put
        # it at N. Held by LABEL rather than letter so reordering the sheet moves it too.
        self.freeze_before = freeze_before
        # {column field -> what it links to}. Resolved at build time because the target
        # origin is an environment value, not a repo one.
        self.link_fields: Dict[str, str] = link_fields or {}
        # `{ROW_CAP}` rather than a literal, resolved once here. The contract's LIMIT check
        # (AC-G.39) matches `limit <digits>`, so a bind parameter would fail it — the cap has
        # to reach the SQL as a number. `ci/check_page_queries.py` substitutes the same hole
        # before executing, the way it already does for `{side}`.
        # {field -> category name}, which drives the header FILL (R-258). Empty on a sheet
        # with one category, and `build` falls back to the single navy in that case.
        self.field_category: Dict[str, str] = field_category or {}
        # Numeric columns that are naturally whole. Consulted BEFORE the sheet's decimal
        # default, so a count does not acquire ".00".
        self.integer_fields = integer_fields
        # Columns that keep `fmt.precision_for` even on a sheet with its own decimal default.
        self.site_precision = site_precision
        # This sheet's default decimal places for a real decimal, or None to keep deriving
        # them from `fmt.precision_for` as Schedule does.
        self.decimals = decimals
        # The column whose PARITY bands the rows. A value, never a row position — see
        # SCORES_BAND_FIELD for why the position-based version silently lies after a sort.
        self.band_field = band_field
        # Holes resolved once, here. The contract's LIMIT check (AC-G.39) matches
        # `limit <digits>`, so the cap has to reach the SQL as a number rather than a bind
        # parameter; the sort key is shared between the ORDER BY and `game_no`'s window and
        # is interpolated so the two cannot disagree. `ci/check_page_queries.py` resolves
        # both the same way before executing.
        for hole, value in SQL_HOLES.items():
            sql = sql.replace("{" + hole + "}", value)
        self.sql = sql
        self.columns, self.has_predictions = columns, has_predictions
        self.note, self.scoped = note, scoped

    @property
    def division_scoped(self) -> bool:
        """Whether this sheet's query can honour the Division filter.

        Only `srv_game` carries `is_fbs_game` — checked against the serving schema, not
        assumed. A sheet that cannot narrow says so on the Index rather than letting the
        scope line claim a filter it did not apply.
        """
        return ":division" in self.sql

    @property
    def fields(self) -> List[str]:
        return [field for field, _ in self.columns]

    @property
    def selected_fields(self) -> List[str]:
        """The bare column names this query actually selects.

        Not the same list as `fields`: the query also pulls what the sheet DERIVES from
        (`game_date`, `is_completed`), what it LINKS with (the two slugs) and the window
        count. A test fixture built from `fields` alone silently omits all of those, and the
        hyperlinks then look absent rather than unstubbed.
        """
        body = re.search(r"select\s+(.*?)\s+from\s", " ".join(self.sql.split()),
                         re.IGNORECASE | re.DOTALL)
        if not body:
            return []
        # SPLIT ON TOP-LEVEL COMMAS ONLY. A plain `.split(",")` walks straight into a window
        # function's own commas: Scores' `dense_rank() over (order by season, <case>, week,
        # game_date, game_id) as game_no` came apart into six pieces, four of which look like
        # bare column names. Two of them — `end` and `game_no` — carry no bracket at all and
        # were reported as selected columns that do not exist, which is precisely the class of
        # error this property exists to prevent a fixture from inheriting.
        out, depth, piece = [], 0, ""
        for character in body.group(1):
            if character == "(":
                depth += 1
            elif character == ")":
                depth -= 1
            if character == "," and depth == 0:
                out.append(piece)
                piece = ""
            else:
                piece += character
        out.append(piece)
        return [p.split()[-1] for p in (t.strip() for t in out) if p and "(" not in p]

    @property
    def centred(self) -> set:
        """Columns whose values are marks, and are therefore centred.

        Derived from `display` rather than listed again: a column that renders a shape is
        exactly a column that should be centred, and two lists would drift.
        """
        return {field for field, renderer in self.display.items()
                if getattr(renderer, "__qualname__", "").startswith("_marked")}

    @property
    def open_mark_fields(self) -> set:
        return self.centred

    @property
    def url_value_fields(self) -> set:
        """Columns built FROM a url, whether or not they show it.

        The original argument for this column was that a cell hyperlink is invisible to
        anything reading the file as data, so one column should show the URL as text.
        MARC OVERRODE THAT after seeing it: a 90-character URL in every row of a 56-column
        sheet costs more width and legibility than the machine-readability is worth to him,
        and the `Game id` column beside it already reconstructs the link.

        So the cell reads `URL_CELL_LABEL` and carries the link. The distinction the property
        name draws — value comes from a url builder, not from the frame — still holds.
        """
        return {f for f in self.link_fields if f.endswith("_url")}

    def value_for(self, field: str, record):
        """One cell's value: computed if this sheet derives it, selected otherwise, then
        rendered if this sheet has a display rule for it."""
        builder = self.derived.get(field)
        value = builder(record) if builder is not None else record.get(field)
        renderer = self.display.get(field)
        return renderer(value) if renderer is not None else value

    def header_fill_for(self, field: str, default: str) -> str:
        """This column's header fill: its category's, or the sheet-wide default."""
        return CATEGORY_FILLS.get(self.field_category.get(field, ""), default)

    def freeze_column(self) -> int:
        """1-based index of the first column that SCROLLS, or 1 for no horizontal freeze."""
        if not self.freeze_before:
            return 1
        labels = [label for _, label in self.columns]
        return labels.index(self.freeze_before) + 1

    def hyperlinks(self, base: str, **scope) -> Dict[str, Callable]:
        """{field -> record -> url}, or {} when nothing told us the site's origin.

        THE SLUG FIELD IS STATED, NOT DERIVED (R-289). This used to read

            slug_field = field.replace("_team_display", "_team_slug")

        which is correct on Schedule and a no-op everywhere else. On Scores the display
        column is `team`, so the replace changed nothing and the TEAM NAME was passed where a
        slug belongs — producing a hyperlink that is written, looks right, and goes nowhere.

        It is the same shape as the defect on the page (R-287): an identifier inferred from a
        naming convention rather than named. A convention that happens to hold on one sheet
        is not a rule, so both sheets now say which column carries the slug.
        """
        out: Dict[str, Callable] = {}
        for field, spec in self.link_fields.items():
            kind, source = spec if isinstance(spec, tuple) else (spec, None)
            if kind == "team":
                if not source:
                    raise ValueError(
                        f"{self.name}.{field} links to a team without naming its slug "
                        f"column; deriving it is what shipped a link to nowhere")
                out[field] = (lambda record, sf=source:
                              _team_url(base, record.get(sf), **scope))
            elif kind == "matchup":
                out[field] = (lambda record:
                              _matchup_url(base, record.get("game_id"), **scope))
        return out


# ==========================================================================================
# THE SCORES SHEET — srv_game_team, 131 FIELDS, GAME x TEAM GRAIN (R-255 … R-259)
#
# MARC: "We are going to veer Scores away from Schedule. Source Scores from srv_game_team."
# The order is `claude_work/supporting_files/cfdb_scores_column_order.csv`, taken literally.
#
# THE GRAIN IS THE THING TO HOLD ON TO: one row per team per game, so A GAME IS TWO ROWS.
# The sort, the banding, the possession check and the row count all follow from that one
# fact, and every one of them breaks if a filter ever splits a pair.
# ==========================================================================================

# THE SEVEN CATEGORY BANDS. THE ORDER IS MARC'S FILE, AND THE FILE IS THE TEST FIXTURE.
#
# `tests/fixtures/cfdb_scores_column_order_v2.csv` — 144 rows of Position/Field/Category, from
# the "Updated format request." sheet of his attachment. It SUPERSEDES the 131-field v1 list,
# which predates the Market block. The structure below is generated from that file and the
# test asserts the two match on order AND category, so the file is authoritative rather than
# merely consulted.
#
# R-264 IS A PERMUTATION, NOT AN EDIT. 144 in, 144 out: seven columns lift out of the front
# block and go to the far right, and every other column keeps its relative position. Verified
# before building, and asserted after.
#
#     team_id · conference · classification · opponent_team_id · opponent_conference
#     is_neutral_site · is_completed
#
# THEY ARE THEIR OWN CATEGORY BECAUSE THE FIRST ANSWER BROKE AN INVARIANT. Marc's initial
# sketch painted all seven the same navy as the Game block — which is one category in two
# separate runs, at columns 1-13 and again at 138-144. `SCORES_BLOCKS` is a structure of
# CONTIGUOUS runs precisely so that a split category is impossible rather than merely tested,
# and two tests forbid it. Shown the conflict, Marc made them their own band (2026-09-05):
# "Can change the block at the end to ancillary data points and use a different color."
#
# The earlier shuffle (R-258) is still in force underneath this one: opponent_conference had
# been stranded at field 117 between the advanced blocks and both havoc runs were cut off from
# their own side. Those three moves are what made each category one run in the first place;
# this one keeps that true while moving the keys out of the reader's way.
SCORES_BLOCKS = (
    ("Game", (
        "game_no", "game_id", "season", "season_type", "week", "game_date", "team",
        "opponent", "is_home", "points_for", "points_against", "margin", "result",
        # R-290. Outcome context, beside the team it describes. Marc chose this over
        # appending to Ancillary: they read as part of the result, and the Game Results tab
        # picks them up with no extra mapping. It shifts everything from `spread_final`
        # rightward, which is the cost he accepted.
        "team_rank", "record_before_display", "pregame_elo", "postgame_elo",
    )),
    ("Market", (
        "spread_final", "total_final", "line_implied_points_final",
        "points_vs_line_implied_final", "ats_margin_final", "covered_final", "spread_open",
        "total_open", "line_implied_points_open", "points_vs_line_implied_open",
        "ats_margin_open", "covered_open",
    )),
    ("Box score", (
        "first_downs", "total_yards", "rushing_yards", "passing_yards", "rushing_attempts",
        "turnovers", "interceptions", "fumbles_lost", "third_down_conversions",
        "third_down_attempts", "fourth_down_conversions", "fourth_down_attempts", "penalties",
        "penalty_yards", "possession_minutes",
    )),
    ("Team advanced", (
        "plays", "ppa_overall_total", "ppa_passing_total", "ppa_rushing_total",
        "cumulative_ppa_overall_total", "cumulative_ppa_passing_total",
        "cumulative_ppa_rushing_total", "success_rate_overall_total",
        "success_rate_standard_downs_total", "success_rate_passing_downs_total",
        "explosiveness_total", "power_success", "stuff_rate", "line_yards",
        "line_yards_average", "second_level_yards", "second_level_yards_average",
        "open_field_yards", "open_field_yards_average", "havoc_total", "havoc_front_seven",
        "havoc_db", "scoring_opportunities", "scoring_opportunity_points",
        "points_per_opportunity", "average_start", "average_starting_predicted_points",
    )),
    ("Offense", (
        "offense_plays", "offense_drives", "offense_ppa", "offense_total_ppa",
        "offense_success_rate", "offense_explosiveness", "offense_power_success",
        "offense_stuff_rate", "offense_line_yards", "offense_line_yards_total",
        "offense_second_level_yards", "offense_second_level_yards_total",
        "offense_open_field_yards", "offense_open_field_yards_total",
        "offense_standard_downs_ppa", "offense_standard_downs_success_rate",
        "offense_standard_downs_explosiveness", "offense_passing_downs_ppa",
        "offense_passing_downs_success_rate", "offense_passing_downs_explosiveness",
        "offense_rushing_plays_ppa", "offense_rushing_plays_total_ppa",
        "offense_rushing_plays_success_rate", "offense_rushing_plays_explosiveness",
        "offense_passing_plays_ppa", "offense_passing_plays_total_ppa",
        "offense_passing_plays_success_rate", "offense_passing_plays_explosiveness",
        "offense_total_plays", "offense_total_havoc_events",
        "offense_front_seven_havoc_events", "offense_db_havoc_events", "offense_havoc_rate",
        "offense_front_seven_havoc_rate", "offense_db_havoc_rate",
    )),
    ("Defense", (
        "defense_plays", "defense_drives", "defense_ppa", "defense_total_ppa",
        "defense_success_rate", "defense_explosiveness", "defense_power_success",
        "defense_stuff_rate", "defense_line_yards", "defense_line_yards_total",
        "defense_second_level_yards", "defense_second_level_yards_total",
        "defense_open_field_yards", "defense_open_field_yards_total",
        "defense_standard_downs_ppa", "defense_standard_downs_success_rate",
        "defense_standard_downs_explosiveness", "defense_passing_downs_ppa",
        "defense_passing_downs_success_rate", "defense_passing_downs_explosiveness",
        "defense_rushing_plays_ppa", "defense_rushing_plays_total_ppa",
        "defense_rushing_plays_success_rate", "defense_rushing_plays_explosiveness",
        "defense_passing_plays_ppa", "defense_passing_plays_total_ppa",
        "defense_passing_plays_success_rate", "defense_passing_plays_explosiveness",
        "defense_total_plays", "defense_total_havoc_events",
        "defense_front_seven_havoc_events", "defense_db_havoc_events", "defense_havoc_rate",
        "defense_front_seven_havoc_rate", "defense_db_havoc_rate",
    )),
    ("Ancillary", (
        "team_id", "conference", "classification", "opponent_team_id", "opponent_conference",
        "is_neutral_site", "is_completed",
        # R-289. EXPORT-ONLY, and the far right is where a reference key belongs. A
        # `cell.hyperlink` is invisible to a formula, to pandas and to anything reading the
        # file as data — which is half of what an extract is for — so the URL is a value too.
        # Never rendered on a page tab; the partition test accounts for it explicitly.
        "matchup_url",
    )),
)

# WHITE BOLD SITS ON ALL FIVE, so each needs 4.5:1 against white — and the bands have to be
# told apart from EACH OTHER, which contrast against white says nothing about. Measured, not
# eyeballed, the way the marks were:
#
#     band              hex        contrast on white
#     Game              2F4858       9.59:1     the existing cfdb navy, kept
#     Ancillary         1C4080      10.03:1     the highest on the sheet
#     Market            5C6B2F       5.83:1
#     Box score         6B4C7A       7.14:1
#     Team advanced     4A6670       6.13:1
#     Offense           A63A16       6.49:1
#     Defense           1E6B54       6.39:1
#
# Across all TWENTY-ONE pairs, under normal vision and both red-green dichromacies, the closest
# two colours are 10.5 dE apart (Market / Offense under protanopia) against a just-noticeable
# difference of about 2.3.
#
# THAT WAS ALSO THE CLOSEST PAIR WITH SIX BANDS, which is the number that mattered when the
# seventh was chosen: `1C4080` does not tighten the palette anywhere. Its own nearest
# neighbour is 18.1 dE (Box score, protanopia), and it sits 31.2 and 37.7 dE from the other
# two blues — it lands in a real hue gap between the navy at 249 degrees and the plum at 317.
#
# AND THE OBVIOUS CHOICE FOR "ANCILLARY" IS THE ONE THAT CANNOT WORK. A quiet neutral grey is
# semantically exactly right and fails every time, measured: graphite 3F4448 at 8.9 dE, mid
# grey 4A4A4A at 8.7, warm graphite 55504B at 4.5. The palette already carries two low-chroma
# bands (Game and Team advanced), and under dichromacy a grey collapses towards them. This is
# the second time that finding has come up — the sixth band saw graphite land between 7 and 9
# — so it is a confirmation rather than a discovery. Ancillary earns its quietness from
# POSITION and NAME, not from desaturation.
#
# THE SIXTH BAND WAS THE HARD ONE AND EYEBALLING WOULD HAVE GOT IT WRONG. A dark gold
# (7A5C00) reads as obviously distinct from the sienna next to it and measures 1.5 dE from it
# under deuteranopia — indistinguishable. Of five candidates only the olive cleared the
# threshold; crimson, graphite and bronze all landed between 7 and 9. The palette is spread over hue AND lightness on
# purpose: hue alone collapses under deuteranopia, which is what makes red-and-green the
# wrong choice for two blocks that sit side by side.
#
# FIVE FILLS RATHER THAN FOUR, and that is a change of mind worth recording. The first pass
# gave Box score and Team advanced one colour on the argument that the split between them is
# which CFBD endpoint served them rather than anything about football. True, but it produced
# a 42-column run with no boundary in it — the widest band on the sheet, and the one place a
# reader most needs a landmark. A visible boundary at every block beats a tidy taxonomy.
#
# THIS SUPERSEDES R-182 TRAP 3's single-fill header, which was settled before category
# colours were asked for. `showRowStripes` stays False — see SCORES_BAND_FIELD.
CATEGORY_FILLS = {
    "Game": "2F4858",
    "Market": "5C6B2F",
    "Box score": "6B4C7A",
    "Team advanced": "4A6670",
    "Offense": "A63A16",
    "Defense": "1E6B54",
    "Ancillary": "1C4080",
}

# Labels are RULE-DERIVED with overrides, not 131 hand-typed strings. 131 chances to typo is
# not a trade worth making when the names follow a pattern; the exceptions are listed and the
# test asserts every label is unique and non-empty.
SCORES_WORDS = {
    "offense": "Off", "defense": "Def", "ppa": "PPA", "db": "DB",
    "first": "1st", "second": "2nd", "third": "3rd", "fourth": "4th",
    "opponent": "Opp", "average": "Avg",
}
SCORES_LABEL_OVERRIDES = {
    "game_no": "Game #", "game_id": "Game id", "season": "Season",
    "season_type": "Season type", "week": "Wk", "game_date": "Date",
    "team_id": "Team id", "team": "Team", "conference": "Conference",
    "classification": "Division", "opponent_team_id": "Opp team id",
    "opponent": "Opponent", "opponent_conference": "Opp conference",
    "is_home": "Home game", "is_neutral_site": "Neutral site",
    "is_completed": "Completed", "points_for": "Pts for",
    "points_against": "Pts against", "margin": "Margin", "result": "Result",
    "possession_minutes": "Possession min",
    "average_start": "Avg start", "average_starting_predicted_points": "Avg start pred pts",
    "scoring_opportunities": "Scoring opps",
    "scoring_opportunity_points": "Scoring opp pts",
    "points_per_opportunity": "Pts per opp",
    "havoc_front_seven": "Havoc front 7", "havoc_db": "Havoc DB",
    # THE MARKET (R-260). "Line-implied", never "expected" or "projected" — R-201 section 2.1
    # settled that as a LICENCE point rather than a style one: these come from the market,
    # not from cfdb's model, and a header reading "Expected pts" presents a bookmaker's
    # number as a prediction of ours. Marc's phrasing was "points expected"; the wording here
    # is the ruling he already has on the books, not a second-guess of the ask.
    "spread_final": "Spread", "total_final": "O/U",
    "line_implied_points_final": "Line-implied pts",
    "points_vs_line_implied_final": "Pts vs implied",
    "ats_margin_final": "ATS margin", "covered_final": "Covered",
    "spread_open": "Spread open", "total_open": "O/U open",
    "line_implied_points_open": "Line-implied pts open",
    "points_vs_line_implied_open": "Pts vs implied open",
    "ats_margin_open": "ATS margin open", "covered_open": "Covered open",
    # R-284/R-285/R-286. `Pregame Elo`, never `Elo before`: the vocabulary rule says grep for
    # an existing prefix before inventing one, and stg_games, fct_game and srv_game all
    # already say pregame/postgame. `Record before` keeps Marc's own word, and keeps the
    # column honest about whether this game is counted in it.
    "team_rank": "Rank", "record_before_display": "Record before",
    "pregame_elo": "Pregame Elo", "postgame_elo": "Postgame Elo",
    "matchup_url": "Matchup URL",
}

# SELECTED BUT NOT PRINTED. The page needs a slug to link a team and a logo to draw one; the
# sheet needs neither, and a 149-column file does not want four more. They ride in the frame
# and out of the column list — a different thing from R-279's HIDDEN_ON_PAGE, where the
# column IS printed in the file and merely not shown on the site.
SCORES_PASSENGERS = ("team_slug", "opponent_team_slug",
                     "team_logo_url", "opponent_logo_url", "opponent_rank")

# COLUMNS THAT KEEP THE SITE'S PRECISION RATHER THAN THIS SHEET'S 2dp DEFAULT.
#
# R-259 refused to flip the 2dp rule globally because a spread is written `-6.5` everywhere in
# football and never `-6.50`. That argument does not stop at the sheet boundary: these six are
# the same market numbers, quoted in halves, and `fmt.precision_for` already gives them 1dp.
# Exempting them here is the rule being applied consistently, not an exception to it.
#
# The line-implied columns are NOT exempt, and the difference is real: a spread of -6.5 with a
# total of 51.5 implies 29.0 and 22.5, but -6.5 with 52.0 implies 29.25 and 22.75. Quarter
# points are a consequence of halving, so those need the second digit.
SCORES_SITE_PRECISION = frozenset({
    "spread_final", "spread_open", "total_final", "total_open",
    "ats_margin_final", "ats_margin_open",
})


def scores_label(field: str) -> str:
    """A readable header for one srv_game_team column."""
    if field in SCORES_LABEL_OVERRIDES:
        return SCORES_LABEL_OVERRIDES[field]
    # `front_seven` is ONE idea and has to be substituted as one. Word-by-word turns
    # `offense_front_seven_havoc_rate` into "Off front front 7 havoc rate" — the `front`
    # survives and the `seven` becomes `7` beside it.
    name = field.replace("front_seven", "front7")
    words = [SCORES_WORDS.get(w, w) for w in name.split("_")]
    text = " ".join(words).replace("front7", "front 7")
    return text[0].upper() + text[1:]


def _flatten_blocks(blocks):
    """(columns, {field -> category}) from the block structure above."""
    columns, category = [], {}
    for name, fields in blocks:
        for field in fields:
            columns.append((field, scores_label(field)))
            category[field] = name
    return columns, category


SCORES_COLUMNS, SCORES_CATEGORY = _flatten_blocks(SCORES_BLOCKS)

# `regular` BEFORE `postseason`, WHICH A PLAIN SORT GETS BACKWARDS.
#
# Marc named the order explicitly — "Season (regular then post season)" — and alphabetically
# 'postseason' < 'regular', so `order by season_type` puts January's bowls ahead of
# September. The CASE is load-bearing, not decoration, and the test removes it to prove that.
#
# WORTH KNOWING: inside ONE export it never fires, because `season_type` is a scope filter
# and every row in the file shares a value. It is right anyway, and it is what makes the
# ordering correct the day the filter widens or a second season_type reaches one sheet.
#
# THE GAME ORDER, shared by the sort and by `game_no` so the numbering and the row order
# cannot disagree. `is_home` is appended only to the ORDER BY: false sorts first in Postgres,
# which is away-then-home, which is how a scoreboard is written.
#
# ONE PLAIN LITERAL, not an f-string assembled from parts. `ci/check_page_queries.py` resolves
# these holes by reading module-level `NAME = "..."` constants out of the AST, and it reads
# `ast.Constant` only — an f-string is a `JoinedStr` and would be invisible to it, so the
# Scores query would reach the checker with an unresolved `{SCORES_GAME_ORDER}` and be
# reported as broken rather than executed.
SCORES_GAME_ORDER = ("season, case season_type when 'regular' then 1 "
                     "when 'postseason' then 2 else 3 end, week, game_date, game_id")

# BAND ON A REAL COLUMN, NOT ON ROW POSITION (R-257).
#
# Marc asked for "some kind of subgrouping to show the change between games". Two things that
# look obvious are both wrong here:
#
#   Excel's own row stripes alternate every ROW, so with two rows per game they would shade
#   the away row of every game and leave the home row bare — banding the wrong unit entirely.
#   R-182 already set showRowStripes=False for the navy header; it now does double duty.
#
#   A rule like `=$A2<>$A1` is computed from row POSITION, so the moment a reader re-sorts
#   the Table — which is the entire point of shipping a Table — the lines land between the
#   wrong rows. A format that silently lies after a sort is worse than no format.
#
# `game_no` is DATA. It travels with its row, so the band survives any sort, and it is useful
# on its own: filter to one game, or read off how many games are in scope. Not hidden — a
# hidden column driving visible formatting is the thing nobody can debug six months later.
SCORES_BAND_FIELD = "game_no"

# PAINTED ON THE CELLS, NOT BY A CONDITIONAL FORMAT — AND THAT IS THE THIRD ATTEMPT.
#
# Round 1 shipped a conditional-format rule at F2F5F7. Marc: the separation is "missing
# between every other game". Diagnosed as too faint (3.9 dE from white against a
# just-noticeable difference of 2.3) and re-shipped at DEE2E4, 10.5 dE. Marc, again: still
# not discernible.
#
# WHAT THE SECOND REPORT RULES OUT. A colour that measures 10.5 dE is not invisible. Two
# rounds of the same complaint at two very different lightnesses says the fill is not being
# DRAWN, not that it is hard to see — and the corroborating detail is Marc's own: the header
# colours "look excellent". Those are direct cell fills. The band was the only thing on the
# sheet painted by a conditional format.
#
# I COULD NOT CONFIRM THE MECHANISM AND STOPPED TRYING. Excel's AppleScript bridge returns
# `missing value` for the interior colour of a format condition AND for the header fills that
# demonstrably render, so it cannot distinguish the two; `do Visual Basic` was removed from
# Excel for Mac, so DisplayFormat is unreachable. The likely culprit is that a dxf's solid
# fill takes its colour from `bgColor` while openpyxl writes `fgColor`, but that is a
# hypothesis I have no way to test from here, and Marc has now reported this twice.
#
# So the band stops depending on anything being evaluated at open time. A direct fill is a
# real entry in cellXfs; it renders or the file is corrupt, and it can be read straight back
# out of the saved workbook by the test below.
#
# THE COST, STATED PLAINLY: a painted fill does NOT follow a row when the reader re-sorts the
# Table, which is exactly what R-257 wanted the rule for. The band now describes the order the
# file was written in. `Game #` is the column that survives a sort, which is why R-257 put it
# in the sheet as data rather than hiding it — the durable answer was always the column, and
# the shading was always only a convenience for the default view.
SCORES_BAND_FILL = "DEE2E4"

# NATURALLY INTEGER, MEASURED OVER ALL 110,879 ROWS rather than inferred from the column name.
#
# Marc: "If the number is numeric and not naturally an integer, only show 2 decimal points."
# Deciding that from names would have been wrong four times, and the count is the point:
#
#   `line_yards`, `second_level_yards`, `open_field_yards` are numeric-typed and WHOLE in
#   every row — they are totals. Their `_average` siblings are not.
#   `havoc_total` READS like a count and is fractional in 3,685 rows — it is a rate.
#   `offense_total_havoc_events` is whole in all but FIFTEEN rows, so it is not an integer
#   either; formatting it as one would round those fifteen away in silence.
#   `offense_db_havoc_events` IS whole everywhere, while its front-seven sibling is not.
#
# R-216's rule still decides the SEPARATOR: a comma is for a quantity you might total, so a
# season and an id stay bare. This set only decides the DECIMAL POINT.
SCORES_INTEGER_FIELDS = frozenset({
    "game_no", "points_for", "points_against", "margin",
    "first_downs", "total_yards", "rushing_yards", "passing_yards", "rushing_attempts",
    "turnovers", "interceptions", "fumbles_lost",
    "third_down_conversions", "third_down_attempts",
    "fourth_down_conversions", "fourth_down_attempts", "penalties", "penalty_yards",
    "plays", "line_yards", "second_level_yards", "open_field_yards",
    "scoring_opportunities", "scoring_opportunity_points",
    "offense_plays", "offense_drives", "offense_total_plays", "offense_db_havoc_events",
    "offense_line_yards_total", "offense_second_level_yards_total",
    "offense_open_field_yards_total",
    "defense_plays", "defense_drives", "defense_total_plays", "defense_db_havoc_events",
    "defense_line_yards_total", "defense_second_level_yards_total",
    "defense_open_field_yards_total",
})

# TWO DECIMALS ON THIS SHEET, AND DELIBERATELY NOT GLOBALLY (R-259).
#
# Schedule derives precision from `fmt.precision_for` so a column reads the same in the
# workbook as on the site (AC-G.31) — and a spread is written `-6.5` everywhere in football,
# never `-6.50`. Two sheets with two stated defaults is not drift; one rule applied where it
# is wrong is.
#
# ⚠ REPORTED TO MARC, NOT DECIDED HERE: the rates are stored as PROPORTIONS (0-1), so 2dp
# renders 0.2549 as 0.25. Measured on week 2 of 2025, 133 FBS team-rows: offense_havoc_rate
# holds 91 distinct values at 3dp and 25 at 2dp; offense_success_rate, 111 and 45. Marc's
# call — this constant is the whole change.
SCORES_DECIMALS = 2


# The f-string holes any sheet's SQL may carry, resolved once in `Sheet.__init__`.
# `ci/check_page_queries.py` resolves the same two before executing — ROW_CAP from its own
# SUBSTITUTIONS table, SCORES_GAME_ORDER by reading this module's constants — so the checker
# runs the query the workbook runs rather than a placeholder-shaped approximation of it.
SQL_HOLES = {"ROW_CAP": str(ROW_CAP), "SCORES_GAME_ORDER": SCORES_GAME_ORDER}


# SCHEDULE AND SCORES SHIP; FIVE SHEETS ARE STILL PENDING. Marc took the one-sheet
# decision of 2026-09-03 ("only 1 data sheet for now") and reopened it for Scores
# specifically, on a new source — R-255. The remaining five are unchanged.
#
# The other six are NOT deleted. Their SQL and column lists are real work and each gets the
# same treatment one at a time; they sit in PENDING_SHEETS, are named on the Index so the
# absence is an answer rather than a gap, and their queries stay under the CI query checker.
#
# Why not keep them shipping in the OLD layout: a workbook with one sheet whose header row is
# computed and six whose header row is 4 is a file that contradicts itself about where its
# data starts, which is exactly the thing R-181 exists to fix.
_ALL_SHEETS = [
    # ======================================================================================
    # THE SCHEDULE SHEET — FIFTY-SIX COLUMNS IN MARC'S ORDER (R-214, R-215).
    #
    # THE ORDER IS `claude_work/supporting_files/cfdb_schedule_column_order.csv`, TAKEN
    # LITERALLY. It supersedes spec §3.2's block ordering. Reconciled before building: 67
    # columns built, 11 removed, 56 remain, and his file has exactly 56 rows with no extras
    # and nothing missing.
    #
    # The shape is worth naming so nobody "improves" it later:
    #     scope → fixture → market → result → context → model → keys
    #
    # ELEVEN COLUMNS CAME OUT (R-214), and two consequences are recorded on the Index rather
    # than left for a reader to discover:
    #
    #   `Spread open` and `O/U open` went, so `Δ Spread` LOSES ITS DISAMBIGUATOR. The delta
    #   is null both when the line did not move and when no opening number was recorded —
    #   two facts, one blank — and the neighbouring column was what told them apart. No
    #   sentinel is invented; the Index says the blank means either.
    #
    #   All six CLOSING-LINE columns went, so `Upset level`, `Winner covered` and
    #   `O/U result` lose the number they were judged against. The verdicts are still
    #   correct; they are no longer auditable from the workbook alone. The Matchup URL is
    #   where a reader goes to see the line.
    # ======================================================================================
    Sheet("Schedule", "srv_game", """
        select season, season_type, week, is_current_week, start_date_et,
               best_rank_in_game, is_completed,
               away_rank, away_team_display, away_team_record_display,
               home_rank, home_team_display, home_team_record_display,
               provider_key, line_snapshot_ts,
               spread_current, spread_move_from_open,
               total_current, total_move_from_open,
               away_points, home_points, winner, actual_margin, final_margin, total_points,
               upset_level, is_upset_by_line, winner_covered_close, favorite_covered,
               over_met,
               temperature_f, weather_condition, wind_speed_mph, precipitation_in,
               excitement_index, attendance,
               is_neutral_site, venue_display, is_indoors, network_abbreviation,
               away_conference, home_conference, away_classification, home_classification,
               is_conference_game, is_fbs_game,
               predicted_margin, home_win_probability, confidence_bucket,
               model_name, model_version_key, is_out_of_sample_week,
               game_id, as_of_ts, attribution,
               home_team_slug, away_team_slug,
               count(*) over () as rows_in_scope
        from srv_game
        where season = :season and season_type = :season_type
          and (:week is null or week = :week)
          /* R-184. The Division filter, which this query used to ignore entirely.
             Predicate copied verbatim from `views/schedule.py` rather than re-derived: the
             export must return the set the page showed, and two spellings of one rule is
             how they drift. EITHER team FBS, not both.
             BLOCK comment, not `--`: read_sheet() flattens this SQL onto ONE line, so a
             line comment would swallow every clause after it. (And the wording here avoids
             the word j-o-i-n, because the query contract's FORBIDDEN pattern reads comments
             too — it caught this comment's first draft.) */
          and (:division = 'all' or is_fbs_game)
          and (:conference is null or home_conference = :conference
               or away_conference = :conference)
        order by start_date_et, game_id
        limit {ROW_CAP}
    """, [
        # --- scope ----------------------------------------------------------------------
        ("season", "Season"),                       # "0": 2025, never "2,025" (R-216)
        ("season_type", "Season type"),
        ("week", "Wk"),
        ("is_current_week", "Current week"),
        # --- the fixture. Freeze lands after `Home record`, which is column M. ------------
        ("start_date_et", "Kickoff"),
        ("best_rank_in_game", "Best rank"),
        ("status", "Status"),
        ("away_rank", "Away rank"),
        ("away_team_display", "Away"),
        ("away_team_record_display", "Away record"),
        ("home_rank", "Home rank"),
        ("home_team_display", "Home"),
        ("home_team_record_display", "Home record"),
        # --- the market -------------------------------------------------------------------
        ("provider_key", "Book"),
        ("line_snapshot_ts", "Line taken"),
        ("spread_current", "Spread"),
        ("spread_move_from_open", "Δ Spread"),
        ("total_current", "O/U"),
        ("total_move_from_open", "Δ O/U"),
        # --- the result -------------------------------------------------------------------
        ("away_points", "Away pts"),
        ("home_points", "Home pts"),
        ("winner", "Winner"),
        ("actual_margin", "Margin (away−home)"),
        ("final_margin", "Final margin"),
        ("total_points", "Total points"),
        ("upset_level", "Upset level"),
        ("is_upset_by_line", "Upset by line"),
        ("winner_covered_close", "Winner covered"),
        ("favorite_covered", "Favorite covered"),
        ("over_met", "O/U result"),
        # --- context ----------------------------------------------------------------------
        ("temperature_f", "Temperature °F"),
        ("weather_condition", "Condition"),
        ("wind_speed_mph", "Wind mph"),
        ("precipitation_in", "Precipitation in"),
        ("excitement_index", "Excitement"),
        ("attendance", "Attendance"),
        ("is_neutral_site", "Neutral site"),
        ("venue_display", "Venue"),
        ("is_indoors", "Indoors"),
        ("network_abbreviation", "TV"),
        ("away_conference", "Away conference"),
        ("home_conference", "Home conference"),
        ("away_classification", "Away division"),
        ("home_classification", "Home division"),
        ("is_conference_game", "Conference game"),
        ("is_fbs_game", "FBS game"),
        # --- the model --------------------------------------------------------------------
        ("predicted_margin", "Pred margin"),
        ("home_win_probability", "Home win prob"),
        ("confidence_bucket", "Confidence"),
        ("model_name", "Model"),
        ("model_version_key", "Model version"),
        # AC-15.4: per ROW, never a footnote. A workbook gets filtered and sorted, and a
        # caption does not survive either.
        ("is_out_of_sample_week", "Out-of-sample week"),
        # --- keys and provenance ------------------------------------------------------------
        ("game_id", "Game id"),
        ("matchup_url", "Matchup URL"),
        ("as_of_ts", "As of"),
        # R-221. The sheet-level disclaimer is gone; THIS column is what replaces it, and it
        # is strictly stronger — attribution carried per row survives filtering, sorting and
        # copy-paste, which a line in row 2 does not.
        ("attribution", "Attribution"),
    ],
        has_predictions=True,          # it does carry predictions...
        sheet_disclaimer=False,        # ...and says so per row instead of in row 2 (R-221)
        derived={"status": _status},
        display={
            # R-218. Booleans as words. `Out-of-sample week` is deliberately NOT here:
            # Marc's CSV marks six fields t/f = Yes/No and leaves that one blank, and the
            # CSV is authoritative. Flagged in the report as the one boolean left raw.
            "is_current_week": _yes_no,
            "is_upset_by_line": _yes_no,
            "is_neutral_site": _yes_no,
            "is_indoors": _yes_no,
            "is_conference_game": _yes_no,
            "is_fbs_game": _yes_no,
            "favorite_covered": _title_case_verdict,
            # R-219. The site's marks.
            "upset_level": _marked(UPSET_MARKS),
            "winner_covered_close": _marked(COVER_MARKS),
            "over_met": _marked(OVER_MARKS),
        },
        # Marc's macro freezes at M2. The COLUMN is his and is kept literally: M is the
        # first column to scroll, so Season through Home stay on screen while the market and
        # the result move. The ROW is not his — the header is not row 1 — so it is computed.
        # (Worth knowing: this splits the two record columns. `Home record` scrolls while
        # `Away record` stays. Moving the freeze one column right is a one-word change.)
        freeze_before="Home record",
        link_fields={"away_team_display": ("team", "away_team_slug"),
                     "home_team_display": ("team", "home_team_slug"),
                     "matchup_url": "matchup"}),

    # ======================================================================================
    # THE SCORES SHEET — VEERED AWAY FROM SCHEDULE (Marc, R-255).
    #
    # It used to be twelve columns of srv_game: the same game-grain fixture Schedule already
    # covers, with fewer columns. That made it a worse Schedule rather than a different
    # sheet, which is presumably why Marc moved it: srv_game_team sits at game x TEAM, so
    # Scores now answers "how did each team play" where Schedule answers "what happened in
    # this fixture". Two grains, two sheets, no overlap to keep in step.
    #
    # THE SELECT LIST IS MARC'S CSV ORDER, VERBATIM. The DISPLAY order is the same 131 fields
    # regrouped into contiguous category bands (SCORES_BLOCKS). Keeping the select in his
    # order means the two artefacts can be read against each other line by line; the test
    # asserts they hold the same set.
    # ======================================================================================
    Sheet(
        "Scores", "srv_game_team", """
        select game_id, season, season_type, week, game_date, team_id, team, conference,
               classification, opponent_team_id, opponent, is_home, is_neutral_site,
               is_completed, points_for, points_against, margin, result, first_downs,
               total_yards, rushing_yards, passing_yards, rushing_attempts, turnovers,
               interceptions, fumbles_lost, third_down_conversions, third_down_attempts,
               fourth_down_conversions, fourth_down_attempts, penalties, penalty_yards,
               possession_seconds, plays, ppa_overall_total, ppa_passing_total,
               ppa_rushing_total, cumulative_ppa_overall_total,
               cumulative_ppa_passing_total, cumulative_ppa_rushing_total,
               success_rate_overall_total, success_rate_standard_downs_total,
               success_rate_passing_downs_total, explosiveness_total, power_success,
               stuff_rate, line_yards, line_yards_average, second_level_yards,
               second_level_yards_average, open_field_yards, open_field_yards_average,
               havoc_total, havoc_front_seven, havoc_db, scoring_opportunities,
               scoring_opportunity_points, points_per_opportunity, average_start,
               average_starting_predicted_points, offense_plays, offense_drives,
               offense_ppa, offense_total_ppa, offense_success_rate, offense_explosiveness,
               offense_power_success, offense_stuff_rate, offense_line_yards,
               offense_line_yards_total, offense_second_level_yards,
               offense_second_level_yards_total, offense_open_field_yards,
               offense_open_field_yards_total, offense_standard_downs_ppa,
               offense_standard_downs_success_rate, offense_standard_downs_explosiveness,
               offense_passing_downs_ppa, offense_passing_downs_success_rate,
               offense_passing_downs_explosiveness, offense_rushing_plays_ppa,
               offense_rushing_plays_total_ppa, offense_rushing_plays_success_rate,
               offense_rushing_plays_explosiveness, offense_passing_plays_ppa,
               offense_passing_plays_total_ppa, offense_passing_plays_success_rate,
               offense_passing_plays_explosiveness, defense_plays, defense_drives,
               defense_ppa, defense_total_ppa, defense_success_rate, defense_explosiveness,
               defense_power_success, defense_stuff_rate, defense_line_yards,
               defense_line_yards_total, defense_second_level_yards,
               defense_second_level_yards_total, defense_open_field_yards,
               defense_open_field_yards_total, defense_standard_downs_ppa,
               defense_standard_downs_success_rate, defense_standard_downs_explosiveness,
               defense_passing_downs_ppa, defense_passing_downs_success_rate,
               defense_passing_downs_explosiveness, defense_rushing_plays_ppa,
               defense_rushing_plays_total_ppa, defense_rushing_plays_success_rate,
               defense_rushing_plays_explosiveness, defense_passing_plays_ppa,
               defense_passing_plays_total_ppa, defense_passing_plays_success_rate,
               defense_passing_plays_explosiveness, opponent_conference,
               offense_total_plays, offense_total_havoc_events,
               offense_front_seven_havoc_events, offense_db_havoc_events,
               offense_havoc_rate, offense_front_seven_havoc_rate, offense_db_havoc_rate,
               defense_total_plays, defense_total_havoc_events,
               defense_front_seven_havoc_events, defense_db_havoc_events,
               defense_havoc_rate, defense_front_seven_havoc_rate, defense_db_havoc_rate,
               spread_final, total_final, line_implied_points_final,
               points_vs_line_implied_final, ats_margin_final, covered_final,
               spread_open, total_open, line_implied_points_open,
               points_vs_line_implied_open, ats_margin_open, covered_open,
               team_rank, record_before_display, pregame_elo, postgame_elo,
               /* Passengers: the page links and draws with these; the sheet prints
                  neither. R-287 — the page had 996 team anchors and one destination
                  because this list was never widened when the view was. */
               team_slug, opponent_team_slug, team_logo_url, opponent_logo_url,
               opponent_rank,
               dense_rank() over (order by {SCORES_GAME_ORDER}) as game_no,
               count(*) over () as rows_in_scope
        from srv_game_team
        where season = :season and season_type = :season_type
          and (:week is null or week = :week)
          /* The Division filter, mapped through THIS view's spine. srv_game answers it with
             `is_fbs_game`; srv_game_team had only the TEAM's `classification`, and narrowing
             on that would keep the FBS half of an FBS-vs-FCS game and lose the other — one
             row for a game that has two. So `is_fbs_game` was added to srv_game_team with
             srv_game's predicate copied literally, and both rows of a game carry the same
             value.
             BLOCK comment, not `--`: read_sheet() flattens this onto ONE line, so a line
             comment would swallow every clause after it. And the wording avoids d-r-o-p as
             well as j-o-i-n — the contract's FORBIDDEN regex reads comments, and it caught
             this comment's first draft on the second of those. */
          and (:division = 'all' or is_fbs_game)
          and (:conference is null or conference = :conference
               or opponent_conference = :conference)
          /* R-278. THE ONE PLACE THE TWO SURFACES LEGITIMATELY DIFFER, PARAMETERISED SO IT
             IS VISIBLE AS ONE. The workbook is a data extract and carries `Completed` as a
             column, so it wants every row; the Scores PAGE is a results page and wants
             results. A second query for the page is how Schedule and the export drifted in
             R-184, so the divergence is a bound value rather than a second statement.
             Defaulted to false — the extract's behaviour is the unchanged one. */
          and (not :completed_only or is_completed)
        order by {SCORES_GAME_ORDER}, is_home
        limit {ROW_CAP}
    """,
        SCORES_COLUMNS,
        # The header block is the only place a reader learns the grain, and every surprise on
        # this sheet comes from it.
        note="One row per team per game, so a game is TWO rows and a week of 83 games is 166 "
             "rows. Sorted chronologically, away row then home row, with Game # banding the "
             "pair. Season totals are wrong here unless you filter to one team. Possession "
             "min should total 60.00 across a game's two rows; about 4% of games do not, "
             "which is the source data rather than the arithmetic.",
        derived={"possession_minutes": _possession_minutes},
        display={"is_home": _yes_no, "is_neutral_site": _yes_no, "is_completed": _yes_no,
                 # WORDS, NOT MARKS (Marc, R-262). Schedule uses ■/□ for `Winner covered`
                 # because it sits in a dense block of verdict columns that share one legend
                 # and one glyph vocabulary. Scores has two cover columns among 144, most of
                 # them numeric, and a reader arriving at column AC has no reason to have
                 # read a legend on the Index. A word needs no key.
                 #
                 # It also means the column filters and sorts on something a human types:
                 # `Yes` in a filter box beats hunting for a glyph in a drop-down list.
                 "covered_final": _cover_word,
                 # BLANK, not a dash, when the spread never moved — see _cover_word_or_blank.
                 "covered_open": _cover_word_or_blank},
        field_category=SCORES_CATEGORY,
        integer_fields=SCORES_INTEGER_FIELDS,
        site_precision=SCORES_SITE_PRECISION,
        decimals=SCORES_DECIMALS,
        band_field=SCORES_BAND_FIELD,
        # R-289. Both links Marc asked for, with the slug column NAMED rather than derived.
        link_fields={"team": ("team", "team_slug"), "matchup_url": "matchup"},
        # TEN COLUMNS, ENDING ON `Pts for` — SET BY THE SITE, NOT BY THIS SHEET (R-265).
        #
        # Marc chose the page's frozen block as Game # · Date · Team · Pts for. The site can
        # freeze four columns that are not adjacent; Excel can only freeze a contiguous
        # PREFIX, so the closest honest analogue is everything up to and including the last
        # of those four. A through J.
        #
        # `Pts for` AND NOT `Pts against`, AND THE REASON IS THE GRAIN. Each row is one team,
        # so `Pts for` is that team's score and `Pts against` is the other team's — which is
        # already on screen one row away, in its own `Pts for`. Freezing both would freeze
        # the same two numbers twice and spend a column doing it.
        #
        # It also stays under the twelve-column budget the freeze test measured against real
        # widths (twenty columns came to 200 characters against the ~110-130 an Excel window
        # shows). Held by LABEL and resolved to an index, never a hardcoded letter.
        freeze_before="Pts against"),

    Sheet("Odds", "srv_odds_board", """
        select start_date_et, week, away_team_display, home_team_display,
               provider_display, spread, spread_open, total, total_open,
               home_moneyline, away_moneyline,
               home_implied_probability, away_implied_probability, devig_method,
               is_best_home_spread, is_best_away_spread, snapshot_ts,
               count(*) over () as rows_in_scope
        from srv_odds_board
        where season = :season and is_latest_snapshot
          and (:week is null or week = :week)
        order by start_date_et, game_id, provider_display
        limit {ROW_CAP}
    """, [
        ("start_date_et", "Kickoff"), ("week", "Wk"),
        ("away_team_display", "Away"), ("home_team_display", "Home"),
        ("provider_display", "Book"),
        ("spread", "Spread"), ("spread_open", "Open"),
        ("total", "O/U"), ("total_open", "O/U open"),
        ("home_moneyline", "Home ML"), ("away_moneyline", "Away ML"),
        ("home_implied_probability", "Home implied"),
        ("away_implied_probability", "Away implied"),
        ("devig_method", "De-vig method"),
        ("is_best_home_spread", "Best home"), ("is_best_away_spread", "Best away"),
        ("snapshot_ts", "Snapshot"),
    ]),

    Sheet("Edges", "srv_edge_finder", """
        select week, away_team, home_team, market, edge_unit, edge_value, edge_magnitude,
               model_name, confidence_bucket,
               spread_home_perspective, predicted_margin_home_perspective,
               market_implied_home_win_probability, predicted_home_win_probability,
               actual_home_cover, cover_correct, home_win_correct, is_out_of_sample_week,
               count(*) over () as rows_in_scope
        from srv_edge_finder
        where season = :season
          and (:week is null or week = :week)
        order by edge_magnitude desc
        limit {ROW_CAP}
    """, [
        ("week", "Wk"), ("away_team", "Away"), ("home_team", "Home"),
        ("market", "Market"), ("edge_unit", "Unit"),
        ("edge_value", "Edge"), ("edge_magnitude", "Edge size"),
        ("model_name", "Model"), ("confidence_bucket", "Confidence"),
        ("spread_home_perspective", "Market spread"),
        ("predicted_margin_home_perspective", "Model margin"),
        ("market_implied_home_win_probability", "Market win prob"),
        ("predicted_home_win_probability", "Model win prob"),
        ("actual_home_cover", "Home covered"), ("cover_correct", "Cover hit"),
        ("home_win_correct", "Winner hit"),
        # AC-15.4: the flag is per ROW, not a sheet-level footnote. A workbook gets
        # filtered and sorted, and a caption does not survive that.
        ("is_out_of_sample_week", "Out-of-sample week"),
    ], has_predictions=True),

    Sheet("Standings", "srv_standings", """
        select conference, tiebreak_rank, school, wins, losses, ties,
               conference_wins, conference_losses, win_pct,
               points_for, points_against, point_differential, tiebreak_basis,
               count(*) over () as rows_in_scope
        from srv_standings
        where season = :season and classification in ('fbs','fcs')
        order by conference, tiebreak_rank
        limit {ROW_CAP}
    """, [
        ("conference", "Conference"), ("tiebreak_rank", "#"), ("school", "Team"),
        ("wins", "W"), ("losses", "L"), ("ties", "T"),
        ("conference_wins", "Conf W"), ("conference_losses", "Conf L"),
        ("win_pct", "Win %"),
        ("points_for", "PF"), ("points_against", "PA"),
        ("point_differential", "Diff"), ("tiebreak_basis", "Tiebreak"),
    ], note="Season-scoped: standings are a season figure, not a week one."),

    # Not week-scoped, and deliberately so. These two describe the export rather than adding
    # to it: which models produced the predicted columns, and what every field means. Both
    # are small, and shipping predictions without either would be shipping numbers with no
    # way to check what they are.
    # Every cut, not just the headline. The view stacks overall / week / conference /
    # confidence / probability-decile rows, and segment_type is carried so a filter in the
    # workbook separates them — which is the one thing a spreadsheet is genuinely better at
    # than a page. The headline table on the site reads `segment_type = 'overall'`; here
    # the reader gets the same data plus the breakdowns and does the filtering themselves.
    #
    # cover_scored travels beside ats_accuracy_pct on purpose. A rate without its
    # denominator is the defect AC-G.33 exists to prevent, and it is worse in a workbook,
    # where the column gets averaged.
    Sheet("Model performance", "srv_model_performance", """
        select segment_type, segment_value, model_name, model_version, model_family,
               split, season, is_out_of_sample_week, games,
               mean_absolute_margin_error, winner_accuracy_pct, winner_scored,
               ats_accuracy_pct, cover_scored,
               mean_predicted_home_win_probability, actual_home_win_rate,
               brier_score, log_loss, attribution,
               count(*) over () as rows_in_scope
        from srv_model_performance
        order by model_name, segment_type, segment_order, segment_value
        limit {ROW_CAP}
    """, [
        ("segment_type", "Segment"), ("segment_value", "Segment value"),
        ("model_name", "Model"), ("model_version", "Version"),
        ("model_family", "Family"), ("split", "Split"), ("season", "Season"),
        ("is_out_of_sample_week", "Out-of-sample week"), ("games", "n"),
        ("mean_absolute_margin_error", "Margin MAE"),
        ("winner_accuracy_pct", "SU %"), ("winner_scored", "SU graded"),
        ("ats_accuracy_pct", "ATS %"), ("cover_scored", "ATS graded"),
        ("mean_predicted_home_win_probability", "Model says"),
        ("actual_home_win_rate", "Actually happened"),
        ("brier_score", "Brier"), ("log_loss", "Log loss"),
        ("attribution", "Attribution"),
    ], has_predictions=True, scoped=False,
        note="Every segment: overall, by week, by conference, by confidence, and by "
             "predicted-probability decile. Filter on Segment. Conference rows count a "
             "game under both teams' conferences, so they exceed the overall row."),

    Sheet("Data dictionary", "srv_data_dictionary", """
        select layer, table_name, column_name, data_type, is_nullable,
               description_status, column_description,
               count(*) over () as rows_in_scope
        from srv_data_dictionary
        where layer = 'serving'
        order by table_name, ordinal_position
        limit {ROW_CAP}
    """, [
        ("layer", "Layer"), ("table_name", "Table"), ("column_name", "Column"),
        ("data_type", "Type"), ("is_nullable", "Nullable"),
        ("description_status", "Status"), ("column_description", "Description"),
    ], scoped=False,
        note="AC-15.8 / AC-16.7: generated from the same view the site's Data Dictionary "
             "page reads, so the workbook and the page cannot disagree."),
]

# What the workbook writes, and what it does not write YET. Split rather than filtered, so
# adding a converted sheet is moving one name and cannot be done by accident.
SHIPPED = ("Schedule", "Scores")
SHEETS = [s for s in _ALL_SHEETS if s.name in SHIPPED]
PENDING_SHEETS = [s for s in _ALL_SHEETS if s.name not in SHIPPED]
PENDING_REASON = ("not converted to the new layout yet; it ships in a later pass rather "
                  "than mixing two header layouts in one file")

# Which page each sheet came from, for the Index's link back. Held here rather than on Sheet
# because it is a fact about the SITE, and the pending sheets need it the day they ship.
PAGE_FOR_SHEET = {
    "Schedule": "schedule", "Scores": "scores", "Odds": "odds", "Edges": "edges",
    "Standings": "standings", "Model performance": "performance",
    "Data dictionary": "dictionary",
}
# The split is asserted at IMPORT, not in a test: moving a sheet between the two lists is a
# one-word edit, and this is what makes "I shipped a sheet" and "I lost a sheet" different
# events. Seven sheets exist; two ship.
assert len(SHEETS) == 2 and len(PENDING_SHEETS) == 5
assert set(SHIPPED) <= {s.name for s in _ALL_SHEETS}

# Conditional formatting goes on the columns a reader is scanning for outliers. Anything
# else would be decoration, and a spreadsheet where everything is highlighted highlights
# nothing.
# AC-15.8 says the workbook's headers match the site's. Where they deliberately do not,
# the reason is recorded here rather than left as a discrepancy someone rediscovers — and
# the test enforces that any divergence is one of these and not an accident.
EXPORT_ONLY_LABELS = {
    "away_points": "the site's header is blank, because the score renders beside the team "
                   "name; a spreadsheet column cannot have a blank header",
    "home_points": "as away_points",
    "actual_margin": "the page can say 'Margin' because the column sits between the two "
                     "teams' scores and the direction is obvious in context; a workbook "
                     "travels without that context, so the header states away-minus-home",
    "segment_value": "the page labels this per tab — Week, Conference, Confidence, "
                     "Predicted band — because each tab shows one segment type. The sheet "
                     "stacks all five, so its header has to name the column rather than "
                     "the cut, and Segment sits beside it saying which cut each row is",
    # --- R-182 TRAP 2 made these mandatory to resolve rather than merely tidy. -----------
    # An Excel Table requires unique, non-empty headers, and the Schedule sheet carries both
    # halves of two pairs the page only ever shows one of at a time.
    # The next three are all the same divergence: the page can abbreviate because the column
    # sits inside a labelled block with neighbours giving it context. A workbook column
    # travels alone, gets sorted away from its neighbours, and is read a month later.
    "total_points": "the page's 'Pts' sits in the box score beside the two teams' scores. "
                    "Alone in a spreadsheet, 'Pts' does not distinguish the game total "
                    "from either team's",
    "favorite_covered": "the page's 'Fav cover' is a narrow column in a strip of marks. "
                        "The sheet spells it out because it sits beside 'Winner covered' "
                        "and the two answer DIFFERENT questions — this one can also say "
                        "no_favorite",
    "is_neutral_site": "R-026 then R-101: the Schedule page shows a neutral site as an ICON "
                       "WITH NO LABEL — a deliberate exception to the site's "
                       "glyph-plus-label rule, decided by Marc against a small known user "
                       "base — and R-101 folded that icon into the shared Game column, so "
                       "the site now has NO COLUMN OF ITS OWN for this field at all. A "
                       "spreadsheet has no icon convention and no tooltip, so the column "
                       "keeps a spelled-out header here. This is the divergence being real "
                       "rather than an oversight",
}

# A one-character column is unreadable however short its data is. Measured against the
# narrowest real content on the sheet — "Wk" holds one or two digits — 6 leaves room for the
# filter button a Table puts in every header cell.
MIN_COLUMN_WIDTH = 6

# ...AND A COLUMN IS NEVER NARROWER THAN THE LONGEST WORD IN ITS HEADER, up to this cap.
#
# Measuring the data alone (R-217) was right and went too far: at a flat floor of 6, Excel
# breaks "Favorite" mid-word into "Favori / te", which is harder to read than the wide
# column it replaced. Marc: "shrinking is a little aggressive, to the point it's making some
# of the headers really difficult to read."
#
# So the floor is the longest WORD, not the whole label. "Conference game" gets 10 rather
# than the 15 that started this, and rather than the 6 that broke the word. The cap keeps one
# very long word ("Precipitation") from undoing the change for everyone else — that one still
# breaks, and one broken word is a better trade than a 13-wide column of "0.00".
HEADER_WORD_CAP = 12

# Excel's "character width" is the width of a `0` in the workbook font, and real letters are
# wider than a zero — so a six-character word does not fit in a six-wide column. Measured
# against the headers in this sheet, 0.7 is enough to stop the last letter wrapping.
WORD_PADDING = 0.7


def effective_width(label: str, measured: Optional[float] = None) -> float:
    """The width a column actually gets: hand-set if there is one, measured otherwise, and
    never below what the header's longest word needs.

    A function rather than an expression inline in `build`, because the tests have to be able
    to ask the same question the writer answers — asserting a hardcoded 6.7 would restate the
    arithmetic instead of checking it.
    """
    longest_word = max((len(w) for w in str(label).split()), default=0)
    floor = min(longest_word, HEADER_WORD_CAP) + WORD_PADDING
    base = WIDTH_OVERRIDES.get(label, measured if measured is not None else floor)
    return max(base, floor)


# WIDTHS MARC SET BY HAND, IN EXCEL CHARACTER UNITS.
#
# These OVERRIDE the measurement for these columns and nothing else. Measuring is right by
# default — it is what stopped a 15-character header sizing a 3-character column — but it
# cannot know that a column is worth an extra character because of how the header wraps, or
# worth one fewer because the reader never scans it. Marc read these off Excel's own width
# dialog with the real file in front of him, which is better information than the measurement
# has.
#
# Keyed by LABEL, so they follow the column if the order changes again.
#
# NOTE the US spelling below: Marc's CSV writes "Favourite covered" and he has since asked for
# American spelling throughout, so the shipped label is "Favorite covered". That is the ONE
# place the sheet deliberately departs from the CSV, and CSV_LABEL_OVERRIDES records it so the
# reconciliation test can tell a decision from a typo.
# The single place the built sheet departs from Marc's column-order CSV, with its reason.
# Anything not in here must match the CSV exactly.
CSV_LABEL_OVERRIDES = {
    "Favourite covered": "Favorite covered",   # US spelling, Marc 2026-09-03
}

WIDTH_OVERRIDES = {
    "Kickoff": 11.5,
    "Winner covered": 8.0,
    "Final margin": 5.85,
    "Season": 5.6,
    "Wind mph": 5.5,
    "Pred margin": 6.5,
    "Home win prob": 7.7,
}

# Excel's default row height is 15pt for an 11pt font, and a wrapped line needs about 13pt
# once the header's own padding is allowed for. Both are conventions rather than constants
# openpyxl exposes, so they are named here rather than appearing as bare numbers.
HEADER_LINE_HEIGHT = 13.0

# 50pt, AND IT IS A FLOOR RATHER THAN A FIXED HEIGHT.
#
# Marc: the header needs "enough room to be above the drop-down filters". An Excel Table puts
# a filter button INSIDE the header cell, bottom-right, and it overlaps the last line of a
# wrapped label — so a header sized to exactly fit its text has its final word sitting under
# the button.
#
# Setting it to 50 outright would give back what R-217 bought: a longer header in a future
# column would clip silently again. So the computed height still wins whenever it is taller,
# and 50 is only the minimum. Today the computation returns 39, so 50 applies.
MIN_HEADER_HEIGHT = 50.0


def _header_height(labels_and_widths) -> float:
    """How tall the header row must be for every label to wrap without clipping.

    A column's width is in CHARACTERS, so a label needs ceil(len / width) lines at minimum —
    but Excel wraps on word boundaries, so a long word cannot be split and a short width
    wastes the tail of a line. Wrapping properly rather than dividing is the difference
    between "Margin (away−home)" reporting two lines and reporting the four it really takes.
    """
    worst = 1
    for label, width in labels_and_widths:
        text, usable = str(label), max(int(width), 1)
        lines, current = 1, ""
        for word in text.split():
            candidate = f"{current} {word}".strip()
            if len(candidate) <= usable:
                current = candidate
                continue
            if current:
                lines += 1
            # A single word longer than the column still has to go somewhere: Excel breaks
            # it across lines rather than clipping, so account for that instead of assuming
            # every word fits.
            current = word
            while len(current) > usable:
                lines += 1
                current = current[usable:]
        worst = max(worst, lines)
    # The FLOOR IS NOT APPLIED HERE. This answers "how tall must it be to not clip", and the
    # caller decides whether a taller minimum applies for a different reason (filter buttons).
    # Folding the two together made a unit test comparing a short label to a long one return
    # the same number for both, and it stopped being a test.
    return worst * HEADER_LINE_HEIGHT


# TWO SCALES, BECAUSE THE COLUMNS MEAN DIFFERENT THINGS.
#
# A colour scale asserts good and bad. That is fair on a quantity with a direction — a bigger
# edge, a better point differential — and wrong on one that only has a SIDE.
#
# `actual_margin` and `predicted_margin` are both AWAY MINUS HOME, so negative means the home
# team. Measured across 106,554 completed games: `actual_margin < 0` means a home win with no
# exceptions at all, and it is 64% of FBS games since 2024. The default scale painted red at
# the minimum, so the most common outcome in the sport was alarming on nearly two rows in
# three — Marc's point, and the numbers agree with it.
#
# Reversed for those two, so red marks the AWAY win: the less expected result, which is the
# one worth a reader's eye.
COLOUR_SCALE_FIELDS = {"edge_value", "edge_magnitude", "point_differential"}

# Away-minus-home. Negative is the home team and the common case, so the scale runs the other
# way. Kept as its own set rather than a flag on each field, because the question "which side
# does negative mean" is the whole reason these differ.
REVERSED_SCALE_FIELDS = {"predicted_margin"}

# R-220. Marc supplied a screenshot of the exact rule: Data Bar · automatic min and max ·
# direction Context · SOLID fill · positive BLUE, negative RED · solid black borders on both ·
# axis position MIDPOINT · black axis. `actual_margin` moves off the colour scale to get it;
# `predicted_margin` keeps its scale because he named one column.
DATA_BAR_FIELDS = {"actual_margin"}

# ROUTE (b), KEPT REACHABLE ON PURPOSE.
#
# Prompt 039 said: do the x14 route, and if it fights back ship the single-colour bar and say
# so. It fought back once — the first attempt put its block out of CT_Worksheet order and
# Excel opened the Schedule tab empty. The CAUSE was the insertion point, not the extension,
# and it is fixed and now checkable. But "we know why it broke" is not the same as "it works
# in Excel", so the simpler rendering stays one flag away rather than needing to be rewritten
# under pressure.
#
# `False` = Marc's full rule via x14. `True` = a plain single-colour bar with a midpoint axis,
# entirely through openpyxl's supported API — honest, and not what he drew.
DATA_BAR_SIMPLE = os.getenv("CFDB_SIMPLE_DATA_BAR", "").lower() in ("1", "true", "yes")

# WHY THIS IS HAND-WRITTEN XML AND NOT A DataBarRule.
#
# openpyxl's DataBarRule writes the LEGACY `<dataBar>` element, which carries one colour, a
# min and a max. It has NO negative fill colour, NO border colours and NO axis position — so
# three of the seven things in Marc's screenshot cannot be said with it at all. Those live in
# the CF14 extension (`x14:dataBar`), which openpyxl does not model.
#
# The cost is honest: this is the class of hand-written XML that produces the repair prompt
# AC-15.6 forbids, so it is verified by OPENING THE FILE, not by openpyxl reading it back.
#
# The shape is a legacy `<dataBar>` for readers that only understand CF12, plus an `x14`
# block carrying the rest, tied together by an `x14:id` GUID. Excel prefers the extension and
# ignores the legacy element; anything older degrades to a plain blue bar rather than to
# nothing.
DATA_BAR_GUID_PREFIX = "{FFFFFFFF-CCCC-BBBB-AAAA-"
# THE BAR POINTS THE WAY THE SCORE DID, AND THE COLOUR FOLLOWS THE SAME RULE AS THE SCALES.
#
# `actual_margin` is away minus home, so the bar grows LEFT when the home team won and RIGHT
# when the away team did. Blue for the home win because it is 64% of games and a red common
# case reads as an alarm that never means anything; red for the away win, which is the result
# a reader is scanning for.
#
# Marc's original screenshot said positive blue and negative red. That was before either of us
# had noticed which side of zero the home team lives on.
DATA_BAR_POSITIVE = "FFFF0000"      # away won — the less expected result
DATA_BAR_NEGATIVE = "FF638EC6"      # home won — the common one
DATA_BAR_AXIS = "FF000000"          # black axis and black borders


def _data_bar_guid(index: int) -> str:
    """A stable per-rule GUID. Deterministic, so two builds of one scope are byte-identical —
    a random one would make a diff of two workbooks useless."""
    return f"{DATA_BAR_GUID_PREFIX}{index:012X}}}"


def _mark_text_column(tab, span: str) -> None:
    """Record that this range holds TEXT ON PURPOSE, so Excel stops warning about it.

    "1-0" is a won-lost record. Excel sees text that could be a number or a date, flags every
    cell with a green corner and offers to convert it — which would turn 1-0 into 1 January.
    The advice is correct in general and wrong here, and 166 green triangles across two
    columns is noise the reader has to learn to ignore.

    `ignoredErrors` is the OOXML way to say "I meant this". openpyxl MODELS the element and
    never writes it — `Worksheet` has no `ignored_errors` attribute — so it goes through the
    same post-save injection as the data bar, and past the same order check.
    """
    tab._cfdb_text_columns = getattr(tab, "_cfdb_text_columns", [])
    tab._cfdb_text_columns.append(span)


def _add_data_bar(tab, span: str) -> None:
    """Record that this range wants Marc's data bar. The XML is injected after save.

    Under DATA_BAR_SIMPLE the supported API is used instead and nothing is injected, so the
    fallback shares no code with the thing it is a fallback for.
    """
    if DATA_BAR_SIMPLE:
        from openpyxl.formatting.rule import DataBarRule
        tab.conditional_formatting.add(span, DataBarRule(
            start_type="min", end_type="max", color=DATA_BAR_POSITIVE[2:],
            showValue=True, minLength=None, maxLength=None))
        return
    tab._cfdb_data_bars = getattr(tab, "_cfdb_data_bars", [])
    tab._cfdb_data_bars.append(span)


FLAG_FIELDS = {"cover_correct", "home_win_correct", "is_upset", "is_out_of_sample_week",
               "is_best_home_spread", "is_best_away_spread", "actual_home_cover"}


# R-216. TWO SETS AND A RULE, REPLACING A MEMBERSHIP TUPLE THAT WAS WRONG THREE TIMES.
#
# The old `number_format` tested one hardcoded tuple and let everything else fall through to
# a decimal format. That shape produced three separate defects at once — decimals on a rank,
# decimals on a game id, and `season` rendering as "2,025" because it WAS in the tuple and
# the tuple's only format grouped thousands. A list every new integer column has to be
# remembered into will be wrong again on the next column.
#
# THE RULE THAT DECIDES, stated once so the next column is obvious:
#
#     A THOUSANDS SEPARATOR IS FOR A QUANTITY YOU MIGHT TOTAL.
#     A rank, an id, a season and a week are LABELS THAT HAPPEN TO BE NUMERIC,
#     and a comma in one of them is a bug.
#
# Attendance is a count you would sum, so 74,109 is right. Season 2025 is a name, so "2,025"
# is nonsense. That is the whole distinction.
COUNT_FIELDS = {
    "attendance", "games", "home_points", "away_points", "total_points",
    "wins", "losses", "ties", "conference_wins", "conference_losses",
    "points_for", "points_against", "n", "winner_scored", "cover_scored",
}

# Numbers you never sum. Anything matching these patterns joins them automatically, which is
# the part the tuple could not express — a `_rank` or an `_id` added next month is covered
# without anyone remembering this file exists.
# `best_rank_in_game` is here rather than caught by the suffix, and it is worth saying why:
# the suffix rule sees the END of a name, and this one carries `rank` in the MIDDLE. That
# gap is the reason the test below enumerates the real columns and asserts on meaning rather
# than re-implementing this pattern — a test that repeated the rule would have agreed with
# it and missed the same column.
PLAIN_INTEGER = {"season", "week", "tiebreak_rank", "game_id", "bin_index",
                 "best_rank_in_game"}
PLAIN_INTEGER_SUFFIXES = ("_rank", "_id")


def is_plain_integer(field: str) -> bool:
    """A numeric label: no decimal point, no thousands separator."""
    return field in PLAIN_INTEGER or field.endswith(PLAIN_INTEGER_SUFFIXES)


def number_format(field: str, decimals: Optional[int] = None,
                  integers: frozenset = frozenset(),
                  site_precision: frozenset = frozenset()) -> str:
    """Excel format string at the SAME precision the site renders (AC-G.31, AC-15.7).

    Decimals are derived from fmt.precision_for rather than restated, so a column cannot
    read 1 dp on screen and 2 dp in the workbook.

    A SHEET MAY OVERRIDE BOTH HALVES, and R-259 is why. Scores is 131 columns of PPA,
    success rates and explosiveness, and `precision_for` gives those 3, 1 and 1 — a success
    rate stored as 0-1 would render `0.3`. So that sheet passes `decimals=2` and its own set
    of naturally-integer columns, MEASURED over the data rather than guessed from the names.

    The two arguments are separate on purpose: `integers` decides whether there is a decimal
    point at all, `decimals` decides how many digits follow one. R-216's rule still owns the
    thousands separator either way — a season is not "2,025" on any sheet.
    """
    if is_plain_integer(field):
        return "0"
    if field in integers:
        # A count, so it takes the separator R-216 reserves for quantities you might total.
        return "#,##0"
    if field in COUNT_FIELDS:
        return "#,##0"
    if field.endswith("moneyline"):
        return "+#,##0;-#,##0"
    places = fmt.precision_for(field) if decimals is None or field in site_precision \
        else decimals
    return "#,##0." + "0" * places


# DATE FORMATS PER FIELD, because a kickoff and an audit timestamp are read for different
# reasons.
#
# Marc: force the kickoff to `mmm-dd hh:mm`. A reader scanning a week's slate wants "Sep-05
# 19:00" — the year is redundant when the Season column is right there and the file name
# carries it too. `As of` and `Line taken` keep the full stamp: those exist to be checked
# against something outside the file, and a provenance timestamp without a year is not one.
#
# In Excel `mm` means MINUTES after an hour token and MONTH otherwise, which is why the month
# here is `mmm` and there is no ambiguity to resolve.
DATE_FORMATS = {"start_date_et": "mmm-dd hh:mm"}
DEFAULT_DATE_FORMAT = "yyyy-mm-dd hh:mm"


def date_format(field: str) -> str:
    return DATE_FORMATS.get(field, DEFAULT_DATE_FORMAT)


def rendered_date_width(field: str) -> int:
    """How wide a formatted date actually prints.

    The width measurement used a hardcoded 17 for every datetime, which was right for
    `yyyy-mm-dd hh:mm` and would have left a `mmm-dd hh:mm` column five characters too wide —
    the same defect R-217 fixed for headers, arriving from the date side. Counted from the
    format, so a shorter format takes a narrower column without anyone remembering to.
    """
    pattern = date_format(field)
    # Each `mmm` prints three characters, every other token prints its own length, and the
    # separators print themselves.
    return len(pattern) + (1 if "mmm" in pattern else 0)


def filename(season: int, week: Optional[int], generated: datetime) -> str:
    """AC-15.10: the filename states its own scope.

    A folder of files called `export.xlsx` is a folder of files nobody can tell apart, and
    the scope is the single most important thing about an export whose whole design is
    being bounded.
    """
    scope = f"week{week:02d}" if week is not None else "season"
    return f"cfdb_{scope}_{season}_{generated:%Y%m%d}.xlsx"


def describe_scope(season: int, week: Optional[int], season_type: str,
                   conference: Optional[str], division: str = "fbs") -> str:
    bits = [f"season {season}", season_type]
    bits.append(f"week {week}" if week is not None else "all weeks")
    if conference:
        bits.append(conference)
    # Mirrors `GameScope.describe()`, which also omits the division when it is "all" —
    # naming a filter that excludes nothing reads as a narrowing that did not happen.
    if division != "all":
        bits.append(division.upper())
    return ", ".join(bits)


class IndexRow(NamedTuple):
    """One line of the Index's sheet inventory, and what the caller shows in the preview.

    `rows` and `rows_in_scope` travel together everywhere for the reason SheetRead names.
    """
    name: str
    view: str
    rows: int
    rows_in_scope: int
    note: str = ""

    @property
    def truncated(self) -> bool:
        return self.rows_in_scope > self.rows


class SheetRead(NamedTuple):
    """What one sheet's query returned.

    `rows` and `rows_in_scope` are DELIBERATELY TWO FIELDS. They were one number for as long
    as the cap was invisible, and that is precisely how a season export could hold 400 of
    3,745 games while every number on the page and in the file agreed with every other.
    """
    frame: Optional[pd.DataFrame]
    omission: Optional[str]
    rows_in_scope: int = 0

    @property
    def rows(self) -> int:
        return 0 if self.frame is None else len(self.frame)

    @property
    def truncated(self) -> bool:
        return self.rows_in_scope > self.rows


def read_sheet(sheet, season, week, season_type, conference, division="fbs",
               completed_only: bool = False):
    """Fetch one sheet's rows. Returns a SheetRead.

    The two failures here are NOT the same thing and the first draft of this treated them
    as one. AC-15.5 says a sheet whose SOURCE IS MISSING is omitted with a note — that is a
    view cfdb has not built. A connection failure, a permission error or a typo'd column is
    not that, and reporting it as "omitted, no rows in this scope" produces a workbook that
    calmly says the data is unavailable while the real problem is that nothing can be read
    at all. Every sheet was omitted for exactly that reason on the first run of this module,
    and the message gave no way to tell.

    So: a missing relation is an omission, and anything else is raised.
    """
    params = {"season": season, "week": week if sheet.scoped else None,
              "season_type": season_type, "conference": conference, "division": division,
              # R-278. The workbook takes every row; the page passes True. Named here with
              # the extract's default so a sheet that never mentions it is unaffected.
              "completed_only": completed_only}
    wanted = set(re.findall(r":(\w+)", sheet.sql))
    try:
        df = query(" ".join(sheet.sql.split()),
                   {k: v for k, v in params.items() if k in wanted})
    except Exception as exc:                                       # noqa: BLE001
        message = str(exc).lower()
        if "does not exist" in message or "undefined table" in message:
            return SheetRead(None, f"{sheet.view} has not been built yet")
        raise
    if df.empty:
        return SheetRead(None, f"{sheet.view} returned no rows in this scope")
    # ONE QUERY ANSWERS BOTH QUESTIONS. `count(*) over ()` is a window function, and Postgres
    # evaluates window functions BEFORE the LIMIT — so every returned row carries the size of
    # the full result set. A second `select count(*)` would be a second implementation of
    # "what is in scope", free to disagree with the first; this cannot.
    in_scope = int(df["rows_in_scope"].iloc[0]) if "rows_in_scope" in df.columns else len(df)
    return SheetRead(df, None, in_scope)


def _clean(value):
    """One cell value openpyxl will accept, preserving the null/zero distinction.

    Two things here are not cosmetic. A tz-aware datetime makes openpyxl raise outright,
    and NaN written into a cell is what produces the repair prompt AC-15.6 forbids — the
    file is structurally valid XML and Excel still refuses it. Both are converted rather
    than left to fail at write time, where the traceback would name the cell and not the
    reason.
    """
    if value is None or value is pd.NaT:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, str) and value.strip() == "":
        # The prediction pack writes an empty string where it has no confidence bucket. An
        # empty-string cell is not an empty cell — it is text that looks like nothing, and
        # it breaks COUNTBLANK and sorts ahead of real values. Blank and absent are the
        # same claim here, exactly as they are on the page.
        return None
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime) and value.tzinfo is not None:
        value = value.astimezone(timezone.utc).replace(tzinfo=None)
    if hasattr(value, "item"):                       # numpy scalar
        value = value.item()
    from decimal import Decimal
    if isinstance(value, Decimal):
        return float(value)
    return value


# CT_Worksheet's children are an ORDERED SEQUENCE and Excel validates it. These are the
# elements the schema places AFTER `conditionalFormatting`, in schema order — so the first
# of them that appears in a sheet is the latest point a conditionalFormatting block may be
# inserted. `extLst` is deliberately last: it is the final child, after `tableParts`.
# The preferred anchor: openpyxl's own conditional formatting is provably in the right place,
# so appending after the last one cannot be out of order. Named so the fallback path below is
# reachable in a test — a guard nothing can exercise is a guard nobody should trust.
CF_ANCHOR = "</conditionalFormatting>"

# Everything CT_Worksheet places after `ignoredErrors`.
SHEET_ELEMENTS_AFTER_IGNORED_ERRORS = (
    "<smartTags", "<drawing", "<legacyDrawing", "<picture", "<oleObjects", "<controls",
    "<webPublishItems", "<tableParts", "<extLst", "</worksheet>",
)

SHEET_ELEMENTS_AFTER_CF = (
    "<dataValidations", "<hyperlinks", "<printOptions", "<pageMargins", "<pageSetup",
    "<headerFooter", "<rowBreaks", "<colBreaks", "<customProperties", "<cellWatches",
    "<ignoredErrors", "<smartTags", "<drawing", "<legacyDrawing", "<picture",
    "<oleObjects", "<controls", "<webPublishItems", "<tableParts", "<extLst",
    "</worksheet>",
)

# The FULL sequence, for validation. Excel refuses a sheet whose children are out of order
# and reports it as "a problem with some content" — no clue which element, no line number.
# This project shipped exactly that once, so the order is now checkable rather than reasoned
# about.
CT_WORKSHEET_SEQUENCE = (
    "sheetPr", "dimension", "sheetViews", "sheetFormatPr", "cols", "sheetData",
    "sheetCalcPr", "sheetProtection", "protectedRanges", "scenarios", "autoFilter",
    "sortState", "dataConsolidate", "customSheetViews", "mergeCells", "phoneticPr",
    "conditionalFormatting", "dataValidations", "hyperlinks", "printOptions",
    "pageMargins", "pageSetup", "headerFooter", "rowBreaks", "colBreaks",
    "customProperties", "cellWatches", "ignoredErrors", "smartTags", "drawing",
    "drawingHF", "picture", "oleObjects", "controls", "webPublishItems", "tableParts",
    "extLst",
)


# Elements that are only ever DIRECT children of <worksheet>. Finding one deeper in the tree
# means an injection went into the wrong parent — which is not an ordering fault and would
# otherwise pass an ordering check silently.
TOP_LEVEL_ONLY = ("ignoredErrors", "conditionalFormatting", "hyperlinks", "tableParts",
                  "dataValidations", "pageMargins", "printOptions")


def sheet_order_violations(sheet_xml: bytes) -> list:
    """Ways this sheet part would be rejected or silently ignored by Excel.

    TWO faults, not one, because the second cost a round:

      ORDER     a direct child of <worksheet> out of CT_Worksheet sequence. Excel refuses the
                file, replaces the part and opens the tab empty.
      NESTING   an element that belongs at the top level found somewhere deeper. Excel does
                NOT complain — it simply ignores it. The `ignoredErrors` block spent a whole
                round inside a `<cfRule>`, perfectly well-formed and doing nothing, and the
                first version of this function could not see it because it walked only direct
                children.

    Returns (element, context) pairs; empty means the part is sound.
    """
    from xml.etree import ElementTree
    main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    root = ElementTree.fromstring(sheet_xml)

    def local(tag):
        return tag[len(main):] if tag.startswith(main) else tag

    rank = {n: i for i, n in enumerate(CT_WORKSHEET_SEQUENCE)}
    last_name, last_rank, out = None, -1, []
    for child in root:
        name = local(child.tag)
        position = rank.get(name)
        if position is None:            # not a schema element we model; skip rather than guess
            continue
        if position < last_rank:
            out.append((name, last_name))
        last_name, last_rank = name, max(last_rank, position)

    direct = {id(child) for child in root}
    for element in root.iter():
        name = local(element.tag)
        if name in TOP_LEVEL_ONLY and id(element) not in direct and element is not root:
            out.append((name, "NESTED — not a direct child of <worksheet>"))
    return out


X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
XM_NS = "http://schemas.microsoft.com/office/excel/2006/main"


def _inject_data_bars(payload: bytes, wanted: dict, text_columns=None) -> bytes:
    """Rewrite the saved workbook so the data-bar ranges carry Marc's full rule.

    `wanted` is {sheet name -> [range, ...]}. Each range gets TWO things written into that
    sheet's XML:

      * a legacy `<conditionalFormatting><cfRule type="dataBar">` block carrying an
        `<extLst>` pointer to the x14 rule, so a reader that only speaks CF12 still draws a
        plain blue bar rather than nothing;
      * an `<x14:conditionalFormatting>` inside the sheet's own `<extLst>`, carrying the
        negative colour, both borders, the midpoint axis and the axis colour — the four
        things the legacy element cannot express.

    Done as a post-save rewrite rather than through openpyxl because openpyxl has no model
    for x14 and silently drops unknown elements on write. Rewriting the finished bytes is
    the only place the extension survives.

    ORDER MATTERS AND EXCEL IS STRICT ABOUT IT. In the sheet's schema `conditionalFormatting`
    comes after `mergeCells`/`hyperlinks` and before `dataValidations`... but crucially both
    it and `extLst` must sit in schema order relative to `tableParts`, which is LAST. Getting
    this wrong is the repair prompt, which is why the insertion points are anchored on
    specific elements rather than appended.
    """
    from xml.etree import ElementTree

    text_columns = text_columns or {}
    if not wanted and not text_columns:
        return payload

    source = zipfile.ZipFile(io.BytesIO(payload))
    # Sheet name -> part name, read from the workbook's own relationships rather than guessed
    # from ordering, because `sheet1.xml` is not reliably the first tab. PARSED, not
    # regexed: the first version matched `name=` before `r:id=` and openpyxl does not
    # guarantee attribute order, so it silently produced an empty target and then tried to
    # read a part called "xl/".
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
          "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
          "pr": "http://schemas.openxmlformats.org/package/2006/relationships"}
    book_xml = ElementTree.fromstring(source.read("xl/workbook.xml"))
    rels_xml = ElementTree.fromstring(source.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {rel.get("Id"): rel.get("Target")
                   for rel in rels_xml.findall("pr:Relationship", ns)}
    part_for = {}
    for element in book_xml.findall("m:sheets/m:sheet", ns):
        target = rel_targets.get(element.get(f"{{{ns['r']}}}id"), "")
        if not target:
            continue
        # The Target may be relative to xl/ ("worksheets/sheet2.xml") or absolute
        # ("/xl/worksheets/sheet2.xml") depending on the writer. Normalise rather than
        # assume — assuming produced "xl/xl/worksheets/sheet2.xml".
        target = target.lstrip("/")
        part_for[element.get("name")] = target if target.startswith("xl/") \
            else "xl/" + target

    rewritten = {}
    counter = 0
    for sheet_name in set(wanted) | set(text_columns):
        spans = wanted.get(sheet_name, [])
        part = part_for.get(sheet_name)
        if not part:
            continue
        body = rewritten.get(part) or source.read(part)
        body = body.decode("utf-8") if isinstance(body, bytes) else body
        legacy, extended = [], []
        for span in spans:
            counter += 1
            guid = _data_bar_guid(counter)
            legacy.append(
                f'<conditionalFormatting sqref="{span}">'
                f'<cfRule type="dataBar" priority="{counter}">'
                f'<dataBar><cfvo type="min"/><cfvo type="max"/>'
                f'<color rgb="{DATA_BAR_POSITIVE}"/></dataBar>'
                f'<extLst><ext uri="{{B025F937-C7B1-47D3-B67F-A62EFF666E3E}}" '
                f'xmlns:x14="{X14_NS}"><x14:id>{guid}</x14:id></ext></extLst>'
                f'</cfRule></conditionalFormatting>')
            extended.append(
                f'<x14:conditionalFormatting xmlns:xm="{XM_NS}">'
                f'<x14:cfRule type="dataBar" id="{guid}">'
                f'<x14:dataBar minLength="0" maxLength="100" gradient="0" '
                f'border="1" negativeBarColorSameAsPositive="0" '
                f'negativeBarBorderColorSameAsPositive="0" '
                f'axisPosition="middle">'
                f'<x14:cfvo type="autoMin"/><x14:cfvo type="autoMax"/>'
                f'<x14:borderColor rgb="{DATA_BAR_AXIS}"/>'
                f'<x14:negativeFillColor rgb="{DATA_BAR_NEGATIVE}"/>'
                f'<x14:negativeBorderColor rgb="{DATA_BAR_AXIS}"/>'
                f'<x14:axisColor rgb="{DATA_BAR_AXIS}"/>'
                f'</x14:dataBar></x14:cfRule>'
                f'<xm:sqref>{span}</xm:sqref>'
                f'</x14:conditionalFormatting>')

        # WHERE THE LEGACY BLOCK GOES, AND THIS IS THE PART THAT BROKE EXCEL.
        #
        # `CT_Worksheet` is an ordered SEQUENCE, not a bag. Excel validates it and refuses
        # the file — "We found a problem with some content" — when a child is out of place,
        # which is exactly what happened: the first version inserted before `<pageMargins>`,
        # and on a sheet with hyperlinks that lands AFTER `<hyperlinks>`. The schema puts
        # `conditionalFormatting` BEFORE `hyperlinks`, so Excel replaced the whole sheet part
        # and the Schedule tab opened empty.
        #
        # Preferred insertion point is right after the conditional formatting openpyxl has
        # already written, because that is provably in the right place. Failing that, before
        # the earliest element the schema says must follow it.
        at = body.rfind(CF_ANCHOR)
        if at != -1:
            at += len(CF_ANCHOR)
        else:
            candidates = [body.find(tag) for tag in SHEET_ELEMENTS_AFTER_CF]
            found = [i for i in candidates if i != -1]
            at = min(found) if found else body.find("</worksheet>")
        body = body[:at] + "".join(legacy) + body[at:]

        if extended:
            block = (f'<extLst><ext uri="{{78C0D931-6437-407d-A8EE-F0AAD7539E65}}" '
                     f'xmlns:x14="{X14_NS}">'
                     f'<x14:conditionalFormattings>{"".join(extended)}'
                     f'</x14:conditionalFormattings></ext></extLst>')
            # `extLst` is the LAST child of worksheet, after tableParts.
            body = body.replace("</worksheet>", block + "</worksheet>")

        # `ignoredErrors` sits between `cellWatches` and `smartTags` in CT_Worksheet, which
        # in practice means after conditionalFormatting/hyperlinks and before tableParts and
        # extLst. Anchored on the first element the schema puts after it, never appended.
        ranges = text_columns.get(sheet_name, [])
        if ranges:
            ignored = "".join(
                f'<ignoredError sqref="{span}" numberStoredAsText="1" '
                f'twoDigitTextYear="1"/>' for span in ranges)
            marker = f"<ignoredErrors>{ignored}</ignoredErrors>"
            # ANCHORED ON A TOP-LEVEL ELEMENT, WHICH THE FIRST VERSION WAS NOT.
            #
            # It searched for the earliest of a list that included `<extLst`, and the
            # earliest `<extLst` in this document is the one NESTED INSIDE the data bar's
            # own `<cfRule>`. So the block landed inside cfRule — well-formed XML, valid
            # zip, and completely invisible to Excel, which went on flagging every record
            # cell. It also slipped past the order validator, because that walks the DIRECT
            # children of <worksheet> and this was never one of them.
            #
            # `<tableParts>` is unique and always a direct child, and the schema puts it
            # after ignoredErrors, so it is a safe anchor. `rfind` on extLst finds the
            # worksheet-level one (added above, last in the document) rather than a nested
            # one. Both are checked, earliest wins.
            anchors = [body.find("<tableParts"), body.rfind("<extLst"),
                       body.find("</worksheet>")]
            at = min(i for i in anchors if i != -1)
            body = body[:at] + marker + body[at:]

        rewritten[part] = body.encode("utf-8")

    # FAIL HERE RATHER THAN IN EXCEL. A sheet whose children are out of CT_Worksheet order
    # is a file Excel refuses with "we found a problem with some content", replaces the whole
    # part, and opens with the tab EMPTY — with no indication of which element was wrong.
    # This shipped once. Raising costs a build; the alternative costs the user their data.
    for part, body in rewritten.items():
        violations = sheet_order_violations(body)
        if violations:
            raise RuntimeError(
                f"{part} would be rejected by Excel: element(s) out of CT_Worksheet order "
                f"{violations}. The data-bar injection put a block in the wrong place.")

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            # Directory entries have no content; `read()` on one raises. openpyxl does not
            # normally write them, but a rewrite must not depend on that.
            if item.is_dir():
                target.writestr(item, b"")
                continue
            data = rewritten.get(item.filename, source.read(item.filename))
            target.writestr(item, data)
    source.close()
    return out.getvalue()


def build(season: int, week: Optional[int], season_type: str = "regular",
          conference: Optional[str] = None, division: str = "fbs") -> tuple:
    """Build the workbook for one scope. Returns (bytes, index_rows, omitted).

    Returns the omissions alongside the bytes rather than logging them, because the caller
    has to show the user what is NOT in the file they are about to download.

    `division` was missing from this signature until R-184, while `GameScope` had carried it
    since R-165. A user who filtered Schedule to FBS and walked to Excel Export downloaded
    every Division II fixture in the season — 313 rows for a week that held 83.
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    generated = datetime.now(timezone.utc)
    site_host = site_base_url()
    link_scope = {"season": season, "week": week, "season_type": season_type,
                  "conference": conference, "division": division}
    book = Workbook()
    book.remove(book.active)              # the default sheet, before Index is written

    global CENTRED
    CENTRED = Alignment(horizontal="center", vertical="center")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4858")
    note_font = Font(italic=True, size=9, color="555555")

    index_rows, omitted = [], []
    # The frame the LEGEND is phrased from: the first shipped sheet that carries the
    # threshold columns. One workbook, one set of numbers — they are constant across
    # srv_game, so any row of it answers, and a sheet without them falls back.
    legend_frame = None

    for sheet in SHEETS:
        # AC-15.5: a sheet with nothing to say is omitted and named. The view name goes in
        # the note so the omission points at an object rather than at a mood.
        read = read_sheet(sheet, season, week, season_type, conference, division)
        df = read.frame
        if df is None:
            omitted.append((sheet.name, read.omission))
            continue

        tab = book.create_sheet(sheet.name)

        # AC-15.3 / AC-15.4, written before anything else so no sheet can exist without it.
        # The block is however many lines this sheet needs; the header address follows from
        # it (R-181) rather than being a constant every sheet has to agree with.
        notes = [CFBD_CREDIT] + ([MODEL_DISCLAIMER] if sheet.sheet_disclaimer else [])
        for offset, text in enumerate(notes):
            tab.cell(ROW_CREDIT + offset, 1, text).font = note_font
        row_header = header_row(len(notes))
        row_first_data = first_data_row(len(notes))

        for index, (field, label) in enumerate(sheet.columns, start=1):
            cell = tab.cell(row_header, index, label)
            # R-258. One fill per CATEGORY, so a reader scanning 131 headers left to right
            # can see where offence stops and defence starts without reading a word. A sheet
            # with no categories declared gets the single navy, unchanged.
            cell.font = header_font
            cell.fill = PatternFill("solid",
                                    fgColor=sheet.header_fill_for(field, "2F4858"))
            # Marc: top and centre. Vertical TOP matters more than it sounds — with a wrapped
            # header the row is as tall as the worst label, so a centred one-line header
            # floats in the middle of a four-line row and no two headers share a baseline.
            cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)

        links = sheet.hyperlinks(site_host, **link_scope) if site_host else {}
        # Measured on the RENDERED values, not the raw ones: `is_indoors` is a boolean in the
        # frame and reads "Yes"/"No" in the sheet, and it is the sheet the reader looks at.
        rendered = pd.DataFrame(
            {field: [sheet.value_for(field, record) for _, record in df.iterrows()]
             for field in sheet.fields})
        by_name = {field for field, label in sheet.columns
                   if label in ALWAYS_CENTRED_LABELS}
        centred_fields = (sheet.centred | by_name
                          | _low_cardinality_text(rendered, sheet.fields))
        for offset, (_, record) in enumerate(df.iterrows()):
            for index, field in enumerate(sheet.fields, start=1):
                builder = links.get(field)
                cell_alignment = CENTRED if field in centred_fields else None
                if field in sheet.url_value_fields:
                    # A WORD, not the url, with the url behind it (Marc, round 3). With no
                    # origin resolved the cell is blank rather than a label that goes
                    # nowhere — a link-shaped thing that does nothing is worse than absence.
                    value = URL_CELL_LABEL if (builder and builder(record)) else None
                else:
                    value = _clean(sheet.value_for(field, record))
                cell = tab.cell(row_first_data + offset, index, value)
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = number_format(
                        field, sheet.decimals, sheet.integer_fields,
                        sheet.site_precision)
                elif isinstance(value, datetime):
                    cell.number_format = date_format(field)
                # R-183. The link goes ON the cell whose text is already the label, never as
                # a HYPERLINK() formula: the value stays real text, so the column still
                # sorts, filters and copies as a name.
                if builder is not None and value is not None:
                    target = builder(record)
                    if target:
                        cell.hyperlink = target
                        cell.style = "Hyperlink"
                if cell_alignment is not None:
                    cell.alignment = cell_alignment
                if field in sheet.open_mark_fields and value in MARK_COLOURS:
                    # The open marks carry their own colour, so "did not happen" is legible
                    # at a glance rather than only on close inspection of the shape.
                    cell.font = Font(name=MARK_FONT_NAME, color=MARK_COLOURS[value],
                                     size=MARK_FONT_SIZE)
                elif field in sheet.display and field in sheet.centred:
                    cell.font = Font(name=MARK_FONT_NAME, size=MARK_FONT_SIZE)

        last_row = row_first_data + len(df) - 1
        last_column = get_column_letter(len(sheet.columns))
        # AC-15.12: native Excel affordances, so the file is workable rather than readable.
        tab.freeze_panes = (f"{get_column_letter(sheet.freeze_column())}{row_first_data}")
        # R-182 TRAP 1: NO `tab.auto_filter.ref` HERE. A Table brings its own filter buttons,
        # and openpyxl documents that a table must not overlap the worksheet's autofilter —
        # an overlap is exactly the "we found a problem with some content" repair prompt that
        # AC-15.6 forbids. The Table below supplies the affordance the autofilter used to.
        table = ExcelTable(displayName=table_name(sheet.name),
                           ref=f"A{row_header}:{last_column}{last_row}")
        # TRAP 3: a TableStyleInfo's header band would override the manual navy fill, and the
        # file would then depend on which Excel applied last. Marc's file should look like it
        # came from cfdb, so the manual header stays and the style contributes banding only —
        # with row stripes off, because banding plus a navy header is two header treatments.
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False)
        tab.add_table(table)
        # R-257 / R-261. BAND EVERY OTHER GAME, painted onto the cells.
        #
        # Read from `game_no`, which is real data on the row, so the parity is the game's and
        # not the row number's — two rows per game means a row-parity band would shade the
        # away row of every game and leave the home row bare.
        #
        # Applied to the whole used width including the columns whose own fill is set
        # elsewhere; nothing on the body of this sheet sets one, so there is nothing to
        # overwrite. Rows whose game number is odd are left alone rather than painted white,
        # so the sheet still prints without a full-bleed background.
        if sheet.band_field and sheet.band_field in sheet.fields:
            band_index = sheet.fields.index(sheet.band_field) + 1
            band = PatternFill("solid", fgColor=SCORES_BAND_FILL)
            for offset in range(len(df)):
                row = row_first_data + offset
                number = tab.cell(row, band_index).value
                if isinstance(number, (int, float)) and int(number) % 2 == 0:
                    for index in range(1, len(sheet.columns) + 1):
                        tab.cell(row, index).fill = band
        # WIDTHS FROM THE DATA, AND ONLY FROM THE DATA (R-217).
        #
        # The previous version seeded each width with `len(label)`, and its own comment said
        # the header rows were excluded from the measurement. They were — the 120-character
        # credit line in A1 is not measured. BUT THE COLUMN LABEL WAS, so "Conference game"
        # set a 15-wide column over data that is three characters, and Marc found four of
        # them: Home rank, Current week, Best rank, Conference game. Every one is a long
        # header over short data.
        #
        # R-057 confirmed these widths and was right at the time: the header was not wrapping
        # then, so the label had to fit on one line. It wraps now, so it does not.
        #
        # The floor is the trade. A one-character column is unreadable even when its data is
        # one character, so 6 is the narrowest anything gets.
        widths = {}
        for index, (field, label) in enumerate(sheet.columns, start=1):
            letter = get_column_letter(index)
            longest = 0
            for offset in range(len(df)):
                value = tab.cell(row_first_data + offset, index).value
                if value is None:
                    continue
                if isinstance(value, datetime):
                    rendered = rendered_date_width(field)
                elif isinstance(value, float):
                    if is_plain_integer(field) or field in sheet.integer_fields:
                        # R-216 made these render as "0", so a rank occupies as many
                        # characters as it has digits. The blanket 8-character floor below
                        # was written when every float carried decimals, and it kept a
                        # two-digit rank column ten wide — which is the same defect R-217
                        # is about, arriving from the other direction.
                        rendered = len(f"{int(value)}")
                    else:
                        # A float renders at its FORMAT width, not its repr:
                        # 0.07894736842105 occupies four characters once the number format
                        # is applied.
                        rendered = len(
                            number_format(field, sheet.decimals, sheet.integer_fields,
                                          sheet.site_precision)
                            .replace("#,##", "").replace(";", ""))
                        rendered = max(rendered, 8)
                else:
                    rendered = len(str(value))
                longest = max(longest, rendered)
            ceiling = 60 if field.endswith("description") or field == "attribution" else 28
            longest_word = max((len(w) for w in str(label).split()), default=0)
            floor = max(MIN_COLUMN_WIDTH, min(longest_word, HEADER_WORD_CAP))
            # A HAND-SET WIDTH WINS, BUT NOT BELOW THE HEADER'S OWN LONGEST WORD.
            #
            # Marc set Season to 5.6 and then reported it "isn't wide enough... doesn't look
            # like it was touched". It WAS touched — it was written at exactly 5.6 — and that
            # is narrower than the word "Season", which is six characters. Excel wraps it
            # mid-word, so the column reads "Seaso / n" and looks untouched rather than
            # narrow. A measurement taken from a column whose header already fits cannot
            # anticipate that.
            #
            # So an override is honoured down to the point where its own header stops being
            # readable, and no further. Only two of the six are affected, and by a fraction.
            width = effective_width(label, min(max(longest + 2, floor), ceiling))
            widths[label] = width
            tab.column_dimensions[letter].width = width
            span = f"{letter}{row_first_data}:{letter}{last_row}"
            if any(isinstance(tab.cell(row_first_data + offset, index).value, str)
                   for offset in range(len(df))):
                _mark_text_column(tab, span)
            if field in DATA_BAR_FIELDS:
                _add_data_bar(tab, span)
            elif field in COLOUR_SCALE_FIELDS or field in REVERSED_SCALE_FIELDS:
                # Red at the end a reader should NOTICE. For a quantity with a direction that
                # is the minimum; for an away-minus-home margin it is the positive end, where
                # the away team won.
                low, high = ("F8696B", "63BE7B")
                if field in REVERSED_SCALE_FIELDS:
                    low, high = high, low
                tab.conditional_formatting.add(span, ColorScaleRule(
                    start_type="min", start_color=low,
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color=high))
            elif field in FLAG_FIELDS:
                # R-218 TRAP. Once a boolean column renders "Yes", a rule comparing to TRUE
                # matches nothing — NO ERROR, NO WARNING, the highlight simply disappears.
                # The formula follows what the cell actually holds.
                formula = '"Yes"' if field in sheet.display else "TRUE"
                tab.conditional_formatting.add(span, CellIsRule(
                    operator="equal", formula=[formula],
                    fill=PatternFill("solid", fgColor="D8EFD3")))

        # THE HEADER ROW HEIGHT IS COMPUTED, NOT 38 (R-217).
        #
        # Marc's macro uses 38pt, which is three lines and a good default — but it is a
        # default that happens to fit today's headers, and the next long one would silently
        # clip. Measuring cannot clip: for each label, at that column's FINAL width, work out
        # how many lines it needs and take the worst.
        tab.row_dimensions[row_header].height = max(
            MIN_HEADER_HEIGHT,
            _header_height([(label, widths[label]) for _, label in sheet.columns]))
        # The note carries what the Index has to be able to say about THIS sheet: that its
        # rows were cut, and — separately — that the Division filter could not reach it.
        note = sheet.note
        if not sheet.division_scoped and division != "all":
            extra = (f"The Division filter ({division.upper()}) does not apply to "
                     f"{sheet.view}, which has no game classification to narrow on.")
            note = f"{note} {extra}".strip()
        if legend_frame is None and metrics.BIG_COLUMN in df.columns:
            legend_frame = df
        index_rows.append(IndexRow(sheet.name, sheet.view, read.rows,
                                   read.rows_in_scope, note))

    # THE FRAME GOES WITH IT, so the legend's thresholds come from the rows this
    # workbook actually contains rather than from a constant (R-224).
    _write_index(book, season, week, season_type, conference, division, generated,
                 index_rows, omitted, header_font, header_fill, note_font,
                 site_host, legend_frame)

    buffer = io.BytesIO()
    book.save(buffer)
    wanted = {name: getattr(book[name], "_cfdb_data_bars", []) for name in book.sheetnames}
    text = {name: getattr(book[name], "_cfdb_text_columns", []) for name in book.sheetnames}
    payload = _inject_data_bars(buffer.getvalue(),
                                {k: v for k, v in wanted.items() if v},
                                {k: v for k, v in text.items() if v})
    return payload, index_rows, omitted


def _write_index(book, season, week, season_type, conference, division, generated,
                 index_rows, omitted, header_font, header_fill, note_font,
                 site_host=None, df=None) -> None:
    """AC-15.9. The index is the sheet that makes the workbook auditable a month later.

    It states when the file was made, exactly what scope it covers, how many rows each tab
    holds, which model version produced any predicted column, and — the part that matters
    most — what is NOT here and why.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

    tab = book.create_sheet("Index", 0)
    notes = [CFBD_CREDIT, MODEL_DISCLAIMER]
    for offset, text in enumerate(notes):
        tab.cell(ROW_CREDIT + offset, 1, text).font = note_font

    row = header_row(len(notes))
    tab.cell(row, 1, "cfdb export").font = header_font
    tab.cell(row, 1).fill = header_fill
    row += 2

    # The model version is read from the data rather than stated, so it cannot describe a
    # different run than the one in the file.
    try:
        versions = query("""select distinct model_version, model_name
                            from srv_model_performance
                            where segment_type = 'overall' limit 20""")
        model_version = ", ".join(
            f"{r.model_name} {r.model_version}" for r in versions.itertuples()) or "none"
    except Exception:                                              # noqa: BLE001
        model_version = "unavailable"

    for label, value in (
        ("Generated (UTC)", generated.strftime("%Y-%m-%d %H:%M:%S")),
        ("Scope", describe_scope(season, week, season_type, conference, division)),
        ("Model version(s)", model_version),
        ("Source", "cfdb serving layer; every sheet is one serving view"),
    ):
        tab.cell(row, 1, label).font = header_font
        tab.cell(row, 2, value)
        row += 1

    row += 1
    # "Rows in scope" is its own COLUMN rather than a footnote, so a reader who sorts or
    # filters this block still has the pair side by side. A truncation recorded once in prose
    # at the bottom is a truncation that survives exactly until someone sorts the table.
    for index, label in enumerate(
            ("Sheet", "Serving view", "Rows written", "Rows in scope", "Note"), start=1):
        cell = tab.cell(row, index, label)
        cell.font, cell.fill = header_font, header_fill
    row += 1
    inventory_header = row - 1
    for entry in index_rows:
        # THE INDEX IS A WAY BACK IN, not just a manifest (Marc, round 3).
        #
        # The sheet name links to the page it came from, CARRYING THE SAME FILTERS the file
        # was built with — so someone reading the workbook in November lands on the week it
        # covers rather than on this week. The serving view links to its own entry in the
        # data dictionary, using the site's existing `?table=` convention rather than a
        # second spelling of it.
        name_cell = tab.cell(row, 1, entry.name)
        view_cell = tab.cell(row, 2, entry.view)
        if site_host:
            page = PAGE_FOR_SHEET.get(entry.name)
            if page:
                carried = _scoped_query(season=season, week=week, season_type=season_type,
                                        conference=conference, division=division)
                name_cell.hyperlink = f"{site_host}/{page}?{carried}" if carried \
                    else f"{site_host}/{page}"
                name_cell.style = "Hyperlink"
            view_cell.hyperlink = f"{site_host}/dictionary?table={entry.view}"
            view_cell.style = "Hyperlink"
        tab.cell(row, 3, entry.rows).number_format = "#,##0"
        tab.cell(row, 4, entry.rows_in_scope).number_format = "#,##0"
        tab.cell(row, 5, entry.note)
        row += 1
    # R-182. The inventory is a DATASET — sheet, view, counts, note — so it is a Table and
    # the reader can sort and filter it. The metadata block above it is label/value pairs and
    # stays plain; making that a Table would claim it is data when it is a caption.
    if index_rows:
        index_table = ExcelTable(displayName=table_name("Index"),
                                 ref=f"A{inventory_header}:E{row - 1}")
        index_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False)
        tab.add_table(index_table)

    # R-196. The truncation line. A silently short workbook is the worst failure this file
    # can have, because nothing about it looks wrong — it is a plausible number of rows in a
    # correctly formatted sheet, and the reader has no way to know what is not there.
    cut = [e for e in index_rows if e.truncated]
    row += 1
    if cut:
        tab.cell(row, 1, "⚠ Truncated").font = header_font
        tab.cell(row, 2, f"The row cap is {ROW_CAP:,}. These sheets hit it:")
        row += 1
        for entry in cut:
            tab.cell(row, 1, entry.name)
            tab.cell(row, 2, f"{entry.rows:,} of {entry.rows_in_scope:,} rows written; "
                             f"{entry.rows_in_scope - entry.rows:,} not in this file. "
                             f"Narrow the filters on the Excel Export page to get all of them.")
            row += 1
    else:
        tab.cell(row, 1, "Complete").font = header_font
        tab.cell(row, 2, f"No sheet hit the {ROW_CAP:,}-row cap; every sheet holds every "
                         f"row in scope.")
        row += 1

    row += 1
    tab.cell(row, 1, "Not included in this workbook").font = header_font
    row += 1
    # AC-15.5's omissions (a view with no rows in this scope) and the sheets that are simply
    # not built yet are DIFFERENT ANSWERS to the same question, and a reader who wanted the
    # Odds sheet needs to know which one they got.
    for name, reason in omitted:
        tab.cell(row, 1, name)
        tab.cell(row, 2, reason)
        row += 1
    for sheet in PENDING_SHEETS:
        tab.cell(row, 1, sheet.name)
        tab.cell(row, 2, f"{sheet.view} — {PENDING_REASON}")
        row += 1
    if not omitted and not PENDING_SHEETS:
        tab.cell(row, 1, "Nothing — every sheet had rows in this scope.")
        row += 1

    # R-183. Whether the links are there at all, and the one thing that makes a working
    # link look broken: Cloudflare Access asks the reader to sign in.
    row += 1
    tab.cell(row, 1, "Links").font = header_font
    if site_host:
        tab.cell(row, 2, f"Team and matchup cells link back to {site_host}, carrying this "
                         f"workbook's scope. The site sits behind Cloudflare Access, so a "
                         f"link will ask you to sign in — that is not a broken link.")
    else:
        tab.cell(row, 2, "This workbook has no hyperlinks: the site's address "
                         "(CFDB_SITE_HOST) was not set where it was built. Links are "
                         "omitted rather than guessed, because a wrong link is worse than "
                         "no link.")
    row += 1

    # R-219. THE LEGEND IS NOT OPTIONAL.
    #
    # R-026's icon-only exception on the SITE is defensible because R-102's legend explains
    # it once — that is the stated reason in the code. A workbook travels further than a page
    # does and has no tooltip at all, so the same exception needs the same support or the
    # three verdict columns are undecodable symbols.
    if any(s.display for s in SHEETS):
        row += 1
        tab.cell(row, 1, "Legend").font = header_font
        tab.cell(row, 2, "Upset level, Winner covered and O/U result use the same marks as "
                         "the site. Favorite covered is a word, not a mark, and can also "
                         "read \"No favorite\".")
        row += 1
        for column, mark, meaning in mark_legend(df):
            tab.cell(row, 1, column)
            glyph = tab.cell(row, 2, mark)
            # The legend's glyph must be the SAME SIZE AND COLOUR as the one in the sheet, or
            # it is a picture of a different mark. 12pt because at 11 the difference between
            # ○ and ● is a couple of pixels of ink.
            glyph.font = Font(name=MARK_FONT_NAME, bold=True, size=MARK_FONT_SIZE,
                              color=MARK_COLOURS.get(mark))
            glyph.alignment = Alignment(horizontal="center")
            tab.cell(row, 3, meaning)
            row += 1
        row += 1
        tab.cell(row, 1, "Yes / No columns").font = header_font
        tab.cell(row, 2, 'Boolean columns read "Yes" and "No" rather than TRUE and FALSE, '
                         'which filters and reads better. The cost: they are TEXT, so a '
                         'formula needs =SUM(--(range="Yes")) rather than '
                         '=COUNTIF(range,TRUE).')
        row += 1

    # R-214, consequence 1. `Spread open` and `O/U open` are gone, and they were the
    # disambiguator: a blank delta means two different things and nothing else in the file
    # tells them apart. No sentinel is invented; the reader is told.
    # A BARE COLOUR IS AS UNDECODABLE AS A BARE GLYPH. The marks get a legend for exactly
    # this reason; the bar earns one on the same argument.
    row += 1
    tab.cell(row, 1, "Margin bar").font = header_font
    tab.cell(row, 2, "Margin (away−home) is drawn as a bar from a centre line. It grows LEFT "
                     "in blue when the home team won — 64% of games — and RIGHT in red when "
                     "the away team did, which is the result worth noticing. The colours are "
                     "the direction, not a judgement about the game.")
    row += 1

    row += 1
    tab.cell(row, 1, "Blank Δ Spread / Δ O/U").font = header_font
    tab.cell(row, 2, "A blank delta means EITHER the line did not move OR cfdb holds no "
                     "opening line for that game. The two are not distinguishable in this "
                     "workbook — the opening-line columns are not included. The Matchup URL "
                     "shows which it is.")
    row += 1

    # R-214, consequence 2. The verdicts are still correct; they are no longer checkable
    # from the file alone, and saying so is cheaper than a reader concluding they are wrong.
    row += 1
    tab.cell(row, 1, "Upset / covered / O-U result").font = header_font
    tab.cell(row, 2, "These are judged against the CLOSING line, which is not a column in "
                     "this workbook. The verdicts are correct but cannot be checked against "
                     "a number here; follow the Matchup URL to see the line each one used.")
    row += 1

    # A blank cell means cfdb HOLDS NOTHING, not zero. Said once, because the alternative is
    # writing a dash into the cell, and a dash is text: it breaks COUNTBLANK, sorts ahead of
    # every number and cannot be filtered on "Blanks".
    row += 1
    tab.cell(row, 1, "Blank cells").font = header_font
    tab.cell(row, 2, "A blank cell means cfdb holds no value for it — not zero. The site "
                     "renders the same absence as a dash; a dash in a spreadsheet is text "
                     "and would break sorting and COUNTBLANK.")
    row += 2

    tab.cell(row, 1, "Scope is bounded by the filters on the Excel Export page. cfdb does "
                     "not offer a full-corpus or raw-layer export.").font = note_font

    for index, width in enumerate((26, 34, 14, 14, 60), start=1):
        tab.column_dimensions[get_column_letter(index)].width = width
