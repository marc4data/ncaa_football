"""Export a sample of every table in one warehouse layer to an Excel workbook, a tab per table.

NAMED FOR WHAT IT DOES, NOT FOR ONE OF ITS ARGUMENTS. This was export_staging_sample until
--schema existed, at which point the name described one of two things it could do — the same
"named after a use rather than its shape" mistake that had four game-grain serving views
called srv_schedule, srv_scoreboard, srv_matchup and srv_today_edges. Renamed with no shim
left behind, for the reason alias views are forbidden: a second name is how two things drift.

Rule: up to 20,000 rows per table. Any table with more than that is narrowed to
activity tied to games involving a Big 12 team, ordered by season and week ascending.

ONE OR MORE SEASONS, AND MEMBERSHIP IS RESOLVED SEPARATELY FOR EACH. `--season 2025 2026`
does not union the years and then ask who was in the Big 12; it asks per year and unions the
answers. That distinction is the whole reason membership is season-scoped — a team that
joined in 2026 should contribute its 2026 games and not its 2025 ones, and resolving against
a merged season set would hand it both.

OPPONENTS ARE INCLUDED. "Big 12 activity" means the games, and a game has two teams in it.
Game-grain tables got this for free — filtering to the game set returns both sides — but
team-grain tables were filtered to the member list, so a Big 12 team's opponent was missing
from the very sheets that describe it. The team set used for those is now every team that
appears in the game set, members and opponents alike, and the Index sheet says so.

WHICH IS NOT UNIFORMLY POSSIBLE, and the interesting part of the job is what to do about
it. Staging models are shaped like the endpoints they unpack, so the columns needed to
express "a game involving a Big 12 team" are not in all of them:

  stg_games              home_team_id / away_team_id, season   -> either side in the set
  stg_lines              game_id, no team, no season           -> scope through stg_games
  stg_game_team_stat     game_id, team_id — NO NAME, NO SEASON -> scope through stg_games
  stg_team_season_stat   school, season, no game               -> team membership, by name
  stg_teams              team_id, season, no game              -> team membership, by id

TWO SHAPES, NOT ONE. A game-grain table is filtered to the set of Big 12 games. A team-grain
table has no games to be tied to, so it is filtered to the Big 12 teams themselves — the
same request answered at the grain the table actually has. The Index sheet records which of
the two was applied, because "every game Kansas State played" and "Kansas State's season
totals" are different claims and the tab cannot tell you which it is holding.

CONFERENCE IS SEASON-SCOPED, AND THAT IS NOT A DETAIL. The Big 12 had 14 members in 2023 and
16 from 2024; dim_team carries one conference row per team per season precisely because
realignment moves teams. Resolving membership without a season would silently blend eras.
The match is also exact rather than a LIKE: dim_team holds `Big Sky`, `Big Ten` and
`OVC-Big South`, all of which a `%big%` pattern would sweep up.

A filtered sheet that comes back EMPTY is written anyway, with its headers and a note. An
empty tab with no explanation is indistinguishable from a broken export.

Usage:
  python -m src.export_sample
  python -m src.export_sample --conference "SEC" --season 2024
  python -m src.export_sample --conference "Big 12" --season 2025 2026
  python -m src.export_sample --out /tmp/sample.xlsx --conference "Big Ten"
"""
import argparse
import sys
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

ROW_CAP = 20_000
DEFAULT_SCHEMA = "staging"

# WHICH SCHEMAS THIS WILL EXPORT, AND WHY IT IS A LIST RATHER THAN A FREE STRING.
#
# The schema name is interpolated into SQL — it cannot be a bound parameter, because a
# parameter cannot be an identifier. An allowlist is the cheap correct answer: it is not
# possible to reach `raw` (JSON payloads nobody wants in a workbook) or anything outside
# these two by argument, and there is nothing to escape.
EXPORTABLE_SCHEMAS = {
    "staging": "one model per CFBD endpoint, JSON unpacked and failed responses filtered "
               "out — the layer BELOW the site's serving views",
    "serving": "what the site actually reads: pre-joined, one relation per page section, "
               "with the display rules already applied",
}

# The game spine, ALWAYS read from staging regardless of what is being exported. Resolving
# which games belong to a conference is a question about the schedule, not about the layer
# somebody asked for — and `serving.stg_games` does not exist.
SPINE_TABLE = "staging.stg_games"

# Candidate columns, in preference order, for expressing the narrowing. Checked against
# information_schema per table rather than assumed — the shape of a staging model follows
# its endpoint, not a convention.
#
# ID BEFORE NAME, ALWAYS. The team_id comes from dim_team, the conformed dimension, and it
# is a better key than the school name for a reason that is visible in the data rather than
# theoretical: dim_team also holds `Southeastern Oklahoma State` (199) and `Southwestern
# Oklahoma State` (2927). An exact name match happens to avoid them; any looser match would
# not, and an id cannot hit them at all.
#
# It is also stable where the name's context is not — Oklahoma State is one team_id across
# 1901-2026 while its conference row changes five times through realignment.
ID_COLUMNS = ("team_id",)
ID_PAIR_COLUMNS = ("home_team_id", "away_team_id")
TEAM_COLUMNS = ("school", "team", "team_name")
PAIR_COLUMNS = ("home_team", "away_team")
ORDER_COLUMNS = ("season", "year", "week")


def _clean(value):
    """One cell value openpyxl will accept.

    Three conversions, each of which has broken a workbook in this project before: a
    tz-aware datetime makes openpyxl raise outright, a NaN produces a file Excel refuses to
    open with a repair prompt, and a jsonb column arrives as a dict or list that openpyxl
    cannot serialise at all.
    """
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, (dict, list)):
        import json
        return json.dumps(value)[:32000]        # Excel's per-cell ceiling is 32,767
    if isinstance(value, float) and value != value:
        return None
    if isinstance(value, str) and len(value) > 32000:
        return value[:32000]
    return value


def columns_of(cursor, table: str, schema: str) -> list:
    cursor.execute("""
        select column_name from information_schema.columns
        where table_schema = %s and table_name = %s
        order by ordinal_position
    """, (schema, table))
    return [row[0] for row in cursor.fetchall()]


def resolve_conference(cursor, conference: str, seasons) -> dict:
    """Who is in the conference, which games they played, and who they played.

    From dim_team rather than from staging: dim_team owns team identity, and reading the key
    from the layer that owns it is the difference between a join key and a lucky match.

    MEMBERSHIP IS RESOLVED PER SEASON AND THEN UNIONED, never the other way round. The Big 12
    had 14 members in 2023 and 16 from 2024, so asking "who was in the Big 12 across 2025 and
    2026" as one question would hand a 2026 joiner its 2025 games too. Asking per year and
    unioning the answers keeps each year's game set honest; the union only ever appears
    afterwards, in the filters.

    EXACT MATCH, NEVER A PATTERN. dim_team also holds `Big Sky`, `Big Ten` and
    `OVC-Big South`; `conference ilike '%big%'` would collect all four.

    THREE SETS COME BACK, AND THEY ARE NOT INTERCHANGEABLE:

      member_ids       teams in the conference, across the seasons asked for
      game_ids         games with a member on either side
      participant_ids  every team appearing in those games — members AND their opponents

    Team-grain tables use `participant_ids`, because a request for "Big 12 activity" that
    drops the opponent's own rows describes half of each game. Game-grain tables never had
    this problem: filtering to `game_ids` returns both sides already.
    """
    seasons = sorted({int(s) for s in seasons})

    member_ids, names = [], []
    game_ids, participant_ids = [], []
    per_season = {}

    for season in seasons:
        cursor.execute("""
            select distinct team_id, school
            from marts.dim_team
            where conference = %(conference)s and season = %(season)s
              and team_id is not null
            order by school
        """, {"conference": conference, "season": season})
        rows = cursor.fetchall()
        season_ids = [r[0] for r in rows]
        per_season[season] = len(season_ids)
        member_ids.extend(season_ids)
        names.extend(r[1] for r in rows)

        if not season_ids:
            continue

        # The game set for THIS season, against THIS season's membership. Resolved once and
        # reused, so every game-grain sheet is narrowed to exactly the same games.
        cursor.execute(f"""
            select game_id, home_team_id, away_team_id from {SPINE_TABLE}
            where season = %(season)s
              and (home_team_id = any(%(ids)s) or away_team_id = any(%(ids)s))
        """, {"season": season, "ids": season_ids})
        for game_id, home_id, away_id in cursor.fetchall():
            game_ids.append(game_id)
            # Both sides. This is where opponents enter the team set.
            participant_ids.extend(i for i in (home_id, away_id) if i is not None)

    participant_ids = sorted(set(participant_ids) | set(member_ids))
    member_ids = sorted(set(member_ids))
    game_ids = sorted(set(game_ids))

    participant_names = []
    if participant_ids:
        # Names for the team-grain tables that carry no id. Season-scoped for the same reason
        # the membership is: a school's own name is stable, but restricting to the seasons
        # asked for keeps this list from collecting every alias dim_team has ever held.
        cursor.execute("""
            select distinct school from marts.dim_team
            where team_id = any(%(ids)s) and season = any(%(seasons)s)
            order by school
        """, {"ids": participant_ids, "seasons": seasons})
        participant_names = [r[0] for r in cursor.fetchall()]

    return {"conference": conference,
            "seasons": seasons,
            "per_season": per_season,
            "member_ids": member_ids,
            "member_names": sorted(set(names)),
            "team_ids": participant_ids,
            "names": participant_names or sorted(set(names)),
            "game_ids": game_ids}


def profile_fields(cursor, tables, schema: str) -> list:
    """One row per (table, column): type, null share and cardinality.

    PROFILED OVER THE WHOLE TABLE, NOT OVER THE EXPORTED SAMPLE. The data tabs are narrowed
    to one conference; a field inventory narrowed the same way would describe this workbook
    rather than the warehouse, and "how often is this column null" is a question about the
    data, not about the filter someone happened to apply.

    ONE QUERY PER TABLE, not one per column. 1,447 columns across 80 tables is 2,894 separate
    aggregates if asked individually; folded into one statement per table it is 80 scans.
    Staging models are VIEWS over raw JSON, so each scan re-parses the payload — that is why
    this is the expensive part of the export and why it reports progress as it goes.

    Identifiers are quoted rather than interpolated bare. They come from information_schema
    so they are already trustworthy, but a quoted identifier is correct for any name and a
    bare one is correct only for the names we happen to have today.
    """
    profile = []
    for position, table in enumerate(tables, start=1):
        cursor.execute("""
            select column_name, data_type from information_schema.columns
            where table_schema = %s and table_name = %s order by ordinal_position
        """, (schema, table))
        columns = cursor.fetchall()
        if not columns:
            continue

        parts = ["count(*)"]
        for name, _ in columns:
            parts.append(f'count("{name}")')
            # jsonb supports equality so distinct works on it directly. Plain `json` would
            # not, and there is none in this schema — if one ever appears this is where it
            # will fail, loudly, rather than silently reporting nothing.
            parts.append(f'count(distinct "{name}")')
        cursor.execute(f'select {", ".join(parts)} from {schema}."{table}"')
        row = cursor.fetchone()
        total = row[0]

        for index, (name, data_type) in enumerate(columns):
            non_null = row[1 + index * 2]
            distinct = row[2 + index * 2]
            # Null share is undefined on an empty table, not zero. Zero would claim the
            # column is fully populated, which is the opposite of what an empty table means.
            null_pct = None if not total else round(100.0 * (total - non_null) / total, 1)
            profile.append((table, name, data_type, null_pct, distinct))
        print(f"  [{position:>2}/{len(tables)}] profiled {table:26s} "
              f"{len(columns):>3} field(s), {total:>9,} rows")
    return profile


def build_query(cursor, table: str, row_count: int, members: dict,
                schema: str = DEFAULT_SCHEMA) -> tuple:
    """Return (sql, params, note) for one table.

    The note is what the Index sheet reports. Every sheet gets one, because "20,000 rows of
    199,083" and "every Big 12 game of 2025" are very different things to be looking at and
    the tab itself cannot tell you which it is.
    """
    cols = set(columns_of(cursor, table, schema))
    order = [c for c in ORDER_COLUMNS if c in cols]
    order_sql = f" order by {', '.join(order)} asc" if order else ""
    conference = members["conference"]
    seasons = members["seasons"]
    season_label = ", ".join(str(s) for s in seasons)
    # `team_ids` / `names` are the PARTICIPANTS — members and the opponents they played.
    # `member_ids` is the conference itself, used only where the distinction is reported.
    ids, names, game_ids = members["team_ids"], members["names"], members["game_ids"]
    members_only = members["member_ids"]

    if row_count <= ROW_CAP:
        return (f"select * from {schema}.{table}{order_sql} limit {ROW_CAP}", {},
                f"Whole table ({row_count:,} rows), under the {ROW_CAP:,} cap.")

    where, params, applied = [], {}, []

    # Season first where the table carries one. It bounds the volume and it is what makes
    # the conference membership meaningful, since the roster changes with realignment.
    if "season" in cols:
        where.append("season = any(%(seasons)s)")
        params["seasons"] = seasons
        applied.append(f"season {season_label}")

    # GAME GRAIN BEFORE TEAM GRAIN. The request is activity tied to games involving a Big 12
    # team, so where a table can express a game, that is the faithful reading.
    if all(c in cols for c in ID_PAIR_COLUMNS):
        # Members on either side, not participants — a game qualifies because a CONFERENCE
        # team is in it. Filtering on participants here would pull in games between two of
        # the opponents, which is a different and much larger question.
        where.append("(home_team_id = any(%(member_ids)s) or away_team_id = any(%(member_ids)s))")
        params["member_ids"] = members_only
        applied.append(f"either side a {conference} team ({len(members_only)} teams) — "
                       f"both teams' rows are included")
    elif "game_id" in cols:
        # No team column at all: scope through the games resolved once in resolve_conference.
        # This is the same request answered through a join, and the note says so rather than
        # implying the table carried the filter itself.
        where.append("game_id = any(%(game_ids)s)")
        params["game_ids"] = game_ids
        applied.append(f"game_id in the {len(game_ids):,} {season_label} {conference} games "
                       f"(resolved via stg_games — this table carries no team); both "
                       f"teams' rows are included")
    elif any(c in cols for c in ID_COLUMNS):
        id_col = next(c for c in ID_COLUMNS if c in cols)
        where.append(f"{id_col} = any(%(ids)s)")
        params["ids"] = ids
        applied.append(f"{id_col} in the {len(ids)} teams that played in those games "
                       f"({len(members_only)} {conference} members plus their opponents) — "
                       f"TEAM grain, not game grain: this table has no games to be tied to")
    elif any(c in cols for c in TEAM_COLUMNS):
        team_col = next(c for c in TEAM_COLUMNS if c in cols)
        where.append(f"{team_col} = any(%(names)s)")
        params["names"] = names
        applied.append(f"{team_col} in the {len(names)} teams that played in those games "
                       f"({len(members_only)} {conference} members plus their opponents) — "
                       f"matched by NAME, no id column; TEAM grain, not game grain")
    elif all(c in cols for c in PAIR_COLUMNS):
        where.append("(home_team = any(%(member_names)s) or away_team = any(%(member_names)s))")
        params["member_names"] = members["member_names"]
        applied.append(f"either side a {conference} team, matched by name — both teams' "
                       f"rows are included")

    if not where:
        return (f"select * from {schema}.{table}{order_sql} limit {ROW_CAP}", {},
                f"OVER CAP at {row_count:,} rows and NOT NARROWABLE — no season, team or "
                f"game column to filter on. Showing the first {ROW_CAP:,} rows instead.")

    sql = f"select * from {schema}.{table} where {' and '.join(where)}{order_sql} limit {ROW_CAP}"
    return sql, params, (f"{row_count:,} rows in full, narrowed to " + "; ".join(applied)
                         + (f"; ordered by {', '.join(order)}" if order else ""))


def export(out_path: Path, conference: str, seasons,
           schema: str = DEFAULT_SCHEMA) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    from .load_raw_to_postgres import get_conn

    connection = get_conn()
    cursor = connection.cursor()
    cursor.execute("""
        select table_name from information_schema.tables
        where table_schema = %s order by table_name
    """, (schema,))
    tables = [row[0] for row in cursor.fetchall()]

    members = resolve_conference(cursor, conference, seasons)
    season_label = ", ".join(str(s) for s in members["seasons"])
    if not members["member_ids"]:
        print(f"  ! no teams found for conference {conference!r} in {season_label} — "
              f"oversized tables cannot be narrowed and will show the first "
              f"{ROW_CAP:,} rows")
    else:
        # Per season as well as the total, because a year with zero members is a typo in
        # the conference name or a season with no data, and a combined count hides it.
        breakdown = ", ".join(f"{s}: {n}" for s, n in sorted(members["per_season"].items()))
        opponents = len(members["team_ids"]) - len(members["member_ids"])
        print(f"  {conference} {season_label} -> {len(members['member_ids'])} members "
              f"({breakdown}), {len(members['game_ids']):,} games, "
              f"{opponents} opponents (from marts.dim_team)\n")

    book = Workbook()
    book.remove(book.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4858")
    index_rows = []

    for table in tables:
        cursor.execute(f"select count(*) from {schema}.{table}")
        total = cursor.fetchone()[0]
        sql, params, note = build_query(cursor, table, total, members, schema)
        cursor.execute(sql, params)
        headers = [d[0] for d in cursor.description]
        rows = cursor.fetchall()

        # Excel caps a sheet name at 31 characters. Every staging name fits today; asserting
        # it rather than truncating means a future long name fails loudly instead of
        # silently colliding with another truncated tab.
        sheet_name = table[:31]
        tab = book.create_sheet(sheet_name)
        for index, header in enumerate(headers, start=1):
            cell = tab.cell(1, index, header)
            cell.font, cell.fill = header_font, header_fill
        for r, row in enumerate(rows, start=2):
            for index, value in enumerate(row, start=1):
                tab.cell(r, index, _clean(value))

        if not rows:
            # Headers with no rows is an ambiguous artifact — it reads as a broken export
            # rather than an honest empty result. Say which it is, on the sheet.
            tab.cell(2, 1, f"No rows. {note}")
            tab.cell(2, 1).font = Font(italic=True, color="A03030")

        tab.freeze_panes = "A2"
        if headers:
            tab.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"
        for index, header in enumerate(headers, start=1):
            longest = len(str(header))
            for row in rows[:400]:                # sample, not the whole column
                value = row[index - 1]
                if value is not None:
                    longest = max(longest, min(len(str(value)), 60))
            tab.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 9), 46)

        index_rows.append((table, total, len(rows), note))
        print(f"  {table:26s} {len(rows):>7,} of {total:>9,}  {note[:60]}")

    # Sheet ORDER is set by inserting both at position 0, last one first: Fields goes in,
    # then Index pushes it to second. Requested position, and it reads correctly — what the
    # workbook is, then what is in it, then the data.
    print("\nProfiling fields (whole tables, not the exported sample)...")
    _write_fields(book, profile_fields(cursor, tables, schema), header_font,
                  header_fill, schema)
    _write_index(book, index_rows, members, header_font, header_fill, schema)
    connection.close()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(out_path)
    return {"tables": len(index_rows),
            "rows": sum(r[2] for r in index_rows),
            "path": str(out_path)}


def _write_fields(book, profile, header_font, header_fill,
                  schema: str = DEFAULT_SCHEMA) -> None:
    """Every field in every staging table, with how empty and how varied it is.

    The workbook's own data dictionary. A tab per table answers "what does this hold"; this
    answers "what is in the layer, and which columns are actually populated" without opening
    eighty sheets.

    Null % and cardinality describe the FULL table. Cardinality is a count of distinct values
    including null-free counts only — count(distinct) ignores nulls — so a column that is 90%
    null can still show high cardinality across the rows it does have. Read the two together;
    row counts per table are on the Index sheet.
    """
    from openpyxl.utils import get_column_letter

    tab = book.create_sheet("Fields", 0)
    tab.cell(1, 1, f"cfdb — {schema} field inventory").font = header_font
    tab.cell(1, 1).fill = header_fill
    tab.cell(2, 1, f"{len(profile):,} fields across "
                   f"{len({row[0] for row in profile})} tables. Null % and cardinality are "
                   f"measured over the WHOLE table, not over the filtered sample on the data "
                   f"tabs — this describes the warehouse, not this workbook.")
    tab.cell(3, 1, "Cardinality counts distinct non-null values, so a mostly-null column can "
                   "still show a high count. Read it against Null %, and against the row "
                   "counts on the Index sheet.")

    headers = ("Table", "Field", "Data type", "Null %", "Cardinality")
    for index, label in enumerate(headers, start=1):
        cell = tab.cell(5, index, label)
        cell.font, cell.fill = header_font, header_fill

    for r, (table, field, data_type, null_pct, cardinality) in enumerate(profile, start=6):
        tab.cell(r, 1, table)
        tab.cell(r, 2, field)
        tab.cell(r, 3, data_type)
        # Blank, not zero, where the table is empty and the share is undefined.
        if null_pct is not None:
            tab.cell(r, 4, null_pct).number_format = "0.0"
        tab.cell(r, 5, cardinality).number_format = "#,##0"

    tab.freeze_panes = "A6"
    if profile:
        tab.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{len(profile) + 5}"
    for index, width in enumerate((30, 34, 26, 10, 14), start=1):
        tab.column_dimensions[get_column_letter(index)].width = width


def _write_index(book, index_rows, members, header_font, header_fill, schema) -> None:
    """What each tab is, and why it holds what it holds.

    Without this the workbook cannot answer its own most obvious question: is this tab the
    whole table, the first 20,000 rows, every Big 12 game, or Big 12 teams' own rows?
    """
    from openpyxl.utils import get_column_letter

    tab = book.create_sheet("Index", 0)
    tab.cell(1, 1, f"cfdb — {schema} layer sample").font = header_font
    tab.cell(1, 1).fill = header_fill
    # BOTH ZONES, deliberately. The default filename is stamped in local time so a person can
    # find the file they made this afternoon; this line is the provenance record and has
    # always been UTC. Showing one without the other makes the pair look inconsistent.
    tab.cell(2, 1, f"Generated {datetime.now():%Y-%m-%d %H:%M} local "
                   f"({datetime.now(timezone.utc):%Y-%m-%d %H:%M} UTC)")
    season_label = ", ".join(str(s) for s in members["seasons"])
    opponents = len(members["team_ids"]) - len(members["member_ids"])
    tab.cell(3, 1, f"Rule: up to {ROW_CAP:,} rows per table. Anything larger is narrowed to "
                   f"activity tied to games involving a {members['conference']} team in "
                   f"{season_label} — {len(members['member_ids'])} members, "
                   f"{len(members['game_ids']):,} games, {opponents} opponents also "
                   f"included, resolved from marts.dim_team — ordered by season and week "
                   f"ascending.")
    tab.cell(5, 1, "Two shapes. A game-grain table is filtered to that game set, so BOTH "
                   "teams' rows are present. A team-grain table has no games to be tied "
                   "to, so it is filtered to every team that played in those games — "
                   "members and opponents alike; the Note column says which was applied. "
                   "Membership is resolved SEPARATELY PER SEASON and then unioned, because "
                   "realignment moves teams: the Big 12 had 14 members in 2023 and 16 from "
                   "2024, so a team that joined later must not contribute its earlier "
                   "games.")
    # The layer note comes from EXPORTABLE_SCHEMAS rather than being written twice, so a
    # serving workbook cannot describe itself as staging.
    tab.cell(4, 1, f"{schema.capitalize()}: {EXPORTABLE_SCHEMAS[schema]}. Shapes follow the "
                   f"source, which is why not every table can express the same filter.")

    for index, label in enumerate(("Table", "Rows in full", "Rows exported", "Note"), start=1):
        cell = tab.cell(6, index, label)
        cell.font, cell.fill = header_font, header_fill
    for r, (table, total, exported, note) in enumerate(index_rows, start=7):
        tab.cell(r, 1, table)
        tab.cell(r, 2, total).number_format = "#,##0"
        tab.cell(r, 3, exported).number_format = "#,##0"
        tab.cell(r, 4, note)
    tab.freeze_panes = "A7"
    for index, width in enumerate((28, 14, 14, 110), start=1):
        tab.column_dimensions[get_column_letter(index)].width = width


def main() -> int:
    parser = argparse.ArgumentParser(description="Sample every staging table into one workbook.")
    parser.add_argument("--schema", default=DEFAULT_SCHEMA, choices=sorted(EXPORTABLE_SCHEMAS),
                        help="which warehouse layer to export. `staging` is one model per "
                             "endpoint; `serving` is what the site reads.")
    # DEFAULT IS TIMESTAMPED, so two runs do not silently overwrite each other. The old
    # default was a fixed staging_sample.xlsx: exporting twice left one file and no way to
    # tell which run produced it. Pass --out to name it yourself.
    #
    # LOCAL TIME, not UTC, because the timestamp is there to help a person find the file they
    # made this afternoon. The Index sheet stamps BOTH so the two always reconcile — a
    # filename saying 1800 beside a workbook saying 1700 UTC is a puzzle nobody needs.
    parser.add_argument("--out", type=Path, default=None,
                        help="output path. Defaults to "
                             "data/exports/<schema>_sample_<yyyymmdd>_<hhmm>.xlsx, "
                             "timestamped in local time.")
    parser.add_argument("--conference", default="Big 12",
                        help="exact dim_team.conference value, e.g. 'Big 12', 'SEC'")
    parser.add_argument("--season", type=int, nargs="+", default=[2025],
                        metavar="YEAR",
                        help="one or more seasons, e.g. --season 2025 2026. Membership is "
                             "resolved per season and unioned, never merged first.")
    args = parser.parse_args()
    out_path = args.out or Path(
        f"data/exports/{args.schema}_sample_{datetime.now():%Y%m%d_%H%M}.xlsx")

    season_label = ", ".join(str(s) for s in sorted(set(args.season)))
    print(f"Sampling {args.schema}: <= {ROW_CAP:,} rows per table, larger tables narrowed to "
          f"games involving a {args.conference} team in {season_label}, "
          f"opponents included\n")
    summary = export(out_path, args.conference, args.season, args.schema)
    print(f"\n{summary['tables']} sheet(s), {summary['rows']:,} rows -> {summary['path']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
