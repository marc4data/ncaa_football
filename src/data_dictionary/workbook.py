"""Render the data dictionary as an xlsx workbook.

Sheet order is deliberate: README, then Gaps early enough to be seen, then the reference
sheets. The Fields sheet colours its Provenance column so a reader can tell at a glance how
much of the document is CFBD's word and how much is ours — which, given CFBD describes four
of its 1,017 response fields, is the single most important thing about this artifact.
"""
import collections
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import definitions as dfn

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
SECTION_FONT = Font(name=FONT, bold=True, size=11, color="1F3864")
LABEL_FONT = Font(name=FONT, bold=True, size=10)

WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PROVENANCE_FILL = {
    dfn.GLOSSARY_PROV: PatternFill("solid", fgColor="D9E1F2"),
    dfn.DOCS_PROV: PatternFill("solid", fgColor="E2EFDA"),
    dfn.SPEC_PROV: PatternFill("solid", fgColor="EDEDED"),
    dfn.OBSERVED_PROV: PatternFill("solid", fgColor="DDEBF7"),
    dfn.INFERRED_PROV: PatternFill("solid", fgColor="FFF2CC"),
}

VOCABULARY_NOTES = {
    "SeasonType":
        "QUERY vocabulary. Includes 'both', a filter value that is never stored in a record. "
        "Compare with SeasonTypeDB before typing any column.",
    "SeasonTypeDB":
        "STORED vocabulary. Includes 'preseason', which SeasonType cannot filter on, and omits "
        "'both'. Writing a request parameter into a column typed by this vocabulary can produce "
        "an invalid value.",
    "DivisionClassification":
        "Team/game level. Four values. Compare with ConferenceClassification.",
    "ConferenceClassification":
        "Conference level. Adds 'ii/iii', which has no team-level counterpart — team and "
        "conference classification will NOT join one-to-one.",
}


def _write_table(ws, headers, rows, widths, wrap_columns=()):
    ws.append(headers)
    for column, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in rows:
        ws.append(row)
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in wrap_columns)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(headers)), ws.max_row)
    ws.row_dimensions[1].height = 30


def grain_of(profile: dict):
    """Return (grain, note). Grain is measured, never asserted — say so when it is shaky."""
    combos = profile.get("unique") or []
    if not combos:
        return "", ("Not determined — no combination of up to {} columns was unique in the "
                    "sampled file.".format(3))
    primary = " + ".join(combos[0])
    notes = []
    cross = (profile.get("cross_file") or {}).get("+".join(combos[0]))
    if cross and cross.get("duplicate_keys"):
        notes.append(
            "Unique within a single response file, but {} duplicate keys appear across {} "
            "de-duplicated rows from multiple files — the true grain is wider, or rows mutate "
            "between fetches. Verify before using as a key.".format(
                cross["duplicate_keys"], cross["deduped_rows"]))
    alternates = [" + ".join(c) for c in combos[1:3]]
    if alternates:
        notes.append("Also unique: " + "; ".join(alternates) + ".")
    return primary, " ".join(notes)


def build(spec, endpoints, fields, parameters, profile, registry_by_key,
          dbt_sources, gaps, out_path, captured):
    """Assemble and save the workbook. Returns the row counts written."""
    wb = Workbook()

    # ---------------------------------------------------------------- Fields
    ws = wb.active
    ws.title = "Fields"
    field_rows = []
    for row in fields:
        profiled = profile.get(row["key"], {})
        top = row["field_path"].split(".")[0].replace("[]", "")
        nested = "." in row["field_path"] or "[]" in row["field_path"]
        definition, provenance, confidence = dfn.define(row["field_path"], row["description"])
        observed = (profiled.get("keys") or {}).get(top, {})
        if nested and top in (profiled.get("nested") or []):
            null_pct, distinct, samples = "", "", "(nested — profiled at parent level only)"
        elif observed:
            null_pct = observed.get("null_pct", "")
            distinct = observed.get("distinct", "")
            samples = "; ".join(str(s) for s in observed.get("samples") or [])[:180]
        else:
            null_pct, distinct, samples = "", "", ""
        field_rows.append([
            row["endpoint"], row["key"], row["entity"], row["field_path"], row["type"],
            row["nullable"], row["required"], row["enum_values"],
            definition, provenance, confidence, null_pct, distinct, samples,
            "yes" if row["key"] in dbt_sources else "no",
        ])
    _write_table(ws, [
        "Endpoint", "Raw dir", "Entity", "Field path", "Type", "Nullable (spec)",
        "Required (spec)", "Enum values", "Definition", "Provenance", "Confidence",
        "Observed null %", "Observed distinct", "Observed sample values",
        "Raw table used by dbt",
    ], field_rows, [22, 20, 22, 30, 20, 12, 12, 30, 68, 12, 11, 13, 15, 40, 12],
        wrap_columns=(8, 9, 14))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fill = PROVENANCE_FILL.get(row[9].value)
        if fill:
            row[9].fill = fill
        if row[10].value == dfn.LOW:
            row[10].fill = WARN_FILL
        elif row[10].value == dfn.MEDIUM:
            row[10].fill = NOTE_FILL

    # ---------------------------------------------------------------- Endpoints
    ws = wb.create_sheet("Endpoints")
    field_counts = collections.Counter(r["key"] for r in fields)
    endpoint_rows = []
    for endpoint in sorted(endpoints, key=lambda e: e["key"]):
        key = endpoint["key"]
        profiled = profile.get(key, {})
        grain, grain_note = grain_of(profiled)
        registered = registry_by_key.get(key)
        notes = []
        failed = profiled.get("failed_calls") or {}
        if failed:
            notes.append("FAILED CALLS: {} of {} — status {}.".format(
                sum(failed.values()), profiled.get("manifest_calls", "?"),
                ", ".join(str(s) for s in failed)))
        if grain_note:
            notes.append(grain_note)
        if profiled.get("empty_files"):
            notes.append("{} payload files are empty (status 200, data []). These are "
                         "2026-season queries made before the season started.".format(
                             profiled["empty_files"]))
        if len(profiled.get("shapes") or []) > 1:
            notes.append("Mixed payload shapes on disk: " + ", ".join(profiled["shapes"]) + ".")
        endpoint_rows.append([
            endpoint["endpoint"], key, endpoint["tag"], endpoint["entity"],
            endpoint["description"], grain, "observed" if grain else "not determined",
            registered.strategy if registered else "", registered.bucket if registered else "",
            "yes" if registered and registered.include else "no",
            profiled.get("files", ""), profiled.get("manifest_calls", ""),
            profiled.get("empty_files", ""), profiled.get("records_sampled", ""),
            "; ".join(profiled.get("param_keys") or []), field_counts.get(key, 0),
            "yes" if key in dbt_sources else "no", " | ".join(notes),
        ])
    _write_table(ws, [
        "Endpoint", "Raw dir", "Tag", "Response entity", "CFBD description (docs)",
        "Grain (observed)", "Grain basis", "Registry strategy", "Registry bucket",
        "In default sweep", "Payload files", "Manifest calls", "Empty payloads",
        "Records sampled", "Params used", "Field count", "Sourced by dbt", "Notes",
    ], endpoint_rows,
        [24, 22, 14, 24, 46, 32, 14, 15, 14, 13, 12, 13, 13, 14, 26, 11, 12, 70],
        wrap_columns=(5, 6, 18))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[17].value and "FAILED CALLS" in str(row[17].value):
            for cell in row:
                cell.fill = WARN_FILL
        elif row[16].value == "yes":
            row[1].fill = OK_FILL

    # ---------------------------------------------------------------- Vocabularies
    ws = wb.create_sheet("Vocabularies")
    usage: Dict[str, set] = collections.defaultdict(set)
    for row in fields:
        if row["type"].startswith("enum<"):
            usage[row["type"][5:-1]].add("{}.{}".format(row["endpoint"], row["field_path"]))
    vocabulary_rows = []
    for vocabulary in spec.vocabularies():
        name = vocabulary["name"]
        used = sorted(usage.get(name, ()))
        vocabulary_rows.append([
            name, len(vocabulary["values"]), "; ".join(vocabulary["values"]),
            len(used), "; ".join(used[:6]), VOCABULARY_NOTES.get(name, ""),
        ])
    _write_table(ws, ["Vocabulary", "Values", "Allowed values", "Fields using it",
                      "Example fields", "Notes"],
                 vocabulary_rows, [28, 9, 56, 14, 54, 74], wrap_columns=(3, 5, 6))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[5].value:
            for cell in row:
                cell.fill = WARN_FILL

    # ---------------------------------------------------------------- Glossary
    ws = wb.create_sheet("Glossary")
    reverse = collections.defaultdict(list)
    for leaf, term in dfn.GLOSSARY_FIELDS.items():
        reverse[term].append(leaf)
    glossary_rows = []
    for term, definition in sorted(dfn.GLOSSARY.items()):
        mapped = sorted(reverse.get(term, []))
        glossary_rows.append([
            term, definition, len(mapped),
            "; ".join(mapped) if mapped else
            "No field name maps to this term directly — it is a concept the data relies on "
            "rather than a column.",
        ])
    _write_table(ws, ["Term", "CFBD definition (verbatim)", "Fields mapped",
                      "Field names carrying it"],
                 glossary_rows, [30, 96, 13, 60], wrap_columns=(2, 4))

    # ---------------------------------------------------------------- Parameters
    ws = wb.create_sheet("Parameters")
    parameter_rows = [[
        p["endpoint"], p["key"], p["parameter"], p["required"], p["type"], p["enum"],
        p["description"],
        "yes" if p["parameter"] in (profile.get(p["key"], {}).get("param_keys") or []) else "no",
    ] for p in parameters]
    _write_table(ws, ["Endpoint", "Raw dir", "Parameter", "Required", "Type", "Enum values",
                      "CFBD description (docs)", "Used by cfdb extractor"],
                 parameter_rows, [24, 22, 22, 11, 20, 40, 60, 16], wrap_columns=(6, 7))

    # ---------------------------------------------------------------- Gaps
    ws = wb.create_sheet("Gaps")
    _write_table(ws, ["Category", "Subject", "Detail", "Suggested action"], gaps,
                 [22, 42, 104, 26], wrap_columns=(3,))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if str(row[3].value).startswith(("Fix", "Land", "Check")):
            for cell in row:
                cell.fill = WARN_FILL

    # ---------------------------------------------------------------- README
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 104

    def line(label, value):
        index = ws.max_row + 1
        ws.cell(row=index, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=index, column=2, value=value)
        cell.font = BASE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def section(title):
        ws.cell(row=ws.max_row + 1, column=1, value=title).font = SECTION_FONT

    ws.cell(row=1, column=1, value="CFBD Data Dictionary — cfdb").font = TITLE_FONT
    ws.append([])
    line("Purpose", "What each CFBD endpoint represents, what its grain is, and what every "
                    "field means — for the endpoints cfdb has actually landed.")
    line("Spec version", "College Football Data API {} (OpenAPI 3.0)".format(spec.version))
    line("Spec source", "https://apinext.collegefootballdata.com/api-docs.json")
    line("Glossary source", "https://collegefootballdata.com/Glossary")
    line("Observed data", "data/raw, profiled directly — largest payload files per endpoint")
    line("Regenerate with", "python -m src.data_dictionary --spec <api-docs.json>")
    line("Captured", captured)
    ws.append([])

    section("HOW TO READ THIS")
    line("Provenance", "Every definition on the Fields sheet is tagged with where it came "
                       "from. This workbook does not hide which parts are sourced.")
    line("  glossary", "CFBD's own published definition, verbatim. Authoritative.")
    line("  docs", "Text CFBD publishes in the OpenAPI spec. Authoritative.")
    line("  spec", "Structural fact from the schema — type, nullability, enum. Authoritative.")
    line("  observed", "Measured from the landed raw files. Factual about the sample profiled.")
    line("  inferred", "Ours, not CFBD's. Read the Confidence column before relying on it.")
    ws.append([])
    line("Confidence", "Applies to inferred rows only.")
    line("  high", "Standard term with one plausible meaning.")
    line("  medium", "Meaning is clear but a detail — units, scope, filter — is unverified.")
    line("  low", "Genuinely ambiguous. Confirm before relying on it.")
    ws.append([])

    section("CONTENTS")
    line("Gaps", "What is missing, what failed, and what will bite. Read this first.")
    line("Fields", "One row per leaf field, flattened to dot-paths; [] marks an array.")
    line("Endpoints", "Per endpoint: CFBD's description, observed grain, registry strategy, "
                      "file and call counts.")
    line("Vocabularies", "The controlled vocabularies in the spec. Conflicting pairs highlighted.")
    line("Glossary", "CFBD's published metric definitions and the fields that carry them.")
    line("Parameters", "Every query parameter, and whether cfdb's extractor uses it.")
    ws.append([])

    section("COUNTS")
    last = len(field_rows) + 1
    for label, formula in [
        ("Field rows", "=COUNTA(Fields!D2:D{})".format(last)),
        ("Endpoints documented", "=COUNTA(Endpoints!A2:A{})".format(len(endpoint_rows) + 1)),
        ("Parameters documented", "=COUNTA(Parameters!C2:C{})".format(len(parameter_rows) + 1)),
        ("Controlled vocabularies", "=COUNTA(Vocabularies!A2:A{})".format(len(vocabulary_rows) + 1)),
        ("Glossary terms", "=COUNTA(Glossary!A2:A{})".format(len(glossary_rows) + 1)),
        ("Gaps and issues", "=COUNTA(Gaps!B2:B{})".format(len(gaps) + 1)),
        ("Definitions from CFBD",
         '=COUNTIF(Fields!J2:J{0},"glossary")+COUNTIF(Fields!J2:J{0},"docs")'.format(last)),
        ("Definitions inferred", '=COUNTIF(Fields!J2:J{},"inferred")'.format(last)),
        ("  of which high confidence", '=COUNTIF(Fields!K2:K{},"high")'.format(last)),
        ("  of which medium confidence", '=COUNTIF(Fields!K2:K{},"medium")'.format(last)),
        ("  of which low confidence", '=COUNTIF(Fields!K2:K{},"low")'.format(last)),
    ]:
        index = ws.max_row + 1
        ws.cell(row=index, column=1, value=label).font = LABEL_FONT
        ws.cell(row=index, column=2, value=formula).font = BASE_FONT
    ws.append([])

    section("LIMITS")
    line("Grain is observed", "CFBD states grain nowhere. Each grain here is the smallest "
                              "column combination unique in the largest sampled response file. "
                              "Disagreement across files is noted rather than resolved.")
    line("Sampling", "Profiling reads the largest payload files per endpoint, capped at 6,000 "
                     "records. Null rates and distinct counts describe that sample.")
    line("Nested fields", "Observed columns are blank for nested dot-paths — the profiler "
                          "measures top-level keys only.")
    line("Inference", "Where this says 'inferred', CFBD has published nothing. Confirm before "
                      "quoting those readings as definitions on the website.")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)
    return dict(fields=len(field_rows), endpoints=len(endpoint_rows),
                parameters=len(parameter_rows), vocabularies=len(vocabulary_rows),
                glossary=len(glossary_rows), gaps=len(gaps))
