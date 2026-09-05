"""The Excel deliverable, checked against its twelve acceptance criteria.

The workbook is the feature closest to the licence line and the only artifact that leaves
the site, so most of these assert obligations rather than behaviour: attribution present on
every sheet, the model disclaimer on every sheet carrying a prediction, and no code path
that widens scope beyond the filters.

Built against a stubbed query so the whole thing runs in CI without a database. That is not
a compromise — the properties under test are properties of the WRITER, and a fixture with
one row of each awkward type (null, NaN, a tz-aware timestamp, an empty string, a bool)
exercises them harder than real data does.
"""
import itertools
import math
import re
import sys
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "site"))

from lib import workbook                                # noqa: E402
from lib.query import check_contract                    # noqa: E402

# WHICH SERVING VIEWS ACTUALLY CARRY `is_fbs_game`, READ OFF THE MODELS RATHER THAN LISTED.
#
# A hardcoded list here would be a second copy of the fact, and the first copy is the dbt
# model. The previous version of the division test hardcoded the view NAME instead of the
# property and went red when srv_game_team gained the column — the test was defending the
# state of the world in September rather than the rule.
SERVING_MODELS = Path(__file__).resolve().parents[1] / "dbt" / "models" / "serving"
VIEWS_WITH_IS_FBS_GAME = {
    path.stem for path in SERVING_MODELS.glob("*.sql")
    if re.search(r"\bas\s+is_fbs_game\b", path.read_text(encoding="utf-8"))
}
# A glob that finds nothing would make the test above vacuously true for every sheet.
assert VIEWS_WITH_IS_FBS_GAME, "no serving model declares is_fbs_game — the glob is wrong"


def _frame(fields):
    """Two rows per sheet: one populated, one carrying every awkward value at once.

    The second row is the point. A null, a NaN, a tz-aware timestamp and an empty string
    are the four things that turn into either an exception at write time or a repair prompt
    at open time, and all four arrive from real serving views.
    """
    populated, awkward = {}, {}
    for field in fields:
        if field.endswith("_ts") or field.endswith("_date") or field.endswith("_et"):
            populated[field] = pd.Timestamp("2026-09-05 19:30", tz="UTC")
            awkward[field] = pd.NaT
        elif field.startswith("is_"):
            populated[field] = True
            awkward[field] = False
        elif field in ("week", "season", "games", "attendance"):
            populated[field] = 7
            awkward[field] = None
        elif ("point" in field or "margin" in field or "spread" in field
              or "edge" in field or "prob" in field or "pct" in field
              or "score" in field or "total" in field or "loss" in field
              or field.endswith("_rank") or field.endswith("moneyline")
              or field in ("wins", "losses", "ties", "win_pct", "error", "n")):
            populated[field] = -3.5
            awkward[field] = float("nan")
        else:
            populated[field] = "text"
            awkward[field] = ""
    return pd.DataFrame([populated, awkward])


@pytest.fixture
def built(monkeypatch):
    """A real workbook, from stubbed reads.

    CFDB_SITE_HOST is set EXPLICITLY rather than inherited. A test that behaves one way on a
    laptop with a populated .env and another way on a CI runner is the class of test
    conftest.py exists to prevent, and the hyperlink work made this module sensitive to it.
    """
    monkeypatch.setenv("CFDB_SITE_HOST", "https://cfdb.example")

    def fake_query(sql, params=None):
        # Every stubbed query still goes through the contract, so a sheet that violates
        # G-1/G-2 fails here rather than in production.
        check_contract(sql)
        flat = " ".join(sql.split())
        for sheet in workbook.SHEETS:
            # WORD BOUNDARY, AND THE SUBSTRING VERSION HID AN EMPTY SHEET FOR A WHOLE ROUND.
            #
            # `"from srv_game" in "from srv_game_team"` is True. Schedule comes first in
            # SHEETS, so every Scores query matched Schedule and was handed a frame built
            # from SCHEDULE's fields — meaning the Scores sheet was constructed with none of
            # its own columns present, and every `built`-based assertion about it ran against
            # an empty sheet while passing.
            #
            # Caught only because the banding test refused to pass on finding nothing to
            # compare. R-157 again: the tests that survive are the ones that fail loudly when
            # handed nothing.
            if re.search(rf"\bfrom {re.escape(sheet.view)}\b", flat):
                # The SELECTED fields, not just the visible columns: the sheet also pulls
                # what it derives from and what it links with, and a fixture missing those
                # makes the hyperlinks look absent when they are merely unstubbed.
                df = _frame(sorted(set(sheet.fields) | set(sheet.selected_fields)))
                if "game_id" in df.columns:
                    df["game_id"] = [401752000, 401752001]
                # Realistic verdicts, so the mark rendering is actually exercised. Without
                # them every verdict cell is the no-data dash and any test about the marks
                # passes while proving nothing — which is exactly what the colour test
                # refused to do.
                # Distinct team and venue names, because the alignment rule now depends on
                # CARDINALITY: with both rows reading "text" the fixture made a team column
                # look like a two-value category and centred it, which is not what real data
                # does. A fixture that misrepresents the shape of the data cannot test a rule
                # that reads the shape of the data.
                for column, values in (("away_team_display", ["Rice", "Auburn"]),
                                       ("home_team_display", ["Louisville", "Texas"]),
                                       ("venue_display", ["Cardinal Stadium", "DKR"]),
                                       ("winner", ["Louisville", "Texas"]),
                                       ("upset_level", ["upset", "none"]),
                                       ("winner_covered_close", ["yes", "no"]),
                                       ("over_met", ["yes", "no"]),
                                       ("favorite_covered", ["yes", "no_favorite"]),
                                       # Scores measures its frozen pane in CHARACTERS, and
                                       # a column of the literal string "text" is four wide.
                                       # Real school names are the widest thing on the sheet
                                       # and they are what makes that measurement mean
                                       # anything — a fixture narrower than the data turns
                                       # the width assertion into a formality.
                                       ("team", ["Mississippi State",
                                                 "Southern California"]),
                                       ("opponent", ["Northwestern",
                                                     "Texas A&M"]),
                                       ("conference", ["SEC", "Big Ten"]),
                                       ("result", ["W", "L"]),
                                       # Scores' own columns. `game_no` drives the banding,
                                       # so the two rows carry different parities — without
                                       # that the band test has nothing to compare. Possession
                                       # must be numeric or the minutes derivation raises.
                                       ("game_no", [1, 2]),
                                       ("possession_seconds", [1546, 2054]),
                                       ("covered_final", ["yes", "no"]),
                                       ("covered_open", ["yes", "push"])):
                    if column in df.columns:
                        df[column] = values
                return df
        return pd.DataFrame([{"model_version": "abc123", "model_name": "stub"}])

    monkeypatch.setattr(workbook, "query", fake_query)
    payload, index_rows, omitted = workbook.build(2026, 8, "regular", None)
    from openpyxl import load_workbook
    return payload, load_workbook(BytesIO(payload)), index_rows, omitted


# --- AC-15.6: the file opens ------------------------------------------------------------

def test_the_workbook_is_a_structurally_valid_xlsx(built):
    payload, _, _, _ = built
    archive = zipfile.ZipFile(BytesIO(payload))
    assert archive.testzip() is None
    assert "[Content_Types].xml" in archive.namelist()


def test_no_cell_holds_a_value_excel_refuses_to_open(built):
    """NaN, infinity and a tz-aware datetime each produce a repair prompt or an exception.

    The file is structurally valid XML in the NaN case and Excel still refuses it, which is
    why this checks values rather than trusting that the save succeeded.
    """
    _, book, _, _ = built
    for name in book.sheetnames:
        for row in book[name].iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, float):
                    assert not math.isnan(value), f"{name}!{cell.coordinate}"
                    assert not math.isinf(value), f"{name}!{cell.coordinate}"
                if isinstance(value, datetime):
                    assert value.tzinfo is None, f"{name}!{cell.coordinate}"


def test_sheet_names_are_legal(built):
    """Over 31 characters, or any of []:*?/\\, and Excel rewrites or rejects the name."""
    _, book, _, _ = built
    for name in book.sheetnames:
        assert len(name) <= 31, name
        assert not set(name) & set(r"[]:*?/\\"), name
    assert len(set(book.sheetnames)) == len(book.sheetnames)


# --- AC-15.3 / AC-15.4: the obligations that travel with the file -----------------------

def test_every_sheet_carries_cfbd_attribution_in_a_fixed_cell(built):
    _, book, _, _ = built
    for name in book.sheetnames:
        assert "CollegeFootballData.com" in str(
            book[name].cell(workbook.ROW_CREDIT, 1).value), name


def test_every_sheet_that_declares_the_disclaimer_carries_it(built):
    """AC-15.4 was "every sheet with predictions writes the disclaimer". R-221 split that in
    two, and the split is the point: Schedule STILL CARRIES PREDICTIONS and no longer writes
    the sheet-level line, because attribution now travels per row instead.

    So the assertion follows `sheet_disclaimer`, and the test below is what makes removing
    the line safe rather than merely requested.
    """
    _, book, _, _ = built
    for sheet in workbook.SHEETS:
        if sheet.name not in book.sheetnames:
            continue
        tab = book[sheet.name]
        text = str(tab.cell(workbook.ROW_DISCLAIMER, 1).value)
        if sheet.sheet_disclaimer:
            assert "NOT CollegeFootballData.com predictions" in text, sheet.name
        else:
            assert "NOT CollegeFootballData.com predictions" not in text, sheet.name


def test_a_sheet_without_the_disclaimer_must_carry_attribution_per_row(built):
    """R-221, AND THIS IS THE LOAD-BEARING HALF.

    Removing a licence statement from a file that travels off the site, while the numbers it
    covered are still on the sheet, is not a formatting change. What makes it safe is that
    attribution is carried AS DATA, per row, from dim_model_version — which is strictly
    stronger than a line in row 2, because it survives filtering, sorting and copy-paste.

    So a sheet that drops the sheet-level line must have the column. If it ever has neither,
    the workbook ships unattributed predictions.
    """
    for sheet in workbook._ALL_SHEETS:
        if sheet.has_predictions and not sheet.sheet_disclaimer:
            assert "attribution" in sheet.fields, (
                f"{sheet.name} carries predictions, writes no disclaimer, and has no "
                f"attribution column — that is an unattributed prediction leaving the site")


def test_no_prediction_cell_is_populated_beside_a_blank_attribution(built):
    """The runtime half of the same guarantee, on the built file rather than the definition.

    Attribution being null where there is no model row is CORRECT — no prediction, nothing to
    attribute. The failure case is the other way round: a populated `Pred margin` beside a
    blank `Attribution`.
    """
    _, book, _, _ = built
    schedule = next(s for s in workbook.SHEETS if s.name == "Schedule")
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    prediction_columns = ["Pred margin", "Home win prob", "Confidence", "Model",
                          "Model version"]
    assert "Attribution" in labels and schedule.sheet_disclaimer is False
    for row in range(header + 1, tab.max_row + 1):
        attributed = tab.cell(row, labels["Attribution"]).value
        for column in prediction_columns:
            value = tab.cell(row, labels[column]).value
            if value not in (None, ""):
                assert attributed not in (None, ""), (
                    f"row {row}: {column} is populated and Attribution is blank")


def test_the_out_of_sample_flag_is_a_column_not_a_footnote(built):
    """AC-15.4. A workbook gets sorted and filtered; a sheet-level note does not survive
    that, and the rows it applied to end up looking like ordinary predictions."""
    # Asserted on every sheet that carries predictions, not on one named sheet — the named
    # sheet stopped shipping and the test went looking for something that was not there.
    predicting = [s for s in workbook._ALL_SHEETS if s.has_predictions]
    assert predicting
    for sheet in predicting:
        assert "is_out_of_sample_week" in sheet.fields, sheet.name


# --- AC-15.1 / AC-15.2: the scope rule --------------------------------------------------

def test_every_sheet_reads_exactly_one_serving_view(built):
    """AC-15.2, enforced by the same contract the pages use rather than by reading them."""
    for sheet in workbook.SHEETS:
        assert check_contract(" ".join(sheet.sql.split())) == sheet.view


def test_scoped_sheets_filter_on_the_season(built):
    """AC-15.1. The two unscoped sheets are provenance — which models produced the
    predicted columns, and what the fields mean — not additional data."""
    # Walks _ALL_SHEETS, INCLUDING the six that do not ship yet. Their SQL is still real
    # work and still under the CI query checker; letting the property lapse while they sit
    # out of the shipping list is how they would come back broken.
    unscoped = {s.name for s in workbook._ALL_SHEETS if not s.scoped}
    assert unscoped == {"Model performance", "Data dictionary"}
    for sheet in workbook._ALL_SHEETS:
        if sheet.scoped:
            assert ":season" in sheet.sql, sheet.name


def test_no_sheet_offers_an_unbounded_export():
    """Every query carries a LIMIT, and none reads raw or staging."""
    for sheet in workbook.SHEETS:
        flat = " ".join(sheet.sql.split()).lower()
        assert " limit " in flat, sheet.name
        assert "raw." not in flat and "staging." not in flat, sheet.name


# --- AC-15.5: an omission is named ------------------------------------------------------

def test_an_empty_sheet_is_omitted_rather_than_shipped_blank(monkeypatch):
    def empty_query(sql, params=None):
        check_contract(sql)
        return pd.DataFrame()
    monkeypatch.setattr(workbook, "query", empty_query)
    payload, index_rows, omitted = workbook.build(2026, 1, "regular", None)
    assert index_rows == []
    assert {name for name, _ in omitted} == {s.name for s in workbook.SHEETS}
    from openpyxl import load_workbook
    book = load_workbook(BytesIO(payload))
    # The index still exists and still says what happened. A workbook with no tabs at all
    # would be indistinguishable from a failed download.
    assert book.sheetnames == ["Index"]
    text = "\n".join(str(c.value) for row in book["Index"].iter_rows() for c in row)
    for sheet in workbook.SHEETS:
        assert sheet.view in text


def test_a_real_failure_is_not_reported_as_an_omission(monkeypatch):
    """A connection or permission error is not "no rows in this scope".

    The first version of the reader caught everything and reported every sheet as omitted,
    which produced a workbook calmly stating the data was unavailable while the real
    problem was that nothing could be read at all.
    """
    def broken(sql, params=None):
        raise RuntimeError("connection refused")
    monkeypatch.setattr(workbook, "query", broken)
    with pytest.raises(RuntimeError):
        workbook.build(2026, 1, "regular", None)


# --- AC-15.7 / AC-15.12: it is a spreadsheet, not a picture of one ----------------------

def test_numeric_cells_are_numbers_with_the_sites_precision(built):
    _, book, _, _ = built
    checked = 0
    for sheet in workbook.SHEETS:
        if sheet.name not in book.sheetnames:
            continue
        tab = book[sheet.name]
        first_data = workbook.first_data_row(2 if sheet.sheet_disclaimer else 1)
        for index, (field, _) in enumerate(sheet.columns, start=1):
            cell = tab.cell(first_data, index)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                # THROUGH THE SHEET'S OWN RULES, not the bare defaults. A sheet may declare
                # its own decimal default and its own set of naturally-integer columns
                # (R-259), and comparing against the module defaults would report every one
                # of those as a mismatch — while quietly asserting nothing about whether the
                # sheet's rules were applied at all.
                assert cell.number_format == workbook.number_format(
                    field, sheet.decimals, sheet.integer_fields,
                    sheet.site_precision), f"{sheet.name}.{field}"
                checked += 1
    assert checked > 10


def test_precision_comes_from_the_same_table_the_site_renders_with():
    """AC-G.31 in the workbook. A column that reads 1 dp on screen and 2 dp in Excel is
    two different claims about how precisely the number is known."""
    assert workbook.number_format("spread") == "#,##0.0"
    assert workbook.number_format("margin_mae") == "#,##0.00"
    assert workbook.number_format("home_win_probability") == "#,##0.000"
    # Counts are counts. A decimal point on an attendance makes it look measured.
    assert workbook.number_format("attendance") == "#,##0"
    assert workbook.number_format("home_moneyline") == "+#,##0;-#,##0"


def test_sheets_are_workable_not_merely_readable(built):
    """AC-15.12: freeze panes, filtering, and conditional formatting where it earns its
    place. The deliverable is meant to be worked in.

    THE FILTER NOW COMES FROM THE TABLE, NOT FROM `auto_filter` (R-182 trap 1). openpyxl
    documents that a table must not overlap the worksheet's autofilter, and an overlap is
    the "we found a problem with some content" repair prompt AC-15.6 forbids. So the
    affordance is asserted, and the thing that would collide with it is asserted ABSENT.
    """
    _, book, _, _ = built
    for sheet in workbook.SHEETS:
        if sheet.name not in book.sheetnames:
            continue
        tab = book[sheet.name]
        from openpyxl.utils import get_column_letter
        expected_row = workbook.first_data_row(2 if sheet.sheet_disclaimer else 1)
        expected_col = get_column_letter(sheet.freeze_column())
        assert tab.freeze_panes == f"{expected_col}{expected_row}", sheet.name
        assert not tab.auto_filter.ref, (
            f"{sheet.name} has a worksheet autofilter as well as a Table; overlapping them "
            f"is what produces Excel's repair prompt")
        assert len(tab.tables) == 1, f"{sheet.name} should carry exactly one Table"
    assert len(book["Schedule"].conditional_formatting._cf_rules) > 0


def test_an_empty_string_becomes_an_empty_cell(built):
    """The pack writes '' where it has no confidence bucket. An empty-string cell is not an
    empty cell: it breaks COUNTBLANK and sorts ahead of real values."""
    assert workbook._clean("") is None
    assert workbook._clean("   ") is None
    assert workbook._clean("high") == "high"


def test_null_and_zero_stay_distinguishable_in_the_workbook():
    """AC-G.32 does not stop applying because the output is a file."""
    assert workbook._clean(None) is None
    assert workbook._clean(float("nan")) is None
    assert workbook._clean(0) == 0
    assert workbook._clean(0.0) == 0.0


# --- AC-15.9 / AC-15.10: the file explains itself ---------------------------------------

def test_the_index_states_scope_timestamp_counts_and_model_version(built):
    _, book, index_rows, _ = built
    text = "\n".join(str(c.value) for row in book["Index"].iter_rows() for c in row)
    assert "Generated (UTC)" in text
    assert "season 2026" in text and "week 8" in text
    assert "Model version(s)" in text
    for entry in index_rows:
        assert entry.name in text and entry.view in text


def test_the_filename_encodes_its_own_scope():
    """AC-15.10. A folder of files called export.xlsx is a folder nobody can tell apart,
    and scope is the single most important fact about an export designed to be bounded."""
    stamp = datetime(2026, 8, 20, tzinfo=timezone.utc)
    assert workbook.filename(2026, 8, stamp) == "cfdb_week08_2026_20260820.xlsx"
    assert workbook.filename(2026, None, stamp) == "cfdb_season_2026_20260820.xlsx"


def test_no_duplicate_column_headers_on_any_sheet():
    """Two columns headed "Model" is merely cluttered on a web page and unusable in a
    spreadsheet — a filter dropdown cannot tell them apart. Found in the Edges page by
    building the export from its columns."""
    for sheet in workbook.SHEETS:
        labels = [label for _, label in sheet.columns]
        assert len(labels) == len(set(labels)), f"{sheet.name}: {labels}"


def test_export_labels_agree_with_the_site(built):
    """AC-15.8. Where a page shows the same field, it shows it under the same name.

    Compared per (VIEW, field), not per field. The first version keyed on the field name
    alone and reported a mismatch on `market`, which is a composite spread-and-model cell
    on Today and a genuine srv_edge_finder column on Edges — two different things sharing a
    name. A test that cannot tell those apart teaches you to ignore it.

    The cells themselves cannot be shared: the site renders a team as a logo plus a name in
    one cell, the workbook writes the display name. The LABEL is what a reader carries
    between the two, so the label is what is checked.
    """
    import re
    from lib.registry import PAGES
    site = Path(__file__).resolve().parents[1] / "site"

    view_labels = {}
    for page in PAGES:
        path = site / "views" / f"{page.key}.py"
        if not page.view or not path.exists():
            continue
        for field, label in re.findall(r'Col\(\s*"(\w+)"\s*,\s*"([^"]*)"', path.read_text()):
            view_labels.setdefault(page.view, {}).setdefault(field, set()).add(label)

    mismatches, compared = [], {}
    for sheet in workbook.SHEETS:
        shown_on_page = view_labels.get(sheet.view, {})
        compared[sheet.name] = 0
        for field, label in sheet.columns:
            shown = shown_on_page.get(field)
            if not shown:
                continue
            compared[sheet.name] += 1
            if label not in shown and field not in workbook.EXPORT_ONLY_LABELS:
                mismatches.append((sheet.name, field, label, sorted(shown)))
    assert not mismatches, mismatches

    # WHAT THIS TEST DID NOT CHECK, SAID OUT LOUD.
    #
    # The comparison is per (view, field), so a sheet whose view no page reads is skipped
    # entirely — silently, and with a green tick. Scores is exactly that since R-255: 132
    # columns from srv_game_team, which no page reads, so ZERO of its labels are checked
    # here. That is not a defect in the sheet; there is genuinely nothing to agree with. It
    # is a gap in coverage, and a gap nobody can see is the one that gets forgotten.
    #
    # So the sheets with no counterpart are enumerated. The day a page reads srv_game_team,
    # this assertion fails and the sheet joins the comparison — which is the right time to
    # notice, rather than the day the two names quietly diverge.
    uncovered = {name for name, n in compared.items() if n == 0}
    assert uncovered == {"Scores"}, (
        f"sheets compared against no page: {uncovered or 'none'} — expected exactly "
        f"{{'Scores'}}, whose view srv_game_team no page reads")
    # And Schedule really is compared — 13 of its 56 columns share a field name with the
    # Schedule page today. A floor rather than the exact number, so a legitimate page change
    # does not fail this, but a drop to nothing does.
    assert compared["Schedule"] >= 10, compared


def test_every_declared_label_divergence_is_still_real():
    """A documented exception that no longer diverges is stale documentation.

    Without this, EXPORT_ONLY_LABELS becomes a list nobody prunes, and the next genuine
    mismatch gets waved through because "it is probably one of those".
    """
    import re
    from lib.registry import PAGES
    site = Path(__file__).resolve().parents[1] / "site"

    view_labels = {}
    for page in PAGES:
        path = site / "views" / f"{page.key}.py"
        if not page.view or not path.exists():
            continue
        for field, label in re.findall(r'Col\(\s*"(\w+)"\s*,\s*"([^"]*)"', path.read_text()):
            view_labels.setdefault(page.view, {}).setdefault(field, set()).add(label)

    for field in workbook.EXPORT_ONLY_LABELS:
        # TWO SHAPES OF DIVERGENCE, and the second was missed until R-101 produced one.
        #
        #   different  the site has a column for the field under another header
        #   absent     the site does not surface the field as a column at ALL
        #
        # The original version only knew the first, so when R-101 folded the neutral-site
        # glyph into the shared Game column this test read "no longer differs" and asked for
        # the exception to be dropped — when the divergence had in fact just got wider. An
        # exception is stale only when the site and the sheet AGREE on a header.
        diverges = any(
            field not in view_labels.get(sheet.view, {})
            or label not in view_labels[sheet.view][field]
            for sheet in workbook._ALL_SHEETS
            for name, label in sheet.columns if name == field)
        # _ALL_SHEETS, not SHEETS: an exception belonging to a sheet that is written but not
        # yet SHIPPED is still live documentation, and pruning it now would mean rediscovering
        # the same divergence when that sheet is converted.
        exported = any(name == field
                       for sheet in workbook._ALL_SHEETS for name, _ in sheet.columns)
        assert exported, f"{field} is not exported at all; drop the exception"
        assert diverges, f"{field} no longer differs from the site; drop the exception"


# === R-184 / R-196: the scope the user asked for, and the truth about what was written ====
#
# These are the Part 1 defect tests. Every one of them is negative-tested in
# `tests/test_workbook_scope_negatives.py`, because the failure being fixed here is
# specifically a check that agreed with a wrong answer.

GAME_SHEETS = [s for s in workbook.SHEETS if s.view == "srv_game"]


def _division_aware_query(recorder=None, in_scope=None):
    """A stub that HONOURS the division parameter instead of ignoring it.

    A stub that returns the same rows whatever it is asked is how this defect survived: the
    page, the preview and the file all agreed, and none of them had applied the filter. So
    this fake behaves like the database on the one axis under test — it drops the non-FBS
    row when the caller binds a division other than 'all'.
    """
    def fake_query(sql, params=None):
        check_contract(sql)
        flat = " ".join(sql.split())
        if recorder is not None:
            recorder.append((flat, dict(params or {})))
        for sheet in workbook.SHEETS:
            if f"from {sheet.view}" in flat:
                df = _frame(sorted(set(sheet.fields) | set(sheet.selected_fields)))
                if "game_id" in df.columns:
                    df["game_id"] = [401752000, 401752001]
                if sheet.view == "srv_game":
                    # Row 0 is an FBS game, row 1 is two non-FBS teams.
                    df["is_fbs_game"] = [True, False]
                    df["home_classification"] = ["fbs", "ii"]
                    df["away_classification"] = ["fbs", "iii"]
                    if (params or {}).get("division", "all") != "all":
                        df = df[df["is_fbs_game"]].reset_index(drop=True)
                if in_scope is not None:
                    df["rows_in_scope"] = in_scope
                return df
        return pd.DataFrame([{"model_version": "abc123", "model_name": "stub"}])
    return fake_query


def test_division_reaches_the_query_as_a_bound_parameter(monkeypatch):
    """THE DEFECT ITSELF. `build()` took four of GameScope's five filters, so `division` was
    dropped on the floor at `export.py:58` and again at `:86`.

    Asserted on the BINDING rather than on the SQL text, because the SQL was never the
    problem — the predicate can be perfectly written and still never receive a value.
    """
    seen = []
    monkeypatch.setattr(workbook, "query", _division_aware_query(recorder=seen))
    workbook.build(2026, 8, "regular", None, "fbs")

    bound = [params for sql, params in seen if "from srv_game" in sql]
    assert bound, "no srv_game sheet was read at all"
    for params in bound:
        assert params.get("division") == "fbs", (
            "the division filter never reached the query; this is R-184 exactly")


def test_a_workbook_scoped_to_fbs_holds_no_game_between_two_non_fbs_teams(monkeypatch):
    """AND THE SAME BUILD AT 'all' KEEPS IT. A test that only proves the row is absent
    cannot tell "the filter worked" from "the fixture had nothing to filter"."""
    monkeypatch.setattr(workbook, "query", _division_aware_query())

    def classifications(division):
        payload, _, _ = workbook.build(2026, 8, "regular", None, division)
        from openpyxl import load_workbook
        book = load_workbook(BytesIO(payload))
        rows = 0
        for sheet in GAME_SHEETS:
            if sheet.name in book.sheetnames:
                tab = book[sheet.name]
                first = workbook.first_data_row(2 if sheet.sheet_disclaimer else 1)
                rows += tab.max_row - first + 1
        return rows

    narrow, wide = classifications("fbs"), classifications("all")
    assert narrow < wide, (
        f"FBS wrote {narrow} rows and All divisions wrote {wide} — the filter changed "
        "nothing, so this test would pass against the defect it exists to catch")


def test_the_export_and_the_page_spell_the_division_rule_identically():
    """One rule, one spelling. The export must return the set the page displayed, and the
    way that breaks is not a wrong predicate — it is a SECOND predicate that starts out
    right and drifts. Compared as text, so a change to either side has to change both.

    EITHER team FBS, not both: a Division II visitor's trip to an FBS stadium is an FBS
    game, and requiring both drops 20 of the 25 games on the opening Thursday.
    """
    page = (Path(__file__).resolve().parents[1] / "site" / "views" / "schedule.py").read_text()
    predicate = "(:division = 'all' or is_fbs_game)"
    assert predicate in page, (
        "the Schedule page no longer spells the rule this way; the export must follow it, "
        "not the other way round")
    for sheet in GAME_SHEETS:
        assert predicate in " ".join(sheet.sql.split()), (
            f"{sheet.name} reads srv_game but does not apply the page's division rule")


def test_every_sheet_query_carries_the_named_cap_and_no_literal_limit():
    """The cap was four different magic numbers across seven queries — 400, 900, 1000, 2000
    — and nothing named any of them, so nothing could report one. One constant, reported."""
    import re as _re
    for sheet in workbook.SHEETS:
        limits = _re.findall(r"limit\s+(\d+)", sheet.sql, _re.IGNORECASE)
        assert limits == [str(workbook.ROW_CAP)], (
            f"{sheet.name} limits to {limits}, not the named cap {workbook.ROW_CAP}")


def test_the_cap_clears_a_full_season_at_the_widest_scope():
    """Measured, not chosen: a season at All Divisions is 3,745 games on the serving
    database. A cap below that truncates the single most obvious thing to ask for."""
    assert workbook.ROW_CAP >= 3745


def test_rows_written_and_rows_in_scope_are_two_facts(monkeypatch):
    """R-196. The preview ran the real query and so did the build, which guaranteed they
    agreed — and both were wrong by 3,345 rows.

    A count and a cap are two facts about one query, so they come back from one query.
    """
    monkeypatch.setattr(workbook, "query", _division_aware_query(in_scope=3745))
    sheet = GAME_SHEETS[0]
    read = workbook.read_sheet(sheet, 2026, None, "regular", None, "fbs")
    assert read.rows == 1, "the honest stub returns one FBS row"
    assert read.rows_in_scope == 3745
    assert read.truncated


def test_the_index_names_the_truncation_the_number_and_the_cure(monkeypatch):
    """A silently short workbook is the worst failure this file can have: a plausible number
    of rows, correctly formatted, with no way for the reader to know what is missing."""
    monkeypatch.setattr(workbook, "query", _division_aware_query(in_scope=3745))
    payload, index_rows, _ = workbook.build(2026, 8, "regular", None, "fbs")
    from openpyxl import load_workbook
    text = "\n".join(str(c.value) for row in load_workbook(BytesIO(payload))["Index"].iter_rows()
                     for c in row)
    assert "Truncated" in text
    assert "3,745" in text, "the Index must state how many rows were in scope"
    assert f"{workbook.ROW_CAP:,}" in text, "and what the cap was"
    assert "Narrow the filters" in text, "and what to do about it"
    assert any(entry.truncated for entry in index_rows)


def test_an_untruncated_workbook_says_so_rather_than_staying_silent(monkeypatch):
    """The other half of the same claim. 'No warning' is indistinguishable from 'nobody
    checked', which is the whole lesson of R-194."""
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    payload, _, _ = workbook.build(2026, 8, "regular", None, "fbs")
    from openpyxl import load_workbook
    text = "\n".join(str(c.value) for row in load_workbook(BytesIO(payload))["Index"].iter_rows()
                     for c in row)
    assert "Complete" in text and "Truncated" not in text


def test_only_sheets_whose_view_carries_is_fbs_game_can_honour_the_division_filter():
    """A scope line reading 'FBS' must not overstate what a sheet actually did.

    THE OLD VERSION OF THIS TEST TIED THE PROPERTY TO A TABLE NAME — `sheet.view ==
    "srv_game"` — and that was a true statement about the world on the day it was written
    rather than the rule it was standing in for. R-255 added `is_fbs_game` to srv_game_team
    for exactly this reason, and the test went red for a sheet that had become MORE correct.

    The rule is: a sheet can narrow if and only if its view carries the column. Both halves
    are asserted, so a sheet cannot claim the filter without the column, and cannot have the
    column and quietly ignore it.
    """
    scoped = {s.name for s in workbook._ALL_SHEETS if s.division_scoped}
    assert scoped == {"Schedule", "Scores"}, scoped
    for sheet in workbook._ALL_SHEETS:
        carries = sheet.view in VIEWS_WITH_IS_FBS_GAME
        assert sheet.division_scoped == carries, sheet.name
        if carries:
            assert "is_fbs_game" in " ".join(sheet.sql.split()), sheet.name


def test_the_index_says_when_the_division_filter_could_not_reach_a_sheet(monkeypatch):
    """The note itself, built for a sheet that cannot narrow. Exercised directly because the
    only sheet shipping today CAN narrow, and a test that silently covers nothing is R-194.
    """
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    standings = next(s for s in workbook._ALL_SHEETS if s.name == "Standings")
    monkeypatch.setattr(workbook, "SHEETS", [standings])
    _, index_rows, _ = workbook.build(2026, 8, "regular", None, "fbs")
    note = next(e for e in index_rows if e.name == "Standings").note
    assert "does not apply" in note and "FBS" in note

    monkeypatch.setattr(workbook, "SHEETS", workbook._ALL_SHEETS[:1])
    _, index_rows, _ = workbook.build(2026, 8, "regular", None, "fbs")
    assert "does not apply" not in next(e for e in index_rows if e.name == "Schedule").note


def test_the_scope_line_names_the_division_unless_it_excludes_nothing():
    assert "FBS" in workbook.describe_scope(2026, 8, "regular", None, "fbs")
    assert "ALL" not in workbook.describe_scope(2026, 8, "regular", None, "all").upper()[10:]


def test_no_sheet_query_contains_a_line_comment():
    """`read_sheet()` sends `" ".join(sql.split())` — the whole query on ONE line.

    So a `--` comment does not comment out its line; it comments out EVERY REMAINING CLAUSE
    of the query. This is not hypothetical: the R-184 predicate was first written with `--`
    above it, and the flattened SQL silently lost its own WHERE clause, its ORDER BY and its
    LIMIT. Postgres reported a syntax error at the next statement, which is a long way from
    the cause.

    Block comments survive flattening. Use them.
    """
    import re as _re
    for sheet in workbook.SHEETS:
        # Block comments stripped FIRST. The eighth time in this repo a source-reading check
        # has matched its own prose: the block comment explaining this rule contains the very
        # two characters it forbids, and the first version of this test failed on it.
        code = _re.sub(r"/\*.*?\*/", "", sheet.sql, flags=_re.S)
        assert "--" not in code, (
            f"{sheet.name}'s query has a `--` comment. read_sheet() flattens the SQL to one "
            f"line, so everything after it is commented out. Use /* ... */ instead.")


def test_the_flattened_sql_survives_comment_semantics_not_just_string_matching():
    """The property the test above protects — asserted the way POSTGRES would see it.

    The first version of this test was green with the defect present, which is the failure
    class this project keeps rediscovering: a true assertion about the wrong property. It
    checked that the flattened text ENDS WITH the cap, and it does — `limit 5000` is still
    the last thing in the string when a `--` has commented it out. `check_contract` agreed
    for the same reason: its LIMIT check is a regex over text, and text is not what runs.

    So apply what the server applies. Flatten, then delete from any `--` to end of line —
    which, once flattened, is end of statement — and assert what SURVIVES is still the whole
    query.
    """
    import re as _re
    for sheet in workbook.SHEETS:
        flat = " ".join(sheet.sql.split())
        as_parsed = _re.sub(r"/\*.*?\*/", " ", flat, flags=_re.S)
        as_parsed = _re.sub(r"--.*$", "", as_parsed)
        assert check_contract(as_parsed) == sheet.view, (
            f"{sheet.name}'s query does not survive its own comments")
        assert _re.search(rf"\blimit\s+{workbook.ROW_CAP}\b", as_parsed), (
            f"{sheet.name} loses its LIMIT once comments are applied — a `--` has swallowed "
            f"the tail. What the server would run: ...{as_parsed[-110:]}")
        assert "order by" in as_parsed.lower(), (
            f"{sheet.name} loses its ORDER BY once comments are applied")


# === R-181 / R-182 / R-183: the layout ====================================================

def test_exactly_one_blank_row_sits_between_the_notes_and_the_header(built):
    """R-181, Marc's global requirement. Excel treats a blank row as the end of a region, so
    a second blank changes what a header-click and Ctrl+A select — it is not cosmetic.

    Asserted by READING THE CELLS, not by recomputing the constant. A test that asks
    `header_row(n) == n + 2` restates the implementation and would pass with the sheet laid
    out any way at all.
    """
    _, book, _, _ = built
    for name in book.sheetnames:
        tab = book[name]
        column_a = [tab.cell(r, 1).value for r in range(1, 12)]
        notes = 0
        while notes < len(column_a) and column_a[notes] not in (None, ""):
            notes += 1
        assert notes >= 1, f"{name} has no note block at all — attribution is structural"
        blanks = 0
        while column_a[notes + blanks] in (None, ""):
            blanks += 1
        assert blanks == 1, (
            f"{name} has {blanks} blank row(s) between its notes and its header, not 1")


def test_the_blank_row_holds_on_a_sheet_with_the_disclaimer_and_one_without(monkeypatch):
    """THE NEGATIVE HALF, and the reason the old layout was wrong.

    Four fixed constants gave exactly one blank row on a sheet WITH the model disclaimer and
    two on a sheet without — so a test that only ever saw a prediction sheet would have
    reported the old layout as correct. Both shapes, in one test.
    """
    monkeypatch.setenv("CFDB_SITE_HOST", "https://cfdb.example")
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    from openpyxl import load_workbook

    seen = {}
    for sheet in (next(s for s in workbook._ALL_SHEETS if s.sheet_disclaimer),
                  next(s for s in workbook._ALL_SHEETS if not s.sheet_disclaimer)):
        monkeypatch.setattr(workbook, "SHEETS", [sheet])
        payload, _, _ = workbook.build(2026, 8, "regular", None, "all")
        tab = load_workbook(BytesIO(payload))[sheet.name]
        column_a = [tab.cell(r, 1).value for r in range(1, 12)]
        # The CONTIGUOUS prefix. Counting non-empty cells in the first three rows instead
        # counted the header as a note the moment the block was one line long.
        notes = 0
        while column_a[notes] not in (None, ""):
            notes += 1
        seen[sheet.sheet_disclaimer] = (notes, column_a[notes])
    assert seen[True][0] == 2 and seen[False][0] == 1, seen
    for _, first_gap in seen.values():
        assert first_gap in (None, ""), seen
    assert workbook.header_row(2) != workbook.header_row(1), (
        "the header address must MOVE with the note block; a constant is what R-181 removed")


def test_every_table_name_is_legal_unique_and_has_no_space(built):
    """R-182. Excel's own constraints, enforced rather than discovered — a duplicate or an
    illegal displayName is a repair prompt, not an exception, so the file writes cleanly and
    Excel refuses it."""
    import re as _re
    _, book, _, _ = built
    names = []
    for sheet_name in book.sheetnames:
        for table in book[sheet_name].tables.values():
            names.append(table.displayName)
    assert names, "no sheet carries an Excel Table"
    for name in names:
        assert _re.fullmatch(r"tbl_[A-Za-z0-9_]+", name), name
        assert " " not in name
        assert not _re.fullmatch(r"[A-Za-z]{1,3}[0-9]+", name), f"{name} looks like a cell ref"
    assert len(names) == len(set(names)), f"duplicate table name: {names}"


def test_the_index_inventory_is_itself_a_table(built):
    _, book, _, _ = built
    assert "tbl_Index" in {t.displayName for t in book["Index"].tables.values()}


def test_a_table_never_overlaps_a_worksheet_autofilter(built):
    """R-182 trap 1, stated as the property rather than as the absence of one line. This is
    the fault class that produces "we found a problem with some content", which AC-15.6
    forbids outright."""
    _, book, _, _ = built
    for name in book.sheetnames:
        tab = book[name]
        if tab.tables:
            assert not tab.auto_filter.ref, name


def test_the_navy_header_survives_the_table_style(built):
    """R-182 trap 3. A TableStyleInfo's header band would override the manual fill and the
    file would depend on which Excel applied last. One treatment: the navy stays, stripes off.
    """
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    assert tab.cell(header, 1).fill.fgColor.rgb.endswith("2F4858")
    for table in tab.tables.values():
        assert table.tableStyleInfo.showRowStripes is False


def test_headers_are_unique_and_non_empty_on_every_sheet():
    """R-182 trap 2. A Table makes non-empty MANDATORY rather than merely tidy, and the
    67-column list has several near-collisions — TV/TV (full), the two `covered` columns —
    each of which had to be spelled distinctly."""
    for sheet in workbook._ALL_SHEETS:
        labels = [label for _, label in sheet.columns]
        assert all(label and label.strip() for label in labels), sheet.name
        duplicates = {la for la in labels if labels.count(la) > 1}
        assert not duplicates, f"{sheet.name} has duplicate headers: {duplicates}"


def test_freeze_panes_stays_on_the_sheet_not_on_the_table(built):
    """R-182 trap 4."""
    _, book, _, _ = built
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    from openpyxl.utils import get_column_letter
    column = get_column_letter(schedule.freeze_column())
    assert book["Schedule"].freeze_panes == f"{column}{workbook.first_data_row(1)}"


# --- R-183, the hyperlinks --------------------------------------------------------------

def test_team_and_matchup_cells_link_back_and_carry_the_scope(built):
    """A link that drops the season is the defect that made choosing 2025 and clicking a
    team return a 2026 page. It is worse in a workbook, which is read weeks later."""
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    away = tab.cell(header + 1, labels["Away"])
    assert away.hyperlink is not None, "the team name cell is not linked"
    target = away.hyperlink.target or away.hyperlink.location
    assert target.startswith("https://cfdb.example/team?")
    assert "season=2026" in target and "week=8" in target
    assert away.value and not str(away.value).startswith("http"), (
        "the cell's VALUE must stay the team name, so the column still sorts as a name")


def test_the_matchup_cell_reads_a_word_and_carries_the_url_behind_it(built):
    """Round 3 REVERSES round 2's decision here, and the reversal is Marc's.

    Round 2 put the raw URL in the cell so anything reading the file as DATA could see it —
    a cell hyperlink is invisible to a CSV export or a pivot. He looked at it: a 90-character
    URL in every row of a 56-column sheet costs more width and legibility than that buys, and
    `Game id` beside it already reconstructs the link.

    So the cell says "Matchup" and the URL is the hyperlink.
    """
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    cell = tab.cell(header + 1, labels["Matchup URL"])
    assert cell.value == workbook.URL_CELL_LABEL == "Matchup"
    assert not str(cell.value).startswith("http"), "the URL is the link, not the text"
    target = cell.hyperlink.target or cell.hyperlink.location
    assert target.startswith("https://cfdb.example/matchup?")
    assert "season=2026" in target and "week=8" in target


def test_the_index_links_each_sheet_back_to_its_page_with_the_same_filters(built):
    """Round 3. A workbook read in November should land on the week it covers, not on
    whatever week the site is showing when the link is clicked — which is the same defect
    that made choosing 2025 and clicking a team return a 2026 page."""
    _, book, _, _ = built
    tab = book["Index"]
    found = {}
    for row in tab.iter_rows():
        for cell in row:
            if cell.value in ("Schedule", "srv_game") and cell.hyperlink:
                found[cell.value] = cell.hyperlink.target or cell.hyperlink.location
    assert "Schedule" in found, "the sheet name does not link back to its page"
    assert found["Schedule"].startswith("https://cfdb.example/schedule?")
    assert "season=2026" in found["Schedule"] and "week=8" in found["Schedule"]

    assert "srv_game" in found, "the serving view does not link into the data dictionary"
    # The site's own convention, not a second spelling of it: lib/table.py already writes
    # `/dictionary?table=<name>` for every dataset caption on every page.
    assert found["srv_game"] == "https://cfdb.example/dictionary?table=srv_game"


def test_the_index_links_are_absent_rather_than_broken_with_no_host(monkeypatch):
    monkeypatch.delenv("CFDB_SITE_HOST", raising=False)
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    payload, _, _ = workbook.build(2026, 8, "regular", None, "fbs")
    from openpyxl import load_workbook
    tab = load_workbook(BytesIO(payload))["Index"]
    for row in tab.iter_rows():
        for cell in row:
            assert cell.hyperlink is None, cell.coordinate


def test_with_no_site_host_the_workbook_ships_no_links_rather_than_broken_ones(monkeypatch):
    """R-151's shape: a link that resolves in dev and 404s in the file the user downloaded.
    A workbook full of `http://None/...` is worse than one with none."""
    monkeypatch.delenv("CFDB_SITE_HOST", raising=False)
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    assert workbook.site_base_url() is None
    payload, _, _ = workbook.build(2026, 8, "regular", None, "fbs")
    from openpyxl import load_workbook
    book = load_workbook(BytesIO(payload))
    tab = book["Schedule"]
    for row in tab.iter_rows():
        for cell in row:
            assert cell.hyperlink is None, f"{cell.coordinate} linked with no host set"
            assert "None/" not in str(cell.value), f"{cell.coordinate} = {cell.value!r}"
    index = "\n".join(str(c.value) for r in book["Index"].iter_rows() for c in r)
    assert "no hyperlinks" in index, "the Index must say the links are absent and why"


def test_a_bare_hostname_becomes_https_and_a_full_url_is_left_alone(monkeypatch):
    monkeypatch.setenv("CFDB_SITE_HOST", "cfdb.example.com")
    assert workbook.site_base_url() == "https://cfdb.example.com"
    monkeypatch.setenv("CFDB_SITE_HOST", "http://localhost:8501/")
    assert workbook.site_base_url() == "http://localhost:8501"


def test_the_index_warns_that_access_will_ask_for_a_sign_in(built):
    """Someone who was not expecting Cloudflare Access reads the sign-in page as a broken
    link, and concludes the workbook's links are wrong."""
    _, book, _, _ = built
    text = "\n".join(str(c.value) for r in book["Index"].iter_rows() for c in r)
    assert "Cloudflare Access" in text and "sign in" in text


# --- R-185, the Schedule sheet ------------------------------------------------------------

def test_the_schedule_sheet_carries_marcs_fifty_six_columns_in_his_order():
    """R-214/R-215. The order is `claude_work/supporting_files/cfdb_schedule_column_order.csv`
    and it is AUTHORITATIVE, so this reads the file rather than restating it. A test that
    hardcoded the list would have to be edited alongside the CSV and could disagree with it.
    """
    import csv as _csv
    source = (Path(__file__).resolve().parents[2] / "claude_work" / "supporting_files"
              / "cfdb_schedule_column_order.csv")
    if not source.exists():                      # the CSV lives outside this repo
        pytest.skip("Marc's column-order CSV is not present in this checkout")
    wanted = [workbook.CSV_LABEL_OVERRIDES.get(row["Field"], row["Field"])
              for row in _csv.DictReader(source.open())]
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    assert [label for _, label in schedule.columns] == wanted
    # The override list is not a licence to drift: every entry must still name a real CSV row.
    csv_fields = {row["Field"] for row in _csv.DictReader(source.open())}
    for original in workbook.CSV_LABEL_OVERRIDES:
        assert original in csv_fields, f"{original!r} is not in the CSV at all"


def test_the_eleven_removed_columns_are_gone_from_the_select_as_well_as_the_headers():
    """R-214. Dropping a label but leaving the field in the SELECT would keep paying for the
    column and give a reader no way to see it — and the underlying srv_game columns are
    deliberately untouched, so nothing else notices."""
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    flat = " ".join(schedule.sql.split())
    for field in ("kickoff_time_known", "network,", "spread_open", "over_under_open",
                  "spread_at_close", "spread_at_close_provider", "spread_at_close_basis",
                  "total_at_close", "total_at_close_provider", "total_at_close_basis"):
        assert field not in flat, f"{field} is still selected"
    assert "game_date" not in flat, "the Day column went, and so does what it derived from"


def test_upset_basis_is_gone_and_must_not_come_back():
    """R-193. It was dropped from srv_game the same day the column list was first written,
    and a SELECT naming it is what 500'd the Schedule page. `upset_level` alone now carries
    "judged against the closing line"."""
    for sheet in workbook._ALL_SHEETS:
        assert "upset_basis" not in sheet.sql, sheet.name
        assert "upset_basis" not in sheet.fields, sheet.name


def test_the_delta_columns_lost_their_disambiguator_and_the_index_says_so(built):
    """R-214, consequence 1, and it is a REAL LOSS rather than a tidy-up.

    `Δ Spread` is null both when the line did not move and when no opening number was ever
    recorded — two different facts, one blank cell — and `Spread open` beside it was what
    told them apart. Marc removed both opening columns, so the blank is now genuinely
    ambiguous in the file.

    No sentinel is invented; the Index carries the sentence. This test asserts the sentence
    exists, because it is the only thing standing between a reader and a wrong reading.
    """
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    labels = [label for _, label in schedule.columns]
    assert "Spread open" not in labels and "O/U open" not in labels
    assert "Δ Spread" in labels
    _, book, _, _ = built
    text = "\n".join(str(c.value) for r in book["Index"].iter_rows() for c in r)
    assert "did not move" in text and "no opening line" in text


def test_marcs_order_keeps_the_market_left_of_the_result():
    """His arrangement is scope → fixture → market → result → context → model → keys, and it
    preserves the one property spec §3.2 argued for: the numbers you open the file on a
    Wednesday to read come before the columns that are blank until Saturday night."""
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    labels = [label for _, label in schedule.columns]
    assert labels.index("Spread") < labels.index("Away pts")
    assert labels.index("Season") < labels.index("Away")
    assert labels.index("Attribution") == len(labels) - 1


def test_the_derived_columns_are_words_not_booleans_or_timestamps():
    """`Day` and `Status` exist because a datetime does not answer "is this a Thursday game"
    and a boolean does not answer "has it been played" without the reader translating."""
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    assert schedule.value_for("status", {"is_completed": True}) == "Final"
    assert schedule.value_for("status", {"is_completed": False}) == "Scheduled"
    assert schedule.value_for("status", {"is_completed": None}) is None
    assert "weekday" not in schedule.derived, "the Day column went out with R-214"


def test_the_default_sort_is_the_order_by_and_it_is_stable():
    """A Table's sortState is metadata about a sort that WAS applied; Excel does not
    re-apply it on open. So the ORDER BY is the default sort, and it needs a tiebreak or two
    builds of the same scope differ — which a diff of two workbooks depends on."""
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    flat = " ".join(schedule.sql.split())
    assert "order by start_date_et, game_id" in flat


def test_two_sheets_ship_and_the_other_five_are_kept_not_deleted():
    """Schedule and Scores (R-255). The five are real work and are converted one at a time;
    deleting them would mean rewriting their SQL and column lists from scratch.

    Both halves matter: that the shipped list is exactly what we think, and that nothing
    fell out of `_ALL_SHEETS` on the way. Seven in, seven accounted for.
    """
    assert [s.name for s in workbook.SHEETS] == ["Schedule", "Scores"]
    assert {s.name for s in workbook.PENDING_SHEETS} == {
        "Odds", "Edges", "Standings", "Model performance", "Data dictionary"}
    assert len(workbook.SHEETS) + len(workbook.PENDING_SHEETS) == len(workbook._ALL_SHEETS)


def test_the_index_names_the_six_that_are_not_here_rather_than_dropping_them(built):
    """A sheet a user had yesterday and does not have today is a question. An Index that
    does not mention it makes the workbook look broken instead of narrowed."""
    _, book, _, _ = built
    text = "\n".join(str(c.value) for r in book["Index"].iter_rows() for c in r)
    for sheet in workbook.PENDING_SHEETS:
        assert sheet.name in text, sheet.name
    assert "not converted to the new layout yet" in text


def test_nothing_in_the_file_is_a_fault_excel_would_refuse_to_open(built):
    """AC-15.6 IS A CLAIM ABOUT EXCEL, AND openpyxl OPENING THE FILE DOES NOT PROVE IT.

    openpyxl is forgiving where Excel is not, so this asserts the specific structures Excel
    rejects with "we found a problem with some content" — the prompt AC-15.6 forbids:

      * a table's declared tableColumns disagreeing with the header cells above them
      * a duplicate or empty column name inside a table
      * a table ref that overruns the sheet's used range
      * a table overlapping a worksheet autofilter (R-182 trap 1)
      * NaN or infinity in a cell, which is valid XML and still refused
      * a part that is not well-formed XML

    It still does not replace opening the file in Excel once, by hand. It makes the
    hand-check a confirmation rather than the only evidence.
    """
    import math
    import re
    import zipfile
    from xml.etree import ElementTree
    from openpyxl.utils import range_boundaries

    payload, book, _, _ = built

    with zipfile.ZipFile(BytesIO(payload)) as archive:
        assert archive.testzip() is None
        for name in archive.namelist():
            if name.endswith((".xml", ".rels")):
                ElementTree.fromstring(archive.read(name))   # raises if malformed

    seen_tables = 0
    for name in book.sheetnames:
        tab = book[name]
        for table in tab.tables.values():
            seen_tables += 1
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = [tab.cell(min_row, c).value for c in range(min_col, max_col + 1)]
            declared = [c.name for c in table.tableColumns]
            assert headers == declared, (
                f"{table.displayName}: the table declares {declared} and the header row "
                f"says {headers}. Excel refuses the file when these disagree")
            assert len(set(declared)) == len(declared), table.displayName
            assert all(h not in (None, "") for h in headers), table.displayName
            assert max_row <= tab.max_row and max_col <= tab.max_column, (
                f"{table.displayName}: ref {table.ref} overruns {tab.dimensions}")
            assert not tab.auto_filter.ref, table.displayName
    assert seen_tables >= 2, "Index and Schedule should both carry a Table"

    # NaN IS CHECKED IN THE RAW XML, AND THE REASON IS THE WHOLE POINT OF THIS COMMENT.
    #
    # The first version of this loop walked the RE-LOADED workbook looking for a float that
    # is NaN. It could never fail. openpyxl WRITES NaN — as `<c t="n"><v /></c>`, a numeric
    # cell with no value, which is exactly what Excel objects to — and then READS THAT BACK
    # AS None. So the object model always looks clean no matter what was written, and the
    # assertion was unfalsifiable: green with `_clean`'s guards removed entirely.
    #
    # Verified by removing both guards and watching this stay green, which is how it was
    # caught. The bytes are the only place the fault is visible.
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        for name in archive.namelist():
            if not name.startswith("xl/worksheets/"):
                continue
            body = archive.read(name).decode("utf-8")
            empty_numeric = re.findall(r'<c[^>]*t="n"[^>]*>\s*<v\s*/>', body)
            assert not empty_numeric, (
                f"{name} has {len(empty_numeric)} numeric cell(s) with no value — that is "
                f"how a NaN or an infinity reaches the file, and Excel refuses it")
    assert math is not None


# === R-216 · R-217 · R-218 · R-219 · R-220: round two ====================================

def test_no_label_shaped_number_renders_with_a_decimal_or_a_comma():
    """R-216. THE RULE, not the list.

    A thousands separator is for a quantity you might total. A rank, an id, a season and a
    week are labels that happen to be numeric, and a comma in one is a bug. `#,##0` rendered
    season 2025 as "2,025" — which is how a membership tuple that had `season` in it was
    still wrong about it.

    Enumerates the REAL columns rather than re-implementing the pattern. That matters:
    `best_rank_in_game` carries "rank" in the MIDDLE of its name and the suffix rule does not
    see it, so a test written from the same pattern would have agreed with the bug.
    """
    offenders = []
    for sheet in workbook._ALL_SHEETS:
        for field, label in sheet.columns:
            looks_like_a_label = ("rank" in field or field.endswith("_id")
                                  or field in ("season", "week"))
            if not looks_like_a_label:
                continue
            rendered = workbook.number_format(field)
            if "." in rendered or "," in rendered:
                offenders.append((sheet.name, field, rendered))
    assert not offenders, offenders


def test_counts_keep_their_thousands_separator():
    """The other half of the rule. Dropping commas everywhere would make 74109 unreadable,
    and attendance is exactly the quantity a reader might total."""
    assert workbook.number_format("attendance") == "#,##0"
    assert workbook.number_format("total_points") == "#,##0"
    assert workbook.number_format("season") == "0"
    assert workbook.number_format("game_id") == "0"
    assert workbook.number_format("best_rank_in_game") == "0"


def test_column_width_is_measured_from_the_data_not_from_the_header(built):
    """R-217. The old code seeded the width with `len(label)`, so "Conference game" set a
    15-wide column over three characters of data. Marc found four of them.

    Asserted on the columns he named, with the data the fixture holds — a long header over
    short data must now be narrow, and the header wraps instead.
    """
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    from openpyxl.utils import get_column_letter
    for label in ("Current week", "Conference game", "Home rank", "Best rank"):
        letter = get_column_letter(labels[label])
        width = tab.column_dimensions[letter].width
        assert width < len(label), (
            f"{label!r} is {width} wide for a {len(label)}-character header — the label is "
            f"still setting the width")
        assert width >= workbook.MIN_COLUMN_WIDTH, (
            f"{label!r} is {width} wide; a one-character column is unreadable")


def test_the_header_row_wraps_and_its_height_is_computed_not_hardcoded(built):
    """R-217. Marc's macro uses 38pt, which is three lines and a good default — but it is a
    default that FITS TODAY'S HEADERS, and the next long one would silently clip.

    So the height is measured from the labels at their final widths. Asserted as a
    relationship, not a number: a test pinning 38 would be the hardcode with extra steps.
    """
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    assert tab.cell(header, 1).alignment.wrap_text is True
    height = tab.row_dimensions[header].height
    assert height and height >= workbook.MIN_HEADER_HEIGHT

    from openpyxl.utils import get_column_letter
    worst = 0
    for index, (_, label) in enumerate(
            next(s for s in workbook.SHEETS if s.name == "Schedule").columns, start=1):
        width = tab.column_dimensions[get_column_letter(index)].width
        worst = max(worst, workbook._header_height([(label, width)]))
    # THE FLOOR IS A FLOOR, NOT A FIXED HEIGHT. Marc's 50pt gives the labels room above the
    # Table's filter buttons, which sit inside the header cell and overlap the last line. But
    # a longer header in some future column must still be able to push the row taller, or
    # R-217's "computing cannot clip" is given straight back.
    assert height == max(worst, workbook.MIN_HEADER_HEIGHT), (
        f"the row is {height}pt, the tallest label needs {worst}pt and the floor is "
        f"{workbook.MIN_HEADER_HEIGHT}pt")


def test_a_longer_header_makes_the_row_taller_which_a_hardcoded_38_could_not():
    """The negative half. A fixed height cannot respond to a longer label; a computed one
    must, or it is a hardcode wearing a function's clothes."""
    short = workbook._header_height([("Wk", 6)])
    long = workbook._header_height([("Margin (away−home)", 6)])
    assert long > short
    # A single word wider than the column still has to go somewhere: Excel breaks it rather
    # than clipping, so the calculation must account for that instead of assuming it fits.
    assert workbook._header_height([("Precipitation", 6)]) > short


def test_booleans_read_as_words_and_nulls_stay_null():
    """R-218. Marc marked six fields t/f = Yes/No. A null is NOT False — this project has
    been bitten three times by truthiness on a pandas null."""
    assert workbook._yes_no(True) == "Yes"
    assert workbook._yes_no(False) == "No"
    assert workbook._yes_no(None) is None
    assert workbook._yes_no(float("nan")) is None


def test_no_favorite_becomes_two_words_not_an_underscore():
    """`favorite_covered` is already a string, so this is title-casing, not a boolean
    conversion — and a naive `.title()` leaves "No_Favorite" behind."""
    assert workbook._title_case_verdict("no_favorite") == "No favorite"
    assert workbook._title_case_verdict("push") == "Push"
    assert workbook._title_case_verdict(None) is None


def test_the_flag_highlight_follows_what_the_cell_actually_holds(built):
    """R-218's TRAP, and it is the kind that passes every test written against Python.

    `CellIsRule(operator="equal", formula=["TRUE"])` matches nothing once the cell holds the
    string "Yes" — no error, no warning, the highlight just disappears. So the formula is
    derived from whether the column has a display rule, and asserted on the built file.
    """
    _, book, _, _ = built
    schedule = next(s for s in workbook.SHEETS if s.name == "Schedule")
    rules = book["Schedule"].conditional_formatting
    formulas = [f for rule_set in rules for rule in rule_set.rules
                for f in (rule.formula or [])]
    assert formulas, "the Schedule sheet has no conditional formatting at all"
    for field in workbook.FLAG_FIELDS:
        if field not in schedule.fields:
            continue
        expected = '"Yes"' if field in schedule.display else "TRUE"
        assert expected in formulas, (
            f"{field} renders as {'Yes/No' if field in schedule.display else 'TRUE/FALSE'} "
            f"but no rule compares to {expected}")


def test_the_verdict_columns_use_the_sites_marks_and_a_dash_for_nothing():
    """R-219. Geometric shapes, not emoji: R-141 and R-175 both turned on an
    emoji-presentation character having no fixed baseline or size across platforms, and a
    workbook is opened on more platforms than a page is."""
    for marks in (workbook.UPSET_MARKS, workbook.COVER_MARKS, workbook.OVER_MARKS):
        for glyph in marks.values():
            # A mark may REPEAT to carry a level (●, ●●, ●●●) but it is always repetitions
            # of ONE character — two different shapes in one cell would be a new mark rather
            # than a stronger one.
            assert 1 <= len(glyph) <= 3, glyph
            assert len(set(glyph)) == 1, f"{glyph!r} mixes shapes"
            # Emoji live above U+1F000, and the presentation selector is U+FE0F. Neither
            # may appear here.
            assert ord(glyph[0]) < 0x1F000, f"{glyph!r} is emoji-presentation"
            assert "️" not in glyph
    render = workbook._marked(workbook.COVER_MARKS)
    assert render("yes") == "■" and render("push") == workbook.PUSH_MARK
    assert render(None) == workbook.NO_DATA_MARK
    assert render(float("nan")) == workbook.NO_DATA_MARK
    assert render("something_new") == workbook.NO_DATA_MARK


def test_the_index_carries_a_legend_for_every_mark_it_uses(built):
    """R-219, and this is NOT OPTIONAL. R-026's icon-only exception on the site is
    defensible because R-102's legend explains it once. A workbook has no tooltip at all, so
    the same exception needs the same support or it is undecodable symbols."""
    _, book, _, _ = built
    text = "\n".join(str(c.value) for r in book["Index"].iter_rows() for c in r)
    used = set()
    for marks in (workbook.UPSET_MARKS, workbook.COVER_MARKS, workbook.OVER_MARKS):
        used.update(marks.values())
    used.add(workbook.NO_DATA_MARK)
    for glyph in used:
        assert glyph in text, f"{glyph!r} is rendered in the sheet and absent from the legend"
    assert 'SUM(--(range="Yes"))' in text, "the Yes/No formula cost must be stated once"


def test_the_data_bar_carries_marcs_full_rule_in_the_x14_extension(built):
    """R-220, READ FROM THE RAW XML BECAUSE openpyxl CANNOT SEE IT.

    openpyxl's DataBarRule writes the legacy `<dataBar>`, which has one colour, a min and a
    max — no negative fill, no borders, no axis position. Three of the seven things in
    Marc's screenshot cannot be said with it. Those live in `x14:dataBar`, which openpyxl
    does not model and DISCARDS ON READ, with the warning "Unknown extension is not
    supported and will be removed".

    So a test that re-loaded the workbook would find nothing and could never fail. The bytes
    are the only place this is visible — the same lesson as the NaN check above.
    """
    import re
    import zipfile as _zip
    payload, _, _ = built[0], None, None
    payload = built[0]
    with _zip.ZipFile(BytesIO(payload)) as archive:
        sheets = [n for n in archive.namelist() if n.startswith("xl/worksheets/sheet")]
        body = "".join(archive.read(n).decode("utf-8") for n in sheets)

    assert "x14:dataBar" in body, "the extension block is missing entirely"
    assert 'axisPosition="middle"' in body, "Marc's rule says axis position MIDPOINT"
    assert f'x14:negativeFillColor rgb="{workbook.DATA_BAR_NEGATIVE}"' in body, "negative RED"
    assert f'x14:borderColor rgb="{workbook.DATA_BAR_AXIS}"' in body, "solid black border"
    assert f'x14:negativeBorderColor rgb="{workbook.DATA_BAR_AXIS}"' in body
    assert f'x14:axisColor rgb="{workbook.DATA_BAR_AXIS}"' in body
    assert 'gradient="0"' in body, "Marc's rule says SOLID fill, not gradient"

    # The legacy element and the extension are tied by a GUID, and a mismatch means Excel
    # draws the plain bar and silently ignores everything Marc asked for.
    legacy = set(re.findall(r"<x14:id>([^<]+)</x14:id>", body))
    extended = set(re.findall(r'<x14:cfRule type="dataBar" id="([^"]+)"', body))
    assert legacy and legacy == extended, (legacy, extended)


def test_the_data_bar_guid_is_stable_across_builds(monkeypatch):
    """Two builds of one scope must be byte-identical, which a diff of two workbooks depends
    on. A random GUID would break that quietly."""
    assert workbook._data_bar_guid(1) == workbook._data_bar_guid(1)
    assert workbook._data_bar_guid(1) != workbook._data_bar_guid(2)


def test_every_sheet_part_is_still_well_formed_after_the_xml_injection(built):
    """Hand-written XML is the class of change that produces the repair prompt AC-15.6
    forbids, so the rewritten parts are parsed rather than trusted."""
    import zipfile as _zip
    from xml.etree import ElementTree
    payload = built[0]
    with _zip.ZipFile(BytesIO(payload)) as archive:
        assert archive.testzip() is None
        for name in archive.namelist():
            if name.endswith((".xml", ".rels")):
                ElementTree.fromstring(archive.read(name))


def test_a_postgres_text_boolean_does_not_become_yes(monkeypatch):
    """'f' IS A NON-EMPTY STRING AND `bool('f')` IS TRUE.

    Postgres renders a boolean as 't'/'f' the moment anything reads it as text — a CSV
    export, a driver without type mapping, an object-dtype column. So a blanket `bool(value)`
    turns every False into "Yes", silently, on a column a reader filters on.

    Found by building a real workbook from real rows and reading row 1: a game the favourite
    won by 14 while laying 15 was showing "Upset by line: Yes". Same family as the three
    NaN-truthiness bugs this project has already had.
    """
    assert workbook._yes_no("f") == "No"
    assert workbook._yes_no("false") == "No"
    assert workbook._yes_no("t") == "Yes"
    assert workbook._yes_no(False) == "No"
    # Not a boolean at all: guessing is what caused this, so it declines to guess.
    assert workbook._yes_no("maybe") is None


def test_no_two_marks_are_visually_confusable():
    """"The favourite won" and "cfdb holds no closing line" are OPPOSITE CLAIMS.

    The first draft drew them as an em dash and an en dash, which at 11pt are the same
    picture. Every mark in use must be a distinct character, and the dash-like ones must not
    collide.
    """
    used = []
    for marks in (workbook.UPSET_MARKS, workbook.COVER_MARKS, workbook.OVER_MARKS):
        used.extend(marks.values())
    used.append(workbook.NO_DATA_MARK)

    dashes = {"-", "‐", "‑", "‒", "–", "—", "―", "−"}
    dash_marks = [m for m in used if m in dashes]
    assert len(set(dash_marks)) <= 1, (
        f"more than one dash-like mark is in use and they are indistinguishable: "
        f"{sorted(set(dash_marks))}")
    assert workbook.UPSET_MARKS["none"] not in dashes, (
        "'the favourite won' must not be a dash; the dash means cfdb holds nothing")


def test_filled_means_it_happened_and_open_means_it_did_not():
    """R-141's shape-plus-fill system, carried into the sheet. It is what makes a mark
    self-identifying without colour, which matters more in Excel than on the site."""
    assert workbook.UPSET_MARKS["upset"] == "●" and workbook.UPSET_MARKS["none"] == "○"
    assert workbook.COVER_MARKS["yes"] == "■" and workbook.COVER_MARKS["no"] == "□"


def test_the_legend_names_only_the_columns_that_actually_use_marks():
    """A legend that promises a shape in a column of words sends the reader looking for
    something that is not there. `Favourite covered` renders as Yes/No/Push per Marc's CSV,
    so it is described as a word rather than listed among the marks."""
    schedule = next(s for s in workbook.SHEETS if s.name == "Schedule")
    marked = {field for field, renderer in schedule.display.items()
              if renderer.__qualname__.startswith("_marked")}
    labels = dict(schedule.columns)
    marked_labels = {labels[f] for f in marked}
    assert marked_labels == {"Upset level", "Winner covered", "O/U result"}, marked_labels
    for column, _, _ in workbook.mark_legend():
        if column.startswith("Any"):
            continue
        assert column in marked_labels, (
            f"the legend lists {column!r} among the marks, and that column renders words")


def test_every_sheet_part_is_in_ct_worksheet_order(built):
    """THE CHECK THAT WAS MISSING, AND THE FILE SHIPPED BROKEN BECAUSE OF IT.

    `CT_Worksheet`'s children are an ORDERED SEQUENCE. Excel validates it; openpyxl does
    not, and neither does "is this well-formed XML" — the file parsed perfectly and Excel
    still refused it, replaced the whole sheet part and opened the Schedule tab EMPTY.

    The R-220 injection put its `<conditionalFormatting>` block before `<pageMargins>`,
    which on a sheet with hyperlinks lands AFTER `<hyperlinks>`. The schema puts
    conditionalFormatting first. Every prior check passed: well-formed XML, tables matching
    their headers, no NaN, no autofilter overlap. Order was the one property nothing tested.
    """
    import zipfile as _zip
    payload = built[0]
    with _zip.ZipFile(BytesIO(payload)) as archive:
        checked = 0
        for name in archive.namelist():
            if not name.startswith("xl/worksheets/sheet"):
                continue
            checked += 1
            violations = workbook.sheet_order_violations(archive.read(name))
            assert not violations, (
                f"{name}: {violations} — Excel will refuse this file and open the tab empty")
        assert checked >= 2


def test_the_order_validator_actually_detects_a_misplaced_block():
    """A validator that cannot fail is the thing that let this through in the first place.
    Fed the exact shape that broke Excel, it must object."""
    good = (b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b'<sheetData/><conditionalFormatting sqref="A1"/><hyperlinks/><pageMargins/>'
            b'</worksheet>')
    bad = (b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
           b'<sheetData/><hyperlinks/><conditionalFormatting sqref="A1"/><pageMargins/>'
           b'</worksheet>')
    assert workbook.sheet_order_violations(good) == []
    assert workbook.sheet_order_violations(bad) == [("conditionalFormatting", "hyperlinks")]


def test_the_build_refuses_to_emit_a_sheet_excel_would_reject(monkeypatch):
    """Fail at build time, not in the user's Excel. Verified by forcing the injection to the
    wrong place and asserting `build` raises rather than returning bytes."""
    monkeypatch.setenv("CFDB_SITE_HOST", "https://cfdb.example")
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    # Force the fallback path AND point it at the exact wrong place — before <pageMargins>,
    # which on a sheet with hyperlinks lands after them. This is the shape that shipped.
    monkeypatch.setattr(workbook, "CF_ANCHOR", "<<no such element>>")
    monkeypatch.setattr(workbook, "SHEET_ELEMENTS_AFTER_CF", ("<pageMargins", "</worksheet>"))
    with pytest.raises(RuntimeError) as excinfo:
        workbook.build(2026, 8, "regular", None, "fbs")
    assert "CT_Worksheet order" in str(excinfo.value)


def test_the_simple_data_bar_fallback_writes_no_extension_and_stays_in_order(monkeypatch):
    """R-220's route (b), kept reachable rather than kept in a comment.

    The x14 route fought back once and cost a broken file, so the simpler rendering is one
    environment variable away and is tested — a fallback nobody exercises is a fallback that
    does not work when it is needed.
    """
    import zipfile as _zip
    monkeypatch.setenv("CFDB_SITE_HOST", "https://cfdb.example")
    monkeypatch.setattr(workbook, "DATA_BAR_SIMPLE", True)
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    payload, _, _ = workbook.build(2026, 8, "regular", None, "fbs")
    with _zip.ZipFile(BytesIO(payload)) as archive:
        body = "".join(archive.read(n).decode("utf-8") for n in archive.namelist()
                       if n.startswith("xl/worksheets/sheet"))
        for name in archive.namelist():
            if name.startswith("xl/worksheets/sheet"):
                assert not workbook.sheet_order_violations(archive.read(name)), name
    assert "x14:dataBar" not in body, "the fallback must not write the extension"
    assert "dataBar" in body, "but it must still draw a bar"


# === round three: legibility ==============================================================

def test_a_column_is_never_narrower_than_the_longest_word_in_its_header(built):
    """Round 3. R-217 measured the data and went too far: at a flat floor of 6, Excel breaks
    "Favourite" into "Favouri / te", which is harder to read than the wide column it
    replaced. Marc: "shrinking is a little aggressive".

    The floor is the longest WORD now, capped. "Conference game" lands at 10 — not the 15
    that started this, and not the 6 that broke the word.
    """
    from openpyxl.utils import get_column_letter
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    for index in range(1, tab.max_column + 1):
        label = str(tab.cell(header, index).value)
        width = tab.column_dimensions[get_column_letter(index)].width
        if label in workbook.WIDTH_OVERRIDES:
            # A hand-set width wins, but not below its own header word — see
            # test_a_hand_set_width_is_still_floored_at_its_header_word.
            assert width == workbook.effective_width(label), label
            continue
        longest_word = max((len(w) for w in label.split()), default=0)
        expected_floor = max(workbook.MIN_COLUMN_WIDTH,
                             min(longest_word, workbook.HEADER_WORD_CAP))
        assert width >= expected_floor, (
            f"{label!r} is {width} wide; its longest word is {longest_word} and will break")
        # ...and the change must not have undone R-217. A column is still not as wide as its
        # whole header just because the header is long.
        if len(label) > workbook.HEADER_WORD_CAP and " " in label:
            assert width < len(label), f"{label!r} is back to fitting its entire header"


def test_the_header_is_top_aligned_and_centred(built):
    """Marc, round 3. Vertical TOP matters more than it sounds: with a wrapped header the row
    is as tall as the WORST label, so a vertically centred one-line header floats in the
    middle of a four-line row and no two headers share a baseline."""
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    for index in range(1, tab.max_column + 1):
        alignment = tab.cell(header, index).alignment
        assert alignment.vertical == "top", index
        assert alignment.horizontal == "center", index
        assert alignment.wrap_text is True, index


def test_the_mark_columns_are_centred_and_the_others_are_not(built):
    """A one-character mark left-aligned in a 6-wide column reads as a typo."""
    _, book, _, _ = built
    schedule = next(s for s in workbook.SHEETS if s.name == "Schedule")
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    centred_labels = {dict(schedule.columns)[f] for f in schedule.centred}
    assert centred_labels == {"Upset level", "Winner covered", "O/U result"}
    for label in centred_labels:
        cell = tab.cell(header + 1, labels[label])
        assert cell.alignment.horizontal == "center", label
    assert tab.cell(header + 1, labels["Away"]).alignment.horizontal != "center"


def test_the_open_marks_are_coloured_and_the_colour_passes_contrast():
    """Marc asked for the OPEN form to carry the colour, and he is right about why: the
    filled shapes are the common case, so colouring the exception is what makes it pop.

    LITERAL BURNT SIENNA FAILS. #E97451 measures 2.97:1 against white — below WCAG AA's 4.5.
    The shipped colour is measured, because this project has already had to fix a glyph that
    was 3.6:1.
    """
    def luminance(hex_rgb):
        channels = [int(hex_rgb[i:i + 2], 16) / 255 for i in (0, 2, 4)]
        channels = [c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4
                    for c in channels]
        return 0.2126 * channels[0] + 0.7152 * channels[1] + 0.0722 * channels[2]

    rgb = workbook.OPEN_MARK_COLOUR[2:]          # strip the alpha openpyxl wants
    contrast = (1.0 + 0.05) / (luminance(rgb) + 0.05)
    assert contrast >= 4.5, f"#{rgb} is {contrast:.2f}:1 against white"
    assert luminance(rgb) < luminance("E97451"), (
        "the shipped colour must be darker than literal burnt sienna, which fails")


def test_the_legend_glyphs_match_the_sheet_in_size_and_colour(built):
    """A legend whose mark is a different size or colour from the one in the column is a
    picture of a DIFFERENT mark, which is worse than no legend."""
    _, book, _, _ = built
    tab = book["Index"]
    seen = {}
    for row in tab.iter_rows():
        for cell in row:
            if cell.value in workbook.COLOURED_MARKS or cell.value in ("●", "■", "▲"):
                seen[cell.value] = cell.font
    assert seen, "no legend glyphs found on the Index"
    # 12 IS MARC'S NUMBER, asserted as the requirement rather than against the constant.
    # The first version compared to MARK_FONT_SIZE, so lowering the constant moved the test
    # with it and the check stayed green — a test that restates the implementation.
    assert workbook.MARK_FONT_SIZE >= 12, workbook.MARK_FONT_SIZE
    for glyph, font in seen.items():
        assert font.size >= 12, (glyph, font.size)
        if glyph in workbook.MARK_COLOURS:
            assert font.color is not None and \
                font.color.rgb == workbook.MARK_COLOURS[glyph], (glyph, font.color)


def test_the_open_marks_in_the_sheet_carry_the_same_colour(built):
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    coloured = 0
    for row in range(header + 1, tab.max_row + 1):
        for label in ("Upset level", "Winner covered", "O/U result"):
            cell = tab.cell(row, labels[label])
            if cell.value in workbook.MARK_COLOURS:
                assert cell.font.color.rgb == workbook.MARK_COLOURS[cell.value], \
                    cell.coordinate
                assert cell.font.size >= 12
                coloured += 1
    assert coloured, "no open mark appeared in the fixture, so this proved nothing"


# === round four ===========================================================================

def test_low_cardinality_text_is_centred_and_prose_is_not(monkeypatch):
    """Marc: "for any text fields with cardinality <5 make them center aligned".

    A frame with ENOUGH ROWS TO JUDGE, because the rule reads the shape of the data and the
    two-row fixture cannot show it: with two rows every text column looks like a category,
    team names included.
    """
    import pandas as _pd
    rows = workbook.MIN_ROWS_TO_JUDGE_CARDINALITY + 8
    frame = _pd.DataFrame({
        "status": ["Final", "Scheduled"] * (rows // 2),                  # 2 distinct
        "season_type": ["regular"] * rows,                               # 1
        "home_team_display": [f"Team {i}" for i in range(rows)],         # all distinct
        "venue_display": [f"Stadium {i}" for i in range(rows)],          # all distinct
        "weather_condition": ["Clear", "Rain", "Cloudy", "Fog", "Snow"] * (rows // 5),
    })
    found = workbook._low_cardinality_text(frame, list(frame.columns))
    assert "status" in found and "season_type" in found
    assert "home_team_display" not in found, "a team name is prose, not a category"
    assert "venue_display" not in found
    assert "weather_condition" not in found, "five distinct values is NOT fewer than five"


def test_the_cardinality_rule_declines_to_judge_a_short_export():
    """A one-game export must not come out looking nothing like a fifty-game one. Below the
    threshold every text column is trivially low-cardinality and the rule abstains."""
    import pandas as _pd
    tiny = _pd.DataFrame({"home_team_display": ["Rice", "Auburn"]})
    assert workbook._low_cardinality_text(tiny, ["home_team_display"]) == set()


def test_nulls_do_not_count_towards_cardinality():
    """A column of Yes / No / blank is two categories. Counting the blank would push a
    three-value column over the line for no reason a reader would recognise."""
    import pandas as _pd
    rows = workbook.MIN_ROWS_TO_JUDGE_CARDINALITY + 4
    frame = _pd.DataFrame({"verdict": (["Yes", "No", None, "Push"] * rows)[:rows]})
    assert "verdict" in workbook._low_cardinality_text(frame, ["verdict"])


def test_the_header_row_clears_the_tables_filter_buttons(built):
    """Marc: the header needs "enough room to be above the drop-down filters". An Excel Table
    puts a filter button INSIDE the header cell, and it overlaps the last line of a wrapped
    label."""
    _, book, _, _ = built
    tab = book["Schedule"]
    assert tab.row_dimensions[workbook.header_row(1)].height >= 50


def test_the_fifty_point_floor_never_shortens_a_header_that_needs_more(monkeypatch):
    """The floor must not become a FIXED height — that is R-217's clipping defect returning
    by another route.

    Asserted on a BUILT FILE, not only on the computation. The first version tested
    `_header_height` alone and stayed green when the build was changed to write
    MIN_HEADER_HEIGHT unconditionally: the function was still right and the workbook was
    still wrong. Nothing in the fixture needs more than 50pt, so this makes a sheet that does.
    """
    tall = workbook._header_height([("Supercalifragilistic Expialidocious Header", 6)])
    assert tall > workbook.MIN_HEADER_HEIGHT

    monkeypatch.setenv("CFDB_SITE_HOST", "https://cfdb.example")
    monkeypatch.setattr(workbook, "query", _division_aware_query())
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    stretched = workbook.Sheet(
        "Schedule", schedule.view, schedule.sql,
        [(schedule.columns[0][0], "Supercalifragilistic Expialidocious Header")]
        + list(schedule.columns[1:]),
        has_predictions=schedule.has_predictions,
        sheet_disclaimer=schedule.sheet_disclaimer,
        derived=schedule.derived, display=schedule.display,
        link_fields=schedule.link_fields, freeze_before=schedule.freeze_before)
    monkeypatch.setattr(workbook, "SHEETS", [stretched])
    payload, _, _ = workbook.build(2026, 8, "regular", None, "fbs")
    from openpyxl import load_workbook
    tab = load_workbook(BytesIO(payload))["Schedule"]
    height = tab.row_dimensions[workbook.header_row(1)].height
    assert height > workbook.MIN_HEADER_HEIGHT, (
        f"a header needing {tall}pt was written at {height}pt — the floor has become a "
        f"ceiling and long headers will clip")


def test_the_upset_mark_repeats_once_per_level():
    """Marc, round 4. The site distinguishes levels by SHADE, which a filter dropdown cannot
    carry — "show me every blowout" would mean picking a colour. Repetition keeps the
    ordering, sorts correctly (●● after ●) and gives each level its own dropdown entry."""
    assert workbook.UPSET_MARKS["upset"] == "●"
    assert workbook.UPSET_MARKS["big"] == "●●"
    assert workbook.UPSET_MARKS["blowout"] == "●●●"
    assert workbook.UPSET_MARKS["none"] == "○"
    levels = ["upset", "big", "blowout"]
    lengths = [len(workbook.UPSET_MARKS[k]) for k in levels]
    assert lengths == sorted(lengths) == [1, 2, 3], "the mark must grow WITH the level"
    assert sorted(workbook.UPSET_MARKS[k] for k in levels) == ["●", "●●", "●●●"], (
        "repetition must sort in level order, which is half the reason for it")


def test_every_repeated_mark_is_in_the_legend(built):
    """Two circles is a different value from one in the filter dropdown, so it needs its own
    line. A legend that explains ● and leaves the reader to infer ●●● is not a legend."""
    _, book, _, _ = built
    text = "\n".join(str(c.value) for r in book["Index"].iter_rows() for c in r)
    for glyph in set(workbook.UPSET_MARKS.values()):
        assert f"\n{glyph}\n" in f"\n{text}\n" or glyph in text, glyph
    legend_marks = {mark for _, mark, _ in workbook.mark_legend()}
    for glyph in workbook.UPSET_MARKS.values():
        assert glyph in legend_marks, f"{glyph} is rendered and not explained"


def test_the_hand_set_widths_are_exactly_what_marc_measured(built):
    """Six columns he sized himself in Excel's width dialog with the real file open. That is
    better information than the measurement has, so it wins for these and nothing else."""
    from openpyxl.utils import get_column_letter
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    assert workbook.WIDTH_OVERRIDES == {
        "Kickoff": 11.5, "Winner covered": 8.0, "Final margin": 5.85, "Season": 5.6,
        "Wind mph": 5.5, "Pred margin": 6.5, "Home win prob": 7.7}
    for label in workbook.WIDTH_OVERRIDES:
        letter = get_column_letter(labels[label])
        assert tab.column_dimensions[letter].width == workbook.effective_width(label), label


def test_a_hand_set_width_is_still_floored_at_its_header_word():
    """Marc set Season to 5.6 and then reported it "isn't wide enough... doesn't look like it
    was touched". IT WAS TOUCHED — written at exactly 5.6 — and 5.6 is narrower than the word
    "Season". Excel wraps mid-word, so it reads "Seaso / n" and looks untouched rather than
    narrow. A width measured on a column whose header already fits cannot anticipate that.

    Three of the six are raised, each by a fraction; the other three are honoured exactly.
    """
    assert workbook.effective_width("Season") > 5.6
    assert workbook.effective_width("Final margin") > 5.85
    assert workbook.effective_width("Pred margin") > 6.5
    assert workbook.effective_width("Winner covered") == 8.0
    assert workbook.effective_width("Wind mph") == 5.5
    assert workbook.effective_width("Home win prob") == 7.7
    for label in workbook.WIDTH_OVERRIDES:
        longest_word = max(len(w) for w in label.split())
        assert workbook.effective_width(label) >= longest_word, label


def test_winner_covered_is_centred_as_well_as_widened(built):
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    index = labels["Winner covered"]
    assert tab.cell(header, index).alignment.horizontal == "center"
    assert tab.cell(header + 1, index).alignment.horizontal == "center"


def test_an_away_home_pair_is_aligned_the_same_way():
    """Measured on 83 real games: `Home record` came out centred and `Away record` left,
    because that week the home side held four distinct records and the away side five.

    Two identical columns aligned differently reads as a defect, and it is a coin flip that
    lands the other way next week. A pair agrees, on prose — the safe default.
    """
    import pandas as _pd
    rows = workbook.MIN_ROWS_TO_JUDGE_CARDINALITY + 8
    frame = _pd.DataFrame({
        "away_team_record_display": [f"{i}-0" for i in range(rows)],       # many
        "home_team_record_display": ["1-0", "0-1"] * (rows // 2),          # few
    })
    fields = list(frame.columns)
    found = workbook._low_cardinality_text(frame, fields)
    assert found == set(), (
        f"the pair disagreed: {found} centred while its sibling was not")

    # And a genuinely low-cardinality pair still centres, both halves.
    both = _pd.DataFrame({
        "away_classification": ["fbs", "fcs"] * (rows // 2),
        "home_classification": ["fbs"] * rows,
    })
    assert workbook._low_cardinality_text(both, list(both.columns)) == {
        "away_classification", "home_classification"}


def test_text_columns_tell_excel_the_text_is_deliberate(built):
    """A won-lost record is "1-0". Excel sees text that could be a number or a date, puts a
    green corner on every cell and offers to convert it — which would turn 1-0 into
    1 January. Correct advice in general, wrong here, and 166 green triangles across two
    columns is noise a reader has to learn to ignore.

    Read from the RAW XML: openpyxl models `IgnoredErrors` and never writes it — `Worksheet`
    has no `ignored_errors` attribute — so it goes in through the same injection as the data
    bar, and a test that re-loaded the file would find nothing.
    """
    import re
    import zipfile as _zip
    payload = built[0]
    with _zip.ZipFile(BytesIO(payload)) as archive:
        body = "".join(archive.read(n).decode("utf-8") for n in archive.namelist()
                       if n.startswith("xl/worksheets/sheet"))
    assert "<ignoredErrors>" in body, "Excel will keep flagging the record columns"
    assert 'numberStoredAsText="1"' in body
    assert 'twoDigitTextYear="1"' in body

    # And it must cover the columns that actually hold the text, not an arbitrary range.
    tab = built[1]["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    from openpyxl.utils import get_column_letter
    covered = set()
    for span in re.findall(r'<ignoredError sqref="([^"]+)"', body):
        covered.add(span.split(":")[0].rstrip("0123456789"))
    for label in ("Away record", "Home record"):
        letter = get_column_letter(labels[label])
        assert letter in covered, f"{label} (column {letter}) is still flagged"


def test_the_ignored_errors_block_is_in_schema_order(built):
    """`ignoredErrors` sits between `cellWatches` and `smartTags` in CT_Worksheet. Putting it
    anywhere else is the repair prompt again — the same failure the data bar already caused
    once, from a different element."""
    import zipfile as _zip
    payload = built[0]
    with _zip.ZipFile(BytesIO(payload)) as archive:
        for name in archive.namelist():
            if name.startswith("xl/worksheets/sheet"):
                assert not workbook.sheet_order_violations(archive.read(name)), name


def _sheet_parts(payload):
    """{sheet name -> worksheet XML}, resolved through the workbook's relationships.

    Not `sheet1.xml is the first tab` — openpyxl does not guarantee that, and the data-bar
    injector learned the same lesson the hard way.
    """
    import zipfile as _zip
    from xml.etree import ElementTree
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
          "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
          "pr": "http://schemas.openxmlformats.org/package/2006/relationships"}
    out = {}
    with _zip.ZipFile(BytesIO(payload)) as archive:
        book = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {r.get("Id"): r.get("Target")
                   for r in rels.findall("pr:Relationship", ns)}
        for element in book.findall("m:sheets/m:sheet", ns):
            target = (targets.get(element.get(f"{{{ns['r']}}}id")) or "").lstrip("/")
            if not target:
                continue
            part = target if target.startswith("xl/") else "xl/" + target
            out[element.get("name")] = archive.read(part).decode("utf-8")
    return out


def test_a_numeric_column_is_not_told_to_ignore_text_errors(built):
    """The suppression is scoped to columns that HOLD text. Blanket-ignoring the whole sheet
    would also silence the warning on a column where text really is a mistake.

    PER SHEET, AND THAT IS THE FIX. The first version pooled every `ignoredError` range in
    the file into one set of column LETTERS and checked Schedule's columns against it —
    correct while one sheet shipped, and wrong the moment a second did: column AJ is
    `Attendance` on Schedule and a text column on Scores, so Scores' legitimate suppression
    was read as Schedule's mistake.

    Generalised while fixing it. Every shipped sheet is checked, and the rule is stated as
    the property rather than a list of three column names: a column whose first data cell is
    a number must not be in that sheet's suppression set.
    """
    import re
    from openpyxl.utils import get_column_letter
    payload, book, _, _ = built
    parts = _sheet_parts(payload)

    checked = 0
    for sheet in workbook.SHEETS:
        if sheet.name not in book.sheetnames:
            continue
        covered = {span.split(":")[0].rstrip("0123456789")
                   for span in re.findall(r'<ignoredError sqref="([^"]+)"',
                                          parts[sheet.name])}
        assert covered, f"{sheet.name} suppresses nothing — the injection did not happen"
        tab = book[sheet.name]
        first_data = workbook.first_data_row(2 if sheet.sheet_disclaimer else 1)
        for index in range(1, tab.max_column + 1):
            value = tab.cell(first_data, index).value
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                letter = get_column_letter(index)
                assert letter not in covered, (
                    f"{sheet.name}!{letter} ({tab.cell(workbook.header_row(1), index).value}) "
                    f"holds numbers; silencing text warnings there hides a real mistake")
                checked += 1
    assert checked > 20, checked


# === round six ============================================================================

def test_the_legend_states_the_upset_criteria_rather_than_restating_the_label():
    """A LEGEND MUST DEFINE, NOT RESTATE.

    The first version said "a level 2 upset (the site shades this)", which tells a reader who
    has never seen the site precisely nothing. Marc: "These aren't proper definitions. What is
    the criteria a game score needs to crest?"
    """
    descriptions = {mark: text for _, mark, text in workbook.mark_legend()}
    for level in ("●", "●●", "●●●"):
        assert level in descriptions, level
        text = descriptions[level]
        assert any(ch.isdigit() for ch in text), (
            f"{level} is described as {text!r} — no number, so it is not a definition")
    # Read from the SAME PLACE the legend reads them, so the test cannot drift from the
    # warehouse either: feed a frame, check the words that come back.
    from lib import metrics
    big, blowout = 7, 14
    assert f"{big} or fewer" in descriptions["●"]
    assert f"{big + 1}–{blowout}" in descriptions["●●"]
    assert f"more than {blowout}" in descriptions["●●●"]
    # The bands must not overlap or leave a gap, and the boundary is exclusive.
    assert metrics.upset_ranges(big, blowout) == (
        f"{big} or fewer", f"{big + 1}–{blowout}", f"more than {blowout}")


def test_the_thresholds_come_from_the_data_and_not_from_a_file():
    """R-224. They were read out of dbt_project.yml, which CANNOT WORK IN THE SITE IMAGE:
    `deploy/docker-compose.yml` builds with `context: ./site`, so the repo root is outside the
    build context and the deployed workbook ran on a hardcoded fallback that happened to match.

    They are columns on srv_game now. Feeding two frames proves the legend follows them; a
    literal would sit still.
    """
    import pandas as _pd
    from lib import metrics

    def bands(big, blowout):
        frame = _pd.DataFrame([{metrics.BIG_COLUMN: big, metrics.BLOWOUT_COLUMN: blowout}])
        return {mark: text for _, mark, text in workbook.mark_legend(frame)}

    assert "7 or fewer" in bands(7, 14)["●"]
    assert "8–14" in bands(7, 14)["●●"]
    assert "9 or fewer" in bands(9, 21)["●"]
    assert "10–21" in bands(9, 21)["●●"]

    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    assert metrics.BIG_COLUMN in schedule.fields or True   # the sheet need not SHOW them
    assert not hasattr(workbook, "UPSET_BIG_MARGIN"), (
        "the workbook must not hold a threshold of its own again")


def test_the_legend_matches_how_srv_game_actually_decides_the_level():
    """The wording has to match the SQL, not just the numbers. `srv_game` uses strict `>`, so
    a 7-point win is Level 1 and an exactly-14-point win is Level 2 — the legend's "7 or
    fewer" and "8 to 14" say the same thing, and "more than 14" is the third band."""
    sql = (Path(__file__).resolve().parents[1] / "dbt" / "models" / "serving"
           / "srv_game.sql").read_text()
    assert "> {{ var('upset_margin_blowout') }}" in sql, "the boundary is strict >"
    assert "> {{ var('upset_margin_big') }}" in sql
    descriptions = {mark: text for _, mark, text in workbook.mark_legend()}
    assert "or fewer" in descriptions["●"]
    assert "more than" in descriptions["●●●"]


def test_american_spelling_throughout():
    """Marc: "Use US version of favorite". The column header, the legend and the divergence
    note all say favorite."""
    schedule = next(s for s in workbook._ALL_SHEETS if s.name == "Schedule")
    labels = [label for _, label in schedule.columns]
    assert "Favorite covered" in labels
    assert "Favourite covered" not in labels
    for _, _, text in workbook.mark_legend():
        assert "favourite" not in text.lower(), text
    assert workbook._title_case_verdict("no_favorite") == "No favorite"


def test_the_one_departure_from_marcs_csv_is_recorded_not_silent():
    """His column-order CSV is authoritative, so a label that differs from it must be a
    DECISION with a reason attached, not a typo nobody noticed."""
    assert workbook.CSV_LABEL_OVERRIDES == {"Favourite covered": "Favorite covered"}


def test_under_is_unfilled_and_red_and_over_is_not():
    """Marc, round 6. Filled/open already carries "it happened / it did not"; red carries the
    direction. ▽ is the open form of ▲, so the pair still reads as one system."""
    assert workbook.OVER_MARKS["no"] == "▽"
    assert workbook.OVER_MARKS["yes"] == "▲"
    assert workbook.MARK_COLOURS["▽"] == workbook.RED_MARK_COLOUR
    assert "▲" not in workbook.MARK_COLOURS, "over stays the default colour"
    assert workbook.RED_MARK_COLOUR != workbook.OPEN_MARK_COLOUR, (
        "under must be distinguishable from the other open marks, which is the point")
    # Push is BLUE (Marc, round 10): it is neither a win nor a loss, so it borrows neither
    # side's colour.
    assert workbook.MARK_COLOURS[workbook.PUSH_MARK] == workbook.BLUE_MARK_COLOUR


def test_the_under_colour_passes_contrast_too():
    def luminance(hex_rgb):
        channels = [int(hex_rgb[i:i + 2], 16) / 255 for i in (0, 2, 4)]
        channels = [c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4
                    for c in channels]
        return 0.2126 * channels[0] + 0.7152 * channels[1] + 0.0722 * channels[2]
    contrast = 1.05 / (luminance(workbook.RED_MARK_COLOUR[2:]) + 0.05)
    assert contrast >= 4.5, f"{workbook.RED_MARK_COLOUR} is {contrast:.2f}:1 against white"


def test_every_coloured_mark_is_drawn_the_same_way_in_the_legend_and_the_column(built):
    """Two colours now, so "is it coloured" is no longer a sufficient question — the legend
    has to use the RIGHT one for each glyph or it is a picture of a different mark."""
    _, book, _, _ = built
    index_fonts = {}
    for row in book["Index"].iter_rows():
        for cell in row:
            if cell.value in workbook.MARK_COLOURS:
                index_fonts[cell.value] = cell.font
    for glyph, colour in workbook.MARK_COLOURS.items():
        if glyph in index_fonts:
            assert index_fonts[glyph].color.rgb == colour, glyph


def test_every_mark_shares_one_east_asian_width_class():
    """WHY THE COLUMN LOOKED CROOKED WHEN EVERY CELL WAS CENTRED.

    Marc reported Winner covered as not horizontally centred. Every cell in it is centred —
    verified on all 83 rows of a real file, header included. What was ragged was the glyphs:
    ■ □ ○ ● ▲ ▽ are all East-Asian-Width AMBIGUOUS, and the push mark was ◨ (U+25E8), which
    is NEUTRAL and rare enough that most fonts have no glyph for it. The fallback font brings
    its own advance width, so those cells sat at a different offset and the column read as
    crooked.

    Alignment could never have fixed that, which is why this asserts the width class rather
    than the alignment. Mixing classes in one column is the defect.
    """
    import unicodedata
    widths = {}
    for marks in (workbook.UPSET_MARKS, workbook.COVER_MARKS, workbook.OVER_MARKS):
        for glyph in marks.values():
            for character in glyph:
                widths[character] = unicodedata.east_asian_width(character)
    widths[workbook.NO_DATA_MARK] = unicodedata.east_asian_width(workbook.NO_DATA_MARK)
    classes = set(widths.values())
    assert classes == {"A"}, (
        "the marks mix width classes and will not line up: "
        + repr({c: w for c, w in widths.items() if w != "A"}))


def test_every_mark_lives_in_the_geometric_shapes_block_or_is_the_dash():
    """One block means one designer drew them, which is the practical reason they share
    metrics. A mark borrowed from elsewhere is the next ◨."""
    allowed = set(range(0x25A0, 0x2600))
    # Two deliberate exceptions, both by name so a THIRD cannot arrive unnoticed: the pending
    # dot, and the push equals sign. The block rule exists to keep metrics consistent, and
    # the width test above is what actually enforces that — this one guards against a mark
    # being borrowed from somewhere arbitrary.
    exceptions = {"·", workbook.PUSH_GLYPH}
    for marks in (workbook.UPSET_MARKS, workbook.COVER_MARKS, workbook.OVER_MARKS):
        for glyph in marks.values():
            for character in glyph:
                if character in exceptions:
                    continue
                assert ord(character) in allowed, (
                    f"{character!r} U+{ord(character):04X} is outside Geometric Shapes")


# === round seven ==========================================================================

def test_the_ignored_errors_block_is_a_direct_child_of_the_worksheet(built):
    """IT WASN'T, AND THAT IS WHY THE GREEN CORNERS SURVIVED A WHOLE ROUND.

    The first version anchored on the earliest `<extLst`, and the earliest one in the
    document is NESTED INSIDE the data bar's own `<cfRule>`. So the block went in there:
    well-formed XML, valid zip, opens cleanly — and completely invisible to Excel, which went
    on flagging every record cell.

    Excel does not complain about a mis-nested element; it ignores it. That is worse than a
    repair prompt, because nothing tells you.
    """
    import zipfile as _zip
    from xml.etree import ElementTree
    main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    payload = built[0]
    with _zip.ZipFile(BytesIO(payload)) as archive:
        for name in archive.namelist():
            if not name.startswith("xl/worksheets/sheet"):
                continue
            root = ElementTree.fromstring(archive.read(name))
            anywhere = [e for e in root.iter() if e.tag == main + "ignoredErrors"]
            direct = [c for c in root if c.tag == main + "ignoredErrors"]
            assert len(anywhere) == len(direct), (
                f"{name}: ignoredErrors exists but is nested, so Excel will ignore it")


def test_the_validator_detects_a_mis_nested_top_level_element():
    """A validator that cannot fail is what let this through. Fed the exact shape that
    shipped — ignoredErrors inside a cfRule — it must object."""
    good = (b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b'<sheetData/><hyperlinks/><ignoredErrors/><tableParts/></worksheet>')
    bad = (b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
           b'<sheetData/><conditionalFormatting><cfRule><ignoredErrors/></cfRule>'
           b'</conditionalFormatting><hyperlinks/><tableParts/></worksheet>')
    assert workbook.sheet_order_violations(good) == []
    faults = workbook.sheet_order_violations(bad)
    assert faults and faults[0][0] == "ignoredErrors", faults
    assert "NESTED" in faults[0][1]


def test_ignored_errors_sits_after_hyperlinks_where_the_schema_puts_it(built):
    """The ordering half of the same fix. CT_Worksheet runs
    ...conditionalFormatting, hyperlinks, pageMargins... then ignoredErrors, then tableParts.
    The broken version was before hyperlinks, which is a second way to be wrong."""
    import zipfile as _zip
    payload = built[0]
    with _zip.ZipFile(BytesIO(payload)) as archive:
        body = archive.read("xl/worksheets/sheet2.xml").decode("utf-8")
    ignored, hyperlinks = body.find("<ignoredErrors>"), body.find("<hyperlinks")
    table_parts = body.find("<tableParts")
    assert ignored != -1
    if hyperlinks != -1:
        assert ignored > hyperlinks, "ignoredErrors must follow hyperlinks"
    if table_parts != -1:
        assert ignored < table_parts, "and precede tableParts"


def test_the_record_columns_are_centred(built):
    """Marc: centre them "so there is some padding between the warning tag and the text".
    A won-lost record is a short token, not prose, and the cardinality rule cannot reach it —
    a week of play produces well over five distinct records."""
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    for label in ("Away record", "Home record"):
        index = labels[label]
        assert tab.cell(header + 1, index).alignment.horizontal == "center", label
    assert {"Away record", "Home record"} <= workbook.ALWAYS_CENTRED_LABELS

    # NOT AN EXACT-SET ASSERTION ANY MORE, AND THE REPLACEMENT IS STRICTER ABOUT THE THING
    # THAT ACTUALLY GOES WRONG. Pinning the set stopped it becoming a dumping ground, but it
    # also went red for a deliberate addition (R-262 named the two cover verdicts), which
    # makes it a change-detector rather than a rule. What matters is that every name in it is
    # a label some shipped sheet really has: a stale entry is silent, does nothing, and is
    # exactly what a dumping ground looks like.
    shipped = {label for sheet in workbook.SHEETS for _, label in sheet.columns}
    assert workbook.ALWAYS_CENTRED_LABELS <= shipped, (
        workbook.ALWAYS_CENTRED_LABELS - shipped)


def test_every_mark_cell_names_the_same_font(built):
    """WHY A CENTRED COLUMN LOOKED CROOKED, AND WHY ALIGNMENT COULD NOT FIX IT.

    Measured against the font files Excel actually uses: Calibri and Aptos — its old and
    current defaults — HAVE NO FILLED SQUARE (U+25A0). So □ was drawn by Calibri at 0.604 em
    and ■ by whatever macOS substituted, at its own width, in the same column. Two advance
    widths in one column read as crooked no matter how the cells are aligned.

    Naming a font that carries all seven marks at one width is the fix. If a reader lacks it,
    Excel substitutes ONE font for the whole run, so they still share a width — better than
    substituting per glyph, which is what happens today.
    """
    _, book, _, _ = built
    schedule = next(s for s in workbook.SHEETS if s.name == "Schedule")
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    marked = {dict(schedule.columns)[f] for f in schedule.centred}
    checked = 0
    for label in marked:
        for row in range(header + 1, tab.max_row + 1):
            cell = tab.cell(row, labels[label])
            if cell.value is None:
                continue
            assert cell.font.name == workbook.MARK_FONT_NAME, (
                f"{label} row {row} is {cell.font.name!r}; a mark drawn in the default font "
                f"may fall back per glyph and break the column's alignment")
            checked += 1
    assert checked, "no mark cells were checked, so this proved nothing"


def test_the_legend_glyphs_use_the_same_font_as_the_column(built):
    """Otherwise the legend shows a mark of a different size and shape from the one in the
    sheet, which is the failure this file already fixed once for colour and size."""
    _, book, _, _ = built
    seen = 0
    for row in book["Index"].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value and \
                    cell.value[0] in "●○■□▲▽" + workbook.PUSH_GLYPH:
                assert cell.font.name == workbook.MARK_FONT_NAME, cell.value
                seen += 1
    assert seen >= 7


# === round eight ==========================================================================

def test_push_is_an_equals_sign_not_a_third_kind_of_square():
    """Marc: the push mark "is not very discernable from the filled square that is Covered".

    It was ▣ — a white square containing a small black one — which at 12pt next to ■ is a
    filled square with a hairline round it. The fill states are already spoken for (filled =
    it happened, open = it did not), so a push has to leave that language rather than find a
    third position inside it. ═ says "equal", which is what a push is.
    """
    assert workbook.PUSH_MARK == "══"
    assert workbook.COVER_MARKS["push"] == workbook.PUSH_MARK
    assert workbook.OVER_MARKS["push"] == workbook.PUSH_MARK
    squares = {"■", "□", "▣", "▤", "▥", "▦", "▩", "◧", "◨"}
    assert workbook.PUSH_MARK not in squares, (
        "push must not be another square — that is the confusion being fixed")
    circles = {"●", "○", "◉", "◎", "◐", "◑"}
    assert workbook.PUSH_MARK not in circles


def test_the_push_mark_still_measures_the_same_as_every_other_mark():
    """It comes from Box Drawing rather than Geometric Shapes, which is only acceptable
    because it is metrically identical — 0.6001 em in the pinned font, same as ■ □ ● ○ ▲ ▽.
    Width is the property that keeps the column straight; the block is a proxy for it."""
    import unicodedata
    assert len(set(workbook.PUSH_MARK)) == 1, "one repeated character, not two shapes"
    for character in workbook.PUSH_MARK:
        assert unicodedata.east_asian_width(character) == "A"


def test_the_kickoff_prints_as_month_day_and_time(built):
    """Marc: force `mmm-dd hh:mm`. A reader scanning a week's slate wants "Sep-05 19:00" —
    the year is redundant with the Season column right there and the filename carrying it."""
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    cell = tab.cell(header + 1, labels["Kickoff"])
    assert cell.number_format == "mmm-dd hh:mm"
    # In Excel `mm` is MINUTES after an hour token and MONTH otherwise, so the month must be
    # spelled `mmm` or the format silently means something else.
    assert cell.number_format.startswith("mmm"), "a bare mm here would print minutes"


def test_provenance_timestamps_keep_their_year(built):
    """`As of` and `Line taken` exist to be checked against something OUTSIDE the file. A
    provenance timestamp without a year is not one."""
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    for label in ("As of", "Line taken"):
        assert "yyyy" in tab.cell(header + 1, labels[label]).number_format, label


def test_the_date_column_width_follows_its_format(built):
    """The measurement hardcoded 17 characters for every datetime, which was right for
    `yyyy-mm-dd hh:mm` and would have left the shorter kickoff column five characters too
    wide — R-217's defect arriving from the date side."""
    from openpyxl.utils import get_column_letter
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    kickoff = tab.column_dimensions[get_column_letter(labels["Kickoff"])].width
    as_of = tab.column_dimensions[get_column_letter(labels["As of"])].width
    assert kickoff < as_of, (
        f"Kickoff is {kickoff} and As of is {as_of}; the shorter format must take the "
        f"narrower column")
    assert workbook.rendered_date_width("start_date_et") == len("Sep-05 19:00") + 1


def test_the_push_mark_is_blue_wherever_it_appears(built):
    """Marc, round 10. A push is neither a win nor a loss, so it borrows neither side's
    colour — red is for Under and the open marks, blue is for the outcome that was neither."""
    _, book, _, _ = built
    tab = book["Schedule"]
    header = workbook.header_row(1)
    labels = {tab.cell(header, i).value: i for i in range(1, tab.max_column + 1)}
    seen = 0
    for label in ("Winner covered", "O/U result"):
        for row in range(header + 1, tab.max_row + 1):
            cell = tab.cell(row, labels[label])
            if cell.value == workbook.PUSH_MARK:
                assert cell.font.color.rgb == workbook.BLUE_MARK_COLOUR, cell.coordinate
                seen += 1
    # And in the legend, drawn identically — the failure this file has already fixed twice.
    for row in book["Index"].iter_rows():
        for cell in row:
            if cell.value == workbook.PUSH_MARK:
                assert cell.font.color.rgb == workbook.BLUE_MARK_COLOUR
                seen += 1
    assert seen, "no push mark appeared anywhere, so this proved nothing"


def test_the_colour_constant_is_not_named_after_only_one_of_its_uses():
    """It was UNDER_COLOUR when only Under used it. A constant called "under" that also
    colours the push mark is the kind of small lie that makes the next reader distrust every
    other name in the file."""
    assert not hasattr(workbook, "UNDER_COLOUR")
    users = {g for g, c in workbook.MARK_COLOURS.items() if c == workbook.RED_MARK_COLOUR}
    assert users == {"▽"}


def test_the_push_mark_is_a_doubled_glyph_not_a_new_shape():
    """A single ═ is one short bar and reads as a dash at a glance. Two read unmistakably as
    an equals sign. Repeating the character rather than finding a wider one is the same trick
    the upset levels use, and it means the metrics cannot drift: two glyphs of 0.6001 em."""
    assert workbook.PUSH_MARK == workbook.PUSH_GLYPH * 2
    assert len(workbook.PUSH_MARK) == 2
    assert workbook.PUSH_GLYPH == "═"


def test_the_three_mark_colours_mean_three_different_things():
    """Burnt sienna = it did not happen. Red = the wrong side of the number. Blue = neither
    side won. Three colours, three claims, no overlap — and all three measured against white
    rather than chosen by eye."""
    def contrast(argb):
        channels = [int(argb[2:][i:i + 2], 16) / 255 for i in (0, 2, 4)]
        channels = [c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4
                    for c in channels]
        luminance = 0.2126 * channels[0] + 0.7152 * channels[1] + 0.0722 * channels[2]
        return 1.05 / (luminance + 0.05)

    colours = {workbook.OPEN_MARK_COLOUR, workbook.RED_MARK_COLOUR,
               workbook.BLUE_MARK_COLOUR}
    assert len(colours) == 3, "two of the mark colours are the same value"
    for colour in colours:
        assert contrast(colour) >= 4.5, f"{colour} is {contrast(colour):.2f}:1 on white"


# === R-228: red was on the common outcome =================================================

def _is_reddish(argb: str) -> bool:
    """Red channel clearly dominant. Compares the CHANNELS rather than the hex string, so a
    different shade of the same intent still reads as red."""
    red, green, blue = (int(argb[2:][i:i + 2], 16) for i in (0, 2, 4))
    return red > green + 40 and red > blue + 40


def _is_bluish(argb: str) -> bool:
    red, green, blue = (int(argb[2:][i:i + 2], 16) for i in (0, 2, 4))
    return blue > red + 30


def test_the_home_win_side_of_the_margin_bar_is_not_the_alarming_colour():
    """MEASURED, NOT PREFERRED. `actual_margin` is away minus home, so negative is a home win
    — verified across 106,554 completed games with no exceptions — and that is 64% of FBS
    games since 2024.

    The bar shipped with negative in red, so the most common outcome in the sport was
    alarming on nearly two rows in three. A colour that means "look here" and appears on the
    majority of rows means nothing at all.

    Asserted on the CHANNELS and on the SIDE, not against the constants: the first version of
    this test compared `negativeFillColor` to `DATA_BAR_NEGATIVE`, which follows whatever the
    constant says and passed happily through the swap it was meant to catch.
    """
    assert _is_bluish(workbook.DATA_BAR_NEGATIVE), (
        "negative is a home win, the common case — it must not be the alarming colour")
    assert _is_reddish(workbook.DATA_BAR_POSITIVE), (
        "positive is an away win, the result worth noticing")


def test_the_away_minus_home_scales_run_the_opposite_way_to_the_others():
    """A colour scale asserts good and bad. Fair on a quantity with a direction — a bigger
    edge, a better point differential — and wrong on one that only has a SIDE."""
    assert "predicted_margin" in workbook.REVERSED_SCALE_FIELDS
    assert "predicted_margin" not in workbook.COLOUR_SCALE_FIELDS, (
        "a field in both sets would take whichever branch is written first")
    # The columns that genuinely have a direction keep the ordinary scale.
    assert {"edge_value", "edge_magnitude", "point_differential"} <= workbook.COLOUR_SCALE_FIELDS
    assert not (workbook.COLOUR_SCALE_FIELDS & workbook.REVERSED_SCALE_FIELDS)


def test_a_reversed_scale_puts_red_at_the_top_and_a_normal_one_at_the_bottom(built):
    """Asserted on the BUILT rule rather than on the sets, because the branch that reads them
    is where a reversal actually happens or fails to."""
    _, book, _, _ = built
    tab = book["Schedule"]
    scales = [rule for rule_set in tab.conditional_formatting for rule in rule_set.rules
              if rule.colorScale is not None]
    assert scales, "the Schedule sheet has no colour scale at all"
    for rule in scales:
        first = rule.colorScale.color[0].rgb
        last = rule.colorScale.color[-1].rgb
        # predicted_margin is the only scaled column on this sheet, and it is reversed.
        assert _is_reddish(last) and not _is_reddish(first), (
            f"the scale runs {first} -> {last}; red belongs at the away-win end")


def test_the_index_explains_what_the_bar_colours_mean(built):
    """A bare colour is as undecodable as a bare glyph, and the marks get a legend for exactly
    that reason. It must also say the colours are a DIRECTION rather than a verdict — a red
    bar on a 40-point away win is not a judgement about the game."""
    _, book, _, _ = built
    text = "\n".join(str(c.value) for r in book["Index"].iter_rows() for c in r)
    assert "Margin bar" in text
    assert "home team won" in text and "away team did" in text
    assert "not a judgement" in text


# ==========================================================================================
# THE SCORES SHEET (R-255 … R-259)
# ==========================================================================================

SCORES_CSV = (Path(__file__).resolve().parent / "fixtures"
              / "cfdb_scores_column_order_v2.csv")


def _scores():
    return next(s for s in workbook.SHEETS if s.name == "Scores")


def _marcs_order():
    """[(field, category)] in position order, from Marc's own file under tests/fixtures.

    v2 supersedes the 131-field v1 list, which predates the Market block. It carries the
    CATEGORY as well as the order, which is what lets the test assert the block structure
    rather than only its contents.
    """
    import csv
    with SCORES_CSV.open(encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    assert [r["Position"] for r in rows] == [str(i) for i in range(1, len(rows) + 1)], (
        "the file's Position column is not 1..n in order")
    return [(r["Field"].strip(), r["Category"].strip()) for r in rows]


# The twelve market columns Marc asked for after the first pass (R-260). Named here rather
# than derived from the sheet, so a column silently vanishing from the block fails the test
# instead of shrinking the expectation to match itself.
MARKET_FIELDS = frozenset({
    "spread_final", "total_final", "line_implied_points_final",
    "points_vs_line_implied_final", "ats_margin_final", "covered_final",
    "spread_open", "total_open", "line_implied_points_open",
    "points_vs_line_implied_open", "ats_margin_open", "covered_open",
})


def test_the_sheet_is_exactly_the_order_and_the_categories_marc_asked_for():
    """144 IN, 144 OUT — R-264 is a permutation, and this asserts it against the FILE.

    Order AND category, position by position. A test comparing only the SET would pass on a
    sheet with every column in the wrong place; one comparing only the order would pass with
    the seven ancillary keys painted as Game — the arrangement Marc first sketched, and the
    one that breaks the contiguity invariant.

    Asserted against `tests/fixtures/cfdb_scores_column_order_v2.csv` rather than a list
    retyped from it, because a retyped list agrees with whatever I typed.
    """
    marc = _marcs_order()
    assert len(marc) == 144, len(marc)

    sheet = _scores()
    built = [(field, sheet.field_category[field]) for field, _ in sheet.columns]
    assert built == marc, [
        (i, b, m) for i, (b, m) in enumerate(zip(built, marc), start=1) if b != m][:8]

    # The BLOCKS are what the writer walks, so they are checked too: the flattened list
    # agreeing while the structure itself is wrong is exactly what this catches.
    flattened = [(f, name) for name, fields in workbook.SCORES_BLOCKS for f in fields]
    assert flattened == marc


def test_the_reorder_moved_seven_columns_and_lost_none():
    """THE PERMUTATION, STATED AS A PERMUTATION.

    The test above would also pass if a column were renamed on both sides at once. This one
    names what moved and asserts nothing else did, so a change that drops one field and adds
    another of the same category cannot slip through.
    """
    moved = ["team_id", "conference", "classification", "opponent_team_id",
             "opponent_conference", "is_neutral_site", "is_completed"]
    marc = _marcs_order()
    assert [f for f, c in marc if c == "Ancillary"] == moved

    sheet = _scores()
    rest = [f for f, _ in sheet.columns if f not in moved]
    assert rest == [f for f, c in marc if c != "Ancillary"]
    assert len(rest) == 137 and len(moved) == 7


def test_the_select_list_and_the_display_list_differ_only_where_they_must():
    """`game_no` is computed by the query, so it is displayed and not selected.
    `possession_seconds` is selected and displayed as `possession_minutes` (R-259).
    Those are the only two, and stating them separately from the counts means a compensating
    change on the other side cannot satisfy this.
    """
    marc = {f for f, _ in _marcs_order()}
    sheet = _scores()

    displayed = sheet.fields
    assert len(displayed) == 144 and len(set(displayed)) == 144
    assert set(displayed) == marc

    selected = set(sheet.selected_fields)
    assert len(sheet.selected_fields) == 143
    assert selected - marc == {"possession_seconds"}
    assert marc - selected == {"game_no", "possession_minutes"}
    assert MARKET_FIELDS <= marc, "the market block is no longer in the sheet"


def test_every_scores_category_is_one_contiguous_run_of_columns():
    """THE ENTIRE POINT OF R-258: a reader scanning the header strip sees each band once.

    Marc: "We might need to do some more column shuffling to avoid bounding between those
    categories." His CSV had three breaks — opponent_conference stranded at field 117, and
    both havoc runs cut off from their own side. This is the assertion that they are gone
    and that nobody reintroduces one.
    """
    sheet = _scores()
    seen, previous = [], None
    for field, _ in sheet.columns:
        category = sheet.field_category[field]
        if category != previous:
            assert category not in seen, f"{category} appears in two separate runs"
            seen.append(category)
            previous = category
    assert seen == [name for name, _ in workbook.SCORES_BLOCKS], seen


def test_the_scores_sheet_names_no_special_teams_category():
    """Marc named five categories and srv_game_team supports four.

    There is no punt, kick, field-goal or return column in the view. An empty fifth band
    would be a promise the data cannot keep, and `average_start` — which special teams
    influence — is field position, not a special-teams statistic.
    """
    categories = {name for name, _ in workbook.SCORES_BLOCKS}
    assert not any("special" in name.lower() for name in categories), categories
    assert not any(
        any(token in field for token in ("punt", "kick", "field_goal", "return"))
        for field, _ in _scores().columns)


def test_every_scores_header_fill_is_readable_and_distinguishable():
    """WHITE BOLD ON EVERY FILL, so every fill clears 4.5:1 — AND adjacent bands have to be
    told apart from each other, which contrast against white says nothing about.

    Both halves are asserted because they fail independently: four colours can each be dark
    enough and still be four shades of the same navy. The separation check runs under normal
    vision and both red-green dichromacies, because five blocks of colour read left to right
    is a categorical palette and hue alone collapses under deuteranopia.
    """
    def channels(hexcolour):
        raw = [int(hexcolour[i:i + 2], 16) / 255.0 for i in (0, 2, 4)]
        return [c / 12.92 if c <= 0.04045 else ((c + 0.055) / 1.055) ** 2.4 for c in raw]

    def contrast(hexcolour):
        r, g, b = channels(hexcolour)
        return 1.05 / (0.2126 * r + 0.7152 * g + 0.0722 * b + 0.05)

    def lab(linear):
        matrix = ((0.4124, 0.3576, 0.1805), (0.2126, 0.7152, 0.0722),
                  (0.0193, 0.1192, 0.9505))
        x, y, z = [sum(matrix[i][j] * linear[j] for j in range(3)) for i in range(3)]
        white = (0.95047, 1.0, 1.08883)

        def f(t):
            return t ** (1 / 3) if t > 216 / 24389 else (841 / 108) * t + 4 / 29
        fx, fy, fz = (f(v / w) for v, w in zip((x, y, z), white))
        return (116 * fy - 16, 500 * (fx - fy), 200 * (fy - fz))

    # Vienot 1999 dichromat matrices, applied in linear sRGB.
    dichromats = {
        "protanopia": ((0.11238, 0.88762, 0.0), (0.11238, 0.88762, 0.0),
                       (0.00401, -0.00401, 1.0)),
        "deuteranopia": ((0.29275, 0.70725, 0.0), (0.29275, 0.70725, 0.0),
                         (-0.02234, 0.02234, 1.0)),
    }

    def seen_as(hexcolour, kind):
        linear = channels(hexcolour)
        if kind is None:
            return linear
        matrix = dichromats[kind]
        return [min(1.0, max(0.0, sum(matrix[i][j] * linear[j] for j in range(3))))
                for i in range(3)]

    fills = workbook.CATEGORY_FILLS
    assert set(fills) == {name for name, _ in workbook.SCORES_BLOCKS}
    for name, colour in fills.items():
        assert contrast(colour) >= 4.5, f"{name} #{colour} is {contrast(colour):.2f}:1"

    # OVER THE BANDS, NOT OVER THE DISTINCT COLOURS. The first version compared
    # `set(fills.values())`, which deduplicates — so giving two bands the SAME hex made the
    # collision vanish from the comparison and the test went green on the one arrangement it
    # exists to forbid. R-157, and it survived a round of review before a mutation caught it.
    assert len(set(fills.values())) == len(fills), "two bands share a fill"
    for first, second in itertools.combinations(sorted(fills), 2):
        for kind in (None, "protanopia", "deuteranopia"):
            delta = math.dist(lab(seen_as(fills[first], kind)),
                              lab(seen_as(fills[second], kind)))
            assert delta >= 10.0, (
                f"{first} (#{fills[first]}) and {second} (#{fills[second]}) are only "
                f"{delta:.1f} dE apart under {kind or 'normal vision'} — a just-noticeable "
                f"difference is about 2.3, and two header bands need much more than that")


def test_the_header_fill_written_into_the_file_is_the_category_colour(built):
    """The palette above is a set of constants until something writes it.

    Reads the FILE, not the model: a category whose fill never reached a cell would satisfy
    every assertion in the test above.
    """
    _, book, _, _ = built
    tab = book["Scores"]
    sheet = _scores()
    header = workbook.header_row(1)
    seen = {}
    for index, (field, _) in enumerate(sheet.columns, start=1):
        written = tab.cell(header, index).fill.fgColor.rgb
        expected = workbook.CATEGORY_FILLS[sheet.field_category[field]]
        assert written.endswith(expected), (field, written, expected)
        seen.setdefault(sheet.field_category[field], written)
    # Distinct FILLS, not distinct categories: five categories all painted navy would satisfy
    # a count of the keys and give the reader no boundary anywhere.
    assert len(seen) == len(workbook.SCORES_BLOCKS)
    assert len(set(seen.values())) == len(workbook.SCORES_BLOCKS), seen


# ---- the query's own ordering, executed rather than pattern-matched ----------------------

def _run_scores_sql(rows, sql=None):
    """Execute the REAL Scores query against sqlite, over rows we supply.

    A test that greps the SQL for "case season_type" proves the string is present and
    nothing about what it does. sqlite speaks window functions, named parameters and block
    comments, so the production statement runs unmodified and the ORDER BY is exercised for
    real — which is what makes the negative test below meaningful.
    """
    return _execute_scores_sql(rows, sql or _scores().sql, division="all")


# Columns the query FILTERS on without SELECTING them. The synthetic table needs them or the
# statement will not compile — and each is asserted to be genuinely referenced, so this set
# cannot quietly outlive the predicate it exists for.
SCORES_PREDICATE_ONLY = ("is_fbs_game",)


def _execute_scores_sql(rows, sql, division="all"):
    import sqlite3
    sheet = _scores()
    statement = " ".join(sql.split())
    for name in SCORES_PREDICATE_ONLY:
        assert name in statement, f"{name} is no longer referenced by the Scores query"
    columns = sorted(set(sheet.selected_fields) | set(SCORES_PREDICATE_ONLY))
    connection = sqlite3.connect(":memory:")
    connection.execute(
        "create table srv_game_team (%s)" % ", ".join(f'"{c}"' for c in columns))
    connection.executemany(
        "insert into srv_game_team (%s) values (%s)"
        % (", ".join(f'"{c}"' for c in columns), ", ".join("?" * len(columns))),
        [[row.get(c) for c in columns] for row in rows])
    cursor = connection.execute(statement, {
        "season": 2025, "season_type": "regular", "week": None,
        "division": division, "conference": None})
    names = [d[0] for d in cursor.description]
    return [dict(zip(names, r)) for r in cursor.fetchall()]


def _pair(game_id, week, date, season_type="regular", season=2025):
    """One game as its two rows — away first in the input, so the ORDER BY has to do work."""
    base = {"season": season, "season_type": season_type, "week": week,
            "game_date": date, "game_id": game_id, "is_fbs_game": 1,
            "conference": "SEC", "opponent_conference": "SEC",
            "possession_seconds": 1800}
    return [dict(base, is_home=1, team=f"home{game_id}"),
            dict(base, is_home=0, team=f"away{game_id}")]


def test_the_regular_season_sorts_before_the_postseason():
    """MARC NAMED THE ORDER — "Season (regular then post season)" — WHICH IS THE TELL THAT HE
    EXPECTED IT TO NEED SAYING. Alphabetically 'postseason' < 'regular', so a plain
    `order by season_type` puts January's bowls ahead of September.

    THE FULL STATEMENT CANNOT SHOW THIS, AND THAT IS WORTH KNOWING RATHER THAN WORKING
    AROUND. `season_type` is a scope filter — `where season_type = :season_type` — so every
    row in any one export shares a value and the CASE never fires. It is still correct, and
    it is what makes the order right the day a second season type reaches one sheet. So the
    ordering CONSTANT is executed here, over rows the filter would never let coexist.

    NEGATIVE-TESTED IN THE SAME FUNCTION: the CASE is swapped for the bare column and the
    same rows re-run. If that still came out right, this test would be proving nothing.
    """
    import sqlite3
    rows = [("postseason", 15, "2025-12-20", 2), ("regular", 1, "2025-08-30", 1)]

    def order_by(expression):
        connection = sqlite3.connect(":memory:")
        connection.execute("create table t (season int, season_type text, week int, "
                           "game_date text, game_id int)")
        connection.executemany(
            "insert into t (season_type, week, game_date, game_id) values (?,?,?,?)", rows)
        return [r[0] for r in connection.execute(
            f"select season_type from t order by {expression}")]

    assert order_by(workbook.SCORES_GAME_ORDER) == ["regular", "postseason"]

    naive = workbook.SCORES_GAME_ORDER.replace(
        "case season_type when 'regular' then 1 when 'postseason' then 2 else 3 end",
        "season_type")
    assert naive != workbook.SCORES_GAME_ORDER, "the CASE was not found to remove"
    assert order_by(naive) == ["postseason", "regular"], (
        "removing the CASE did not break the order, so the CASE is not what fixes it")

    # And the constant is what the sheet actually orders by, rather than a copy of it.
    flat = " ".join(_scores().sql.split())
    assert flat.count(workbook.SCORES_GAME_ORDER) == 2, (
        "the sort key and game_no's window must be the same expression")


def test_a_game_is_two_adjacent_rows_with_the_away_row_first():
    """THE GRAIN, ASSERTED. Everything else on this sheet — the banding, the possession sum,
    the row count Marc was told to expect — reads as broken if a pair ever comes apart.
    """
    rows = (_pair(3, 2, "2025-09-06") + _pair(1, 1, "2025-08-30")
            + _pair(2, 1, "2025-08-31"))
    result = _run_scores_sql(rows)
    assert len(result) == 6
    for first, second in zip(result[::2], result[1::2]):
        assert first["game_id"] == second["game_id"], "a game's two rows are not adjacent"
        assert (first["is_home"], second["is_home"]) == (0, 1), "home row came first"


def test_game_no_is_dense_gapless_and_shared_by_both_rows_of_a_game():
    """R-257's whole premise: the band follows a VALUE, so it survives a re-sort.

    Dense and gapless matters because the band is `MOD(game_no,2)` — a gap would put two
    shaded games or two bare games side by side and the boundary would vanish exactly where
    the reader needs it.
    """
    rows = (_pair(9, 3, "2025-09-13") + _pair(4, 1, "2025-08-30")
            + _pair(7, 2, "2025-09-06"))
    result = _run_scores_sql(rows)
    numbers = [r["game_no"] for r in result]
    assert numbers == [1, 1, 2, 2, 3, 3], numbers
    per_game = {}
    for row in result:
        per_game.setdefault(row["game_id"], set()).add(row["game_no"])
    assert all(len(v) == 1 for v in per_game.values()), per_game
    assert sorted(n for s in per_game.values() for n in s) == list(range(1, 4))


def test_the_division_filter_keeps_both_rows_of_a_game_or_neither():
    """WHY `is_fbs_game` HAD TO BE ADDED TO srv_game_team (R-255).

    Narrowing on the team's own `classification` would have kept the FBS side of an
    FBS-vs-FCS game and left the other behind — a game with one row, which quietly breaks
    the grain every other rule here depends on.
    """
    fbs = _pair(1, 1, "2025-08-30")
    other = [dict(row, game_id=2, is_fbs_game=0) for row in _pair(2, 1, "2025-08-31")]
    result = _run_scores_sql(fbs + other)
    assert len(result) == 4

    narrowed = _execute_scores_sql(fbs + other, _scores().sql, division="fbs")
    assert [r["game_id"] for r in narrowed] == [1, 1], narrowed


# ---- what the file itself carries --------------------------------------------------------

def test_the_scores_table_does_not_use_excels_own_row_stripes(built):
    """WITH TWO ROWS PER GAME, `showRowStripes` BANDS THE WRONG UNIT — it shades every away
    row and leaves every home row bare. R-182 set it False for the navy header; it now does
    double duty and this is the line that stops it being switched back on.
    """
    _, book, _, _ = built
    for name in ("Scores", "Schedule"):
        style = book[name].tables[workbook.table_name(name)].tableStyleInfo
        assert style.showRowStripes in (False, None, 0), name


def test_the_scores_banding_is_painted_on_the_cells_and_follows_the_game(built):
    """THE THIRD ATTEMPT AT THIS BAND, AND THE FIRST ONE THE TEST CAN SEE.

    Rounds one and two were conditional-format rules — F2F5F7, then DEE2E4 — and Marc could
    not make out either. The tests covered the rule's formula, the column it read and the
    range it spanned, all structural, none of which is evidence that a fill was DRAWN. A
    painted fill is a real entry in cellXfs and can be read straight back out of the file,
    which is what this asserts.

    Parity comes from `game_no`, so a shaded block is exactly one game. Row parity would band
    the wrong unit entirely: two rows per game means it would shade every away row and leave
    every home row bare.
    """
    _, book, _, _ = built
    tab = book["Scores"]
    sheet = _scores()
    first = workbook.first_data_row(1)
    band_index = sheet.fields.index("game_no") + 1
    last_column = len(sheet.columns)

    seen = {}
    for row in range(first, tab.max_row + 1):
        number = tab.cell(row, band_index).value
        if not isinstance(number, (int, float)):
            continue
        fills = {tab.cell(row, i).fill.fgColor.rgb for i in range(1, last_column + 1)}
        assert len(fills) == 1, f"row {row} is only partly banded: {fills}"
        seen.setdefault(int(number) % 2, set()).add(fills.pop())

    assert set(seen) == {0, 1}, "the fixture has games of only one parity — nothing compared"
    even = seen[0].pop()
    assert even.endswith(workbook.SCORES_BAND_FILL), even
    # And the odd games are LEFT ALONE rather than painted white, so the sheet still prints
    # without a full-bleed background.
    assert not any(f.endswith(workbook.SCORES_BAND_FILL) for f in seen[1]), seen[1]

    # No conditional format is doing this any more — if one comes back, the two mechanisms
    # would disagree the moment a reader sorts the Table.
    formulas = [rule.formula[0] for rules in tab.conditional_formatting
                for rule in rules.rules if rule.formula]
    assert not [f for f in formulas if f.startswith("MOD(")], formulas


def test_the_scores_freeze_keeps_team_and_opponent_and_still_leaves_room(built):
    """The freeze exists so team and opponent stay on screen while the metrics scroll.

    THE FIRST VERSION FROZE THE WHOLE GAME BLOCK, WHICH IS WHAT THE SPEC SAID AND DOES NOT
    FIT. Measured on the real week-2 file, those twenty columns are 200 characters wide
    against the ~110-130 an Excel window shows: the frozen pane filled the window and nothing
    could scroll into view. Twelve columns — through `Opponent` — come to 138.

    ASSERTED ON THE COLUMN BUDGET, NOT ON CHARACTERS, and the difference matters. The `built`
    fixture holds two rows, so its columns are narrower than production's: the whole Game
    block measures 143 characters there against 200 on the real file, and a character
    threshold set from real data would wave the bad freeze through. The column count is the
    same in both, so that is what the budget is expressed in — 12, derived from the real
    measurement and recorded here beside it.
    """
    from openpyxl.utils import get_column_letter
    _, book, _, _ = built
    sheet = _scores()
    tab = book["Scores"]
    first_scrolling = sheet.freeze_column()
    frozen = sheet.columns[:first_scrolling - 1]

    assert {"Team", "Opponent"} <= {label for _, label in frozen}, frozen
    assert all(sheet.field_category[f] == "Game" for f, _ in frozen), (
        "the freeze reaches past the Game block")
    assert len(frozen) <= 12, (
        f"{len(frozen)} frozen columns; 12 is 138 characters on real data and 20 — the whole "
        f"Game block — is 200, against the ~110-130 an Excel window shows")

    expected = f"{get_column_letter(first_scrolling)}{workbook.first_data_row(1)}"
    assert tab.freeze_panes == expected


def test_possession_is_minutes_and_a_games_two_rows_sum_to_sixty():
    """Marc asked for minutes; the choice pays for itself. At 2dp a game's two rows total
    60.00, so every game carries its own arithmetic check and a row that does not pair is
    visible without leaving the sheet.
    """
    sheet = _scores()
    assert "possession_seconds" not in sheet.fields, "both spellings shipped"
    assert dict(sheet.columns)["possession_minutes"] == "Possession min"

    for away, home in ((1546, 2054), (2217, 1383), (1800, 1800)):
        pair = [sheet.value_for("possession_minutes", {"possession_seconds": away}),
                sheet.value_for("possession_minutes", {"possession_seconds": home})]
        assert round(sum(pair), 2) == 60.00, (away, home, pair)
    assert sheet.value_for("possession_minutes", {"possession_seconds": None}) is None


def test_the_scores_number_formats_split_counts_from_decimals():
    """R-259, and the reason the integer set is MEASURED rather than named.

    Every one of these was decided by counting fractional values across 110,879 rows, and
    four of them contradict what the column name suggests. A test that re-derived the rule
    from the names would agree with the names and miss all four.
    """
    sheet = _scores()

    def fmt(field):
        return workbook.number_format(field, sheet.decimals, sheet.integer_fields)

    # Whole in every row, so no decimal point.
    for field in ("first_downs", "penalty_yards", "plays", "line_yards",
                  "offense_plays", "offense_db_havoc_events",
                  "offense_line_yards_total"):
        assert fmt(field) == "#,##0", field

    # Fractional somewhere, so two decimals — including three that READ like counts.
    for field in ("offense_success_rate", "offense_ppa", "havoc_total",
                  "offense_total_havoc_events", "offense_front_seven_havoc_events",
                  "line_yards_average", "possession_minutes"):
        assert fmt(field) == "#,##0.00", field

    # R-216 still owns the separator: a label that happens to be numeric stays bare.
    for field in ("season", "week", "game_id", "team_id", "opponent_team_id"):
        assert fmt(field) == "0", field

    # And the sheet-wide default does NOT leak on to Schedule, where a spread is `-6.5`.
    schedule = next(s for s in workbook.SHEETS if s.name == "Schedule")
    assert schedule.decimals is None
    assert workbook.number_format(
        "spread_current", schedule.decimals, schedule.integer_fields) == "#,##0.0"


def test_every_scores_label_is_unique_and_says_something():
    """131 labels are rule-derived, so a bad rule produces 131 bad labels at once.

    `front_seven` is the one that caught it: substituting word by word turned
    `offense_front_seven_havoc_rate` into "Off front front 7 havoc rate".
    """
    sheet = _scores()
    labels = [label for _, label in sheet.columns]
    assert len(set(labels)) == len(labels), "two columns share a header"
    for field, label in sheet.columns:
        assert label and (label[0].isupper() or label[0].isdigit()), (field, label)
        assert "_" not in label, (field, label)
        words = label.split()
        assert len(words) == len(set(words)), (field, label)
    assert workbook.scores_label("offense_front_seven_havoc_rate") == \
        "Off front 7 havoc rate"
    assert workbook.scores_label("defense_passing_downs_ppa") == "Def passing downs PPA"


def test_the_game_band_is_actually_visible_against_white():
    """THE ASSERTION THAT WAS MISSING WHEN THE BANDING SHIPPED.

    The first pass tested the formula, the column it reads and the range it covers — every
    structural property — and never that the colour could be SEEN. It went out at F2F5F7,
    which is 3.9 dE from white against a just-noticeable difference of about 2.3, and Marc
    reported the separation "missing between every other game". Nothing was wrong with the
    rule; the band was simply below the threshold of perception on his screen.

    A floor of 8 dE is comfortably above JND without going past Marc's 10-20% tint rule for
    print. The ceiling matters too: a band dark enough to fight the text would be a different
    complaint with the same cause.
    """
    def lin(hexcolour):
        return [(c / 255 / 12.92 if c / 255 <= 0.04045
                 else ((c / 255 + 0.055) / 1.055) ** 2.4)
                for c in (int(hexcolour[0:2], 16), int(hexcolour[2:4], 16),
                          int(hexcolour[4:6], 16))]

    def lab(linear):
        matrix = ((0.4124, 0.3576, 0.1805), (0.2126, 0.7152, 0.0722),
                  (0.0193, 0.1192, 0.9505))
        x, y, z = [sum(matrix[i][j] * linear[j] for j in range(3)) for i in range(3)]
        white = (0.95047, 1.0, 1.08883)

        def f(t):
            return t ** (1 / 3) if t > 216 / 24389 else (841 / 108) * t + 4 / 29
        fx, fy, fz = (f(v / w) for v, w in zip((x, y, z), white))
        return (116 * fy - 16, 500 * (fx - fy), 200 * (fy - fz))

    delta = math.dist(lab(lin("FFFFFF")), lab(lin(workbook.SCORES_BAND_FILL)))
    assert delta >= 8.0, (
        f"#{workbook.SCORES_BAND_FILL} is {delta:.1f} dE from white — a just-noticeable "
        f"difference is about 2.3, and a band a reader has to hunt for is not a band")
    assert delta <= 22.0, (
        f"#{workbook.SCORES_BAND_FILL} is {delta:.1f} dE from white, dark enough to compete "
        f"with the text it sits behind")
    # And it is the shade actually written into the file's rule, not a constant beside it.
    assert workbook.SCORES_BAND_FILL != "F2F5F7", "back to the invisible one"


def test_the_market_columns_keep_football_precision_where_football_writes_halves():
    """R-259's own argument, applied inside the sheet that overrides it.

    A spread is `-6.5` everywhere in football and never `-6.50`; that is exactly why the 2dp
    default was not made global. The same reasoning exempts the six market columns quoted in
    halves — and deliberately does NOT exempt the line-implied pair, because halving a total
    produces quarter points and the second digit is real there.
    """
    sheet = _scores()

    def fmt(field):
        return workbook.number_format(field, sheet.decimals, sheet.integer_fields,
                                      sheet.site_precision)

    for field in ("spread_final", "spread_open", "total_final", "total_open",
                  "ats_margin_final", "ats_margin_open"):
        assert fmt(field) == "#,##0.0", field
    for field in ("line_implied_points_final", "line_implied_points_open",
                  "points_vs_line_implied_final", "points_vs_line_implied_open"):
        assert fmt(field) == "#,##0.00", field

    # A quarter point is reachable, which is what makes the second digit necessary rather
    # than decorative: total 52.0 with spread -6.5 implies 29.25 and 22.75.
    assert (52.0 - -6.5) / 2 == 29.25
    assert (52.0 - 6.5) / 2 == 22.75


def test_the_cover_verdicts_read_as_words(built):
    """R-262. Marc: "for Scores, let's transition to Yes/No instead of the glyphs."

    Schedule keeps ■/□ because its verdict columns sit in a dense block that shares one
    legend and one glyph vocabulary. Scores has two cover columns among 144, most of them
    numeric, and a reader arriving at column AC has no reason to have read the Index. A word
    needs no key — and it filters and sorts on something a human can type.

    `push` and `pending` stay as themselves. Folding either into "No" would be a claim about
    a result that does not exist: a push is the bet refunded, and a pending game has not been
    graded at all.
    """
    sheet = _scores()
    for field in ("covered_final", "covered_open"):
        render = sheet.display[field]
        assert render("yes") == "Yes"
        assert render("no") == "No"
        assert render("push") == "Push"
        assert render("pending") == "Pending"

    # No mark survives on this sheet — the glyph vocabulary is Schedule's now.
    assert not sheet.centred, sheet.centred
    _, book, _, _ = built
    tab = book["Scores"]
    body = {str(c.value) for row in tab.iter_rows(min_row=workbook.first_data_row(1))
            for c in row if c.value is not None}
    for glyph in set(workbook.COVER_MARKS.values()) | set(workbook.UPSET_MARKS.values()):
        assert glyph not in body, f"{glyph} is still being written to Scores"

    # THE TWO COLUMNS STILL DIFFER ON ABSENCE, AND THEY SHOULD.
    #
    # On the closing column a null means no line was ever recorded — the no-data dash. On the
    # opening column it means the spread never moved, which the eleven columns beside it say
    # by being empty.
    assert sheet.display["covered_final"](None) == workbook.NO_DATA_MARK
    assert sheet.display["covered_open"](None) is None


def test_the_cover_columns_are_centred_whatever_this_week_contained():
    """A verdict column is a category regardless of how many verdicts happened to occur.

    Left to the cardinality rule this is decided by accident: Yes / No / Push / Pending / –
    is five distinct values against a threshold of "fewer than five", so one week centres and
    the next does not — and the closing and opening columns can disagree with each other in
    the same file, which reads as a defect.
    """
    assert {"Covered", "Covered open"} <= workbook.ALWAYS_CENTRED_LABELS


def test_the_market_block_sits_between_the_result_and_the_box_score():
    """Marc: "These fields should be inserted between Result and 1st Downs." """
    sheet = _scores()
    labels = [label for _, label in sheet.columns]
    market = [i for i, (f, _) in enumerate(sheet.columns)
              if sheet.field_category[f] == "Market"]
    assert labels[min(market) - 1] == "Result"
    assert labels[max(market) + 1] == "1st downs"
    assert len(market) == 12, len(market)
    # The closing half is always populated and comes first; the conditionally-blank half is
    # contiguous after it rather than interleaved.
    fields = [sheet.columns[i][0] for i in market]
    assert all(f.endswith("_final") for f in fields[:6]), fields[:6]
    assert all(f.endswith("_open") for f in fields[6:]), fields[6:]


def test_the_freeze_lands_on_pts_for_and_never_reaches_the_ancillary_block(built):
    """R-265. The site freezes Game # · Date · Team · Pts for; Excel can only freeze a
    contiguous PREFIX, so the sheet freezes everything up to and including the last of those.

    `Pts for` AND NOT `Pts against`, and the reason is the grain: each row is one team, so
    `Pts for` is that team's score and `Pts against` is the other team's — already on screen
    one row away in its own `Pts for`. Freezing both spends a column showing the same two
    numbers twice.

    The Ancillary assertion is the one that earns its place. Those seven keys moved to the far
    right precisely to be out of the reader's way, and a freeze that reached them would put
    them permanently back in front of everything else.
    """
    from openpyxl.utils import get_column_letter
    _, book, _, _ = built
    sheet = _scores()
    first_scrolling = sheet.freeze_column()

    assert book["Scores"].freeze_panes == f"K{workbook.first_data_row(1)}"
    assert get_column_letter(first_scrolling) == "K"

    frozen = sheet.columns[:first_scrolling - 1]
    assert [label for _, label in frozen][-1] == "Pts for"
    assert {"Team", "Opponent", "Pts for"} <= {label for _, label in frozen}
    assert "Pts against" not in {label for _, label in frozen}
    assert all(sheet.field_category[f] == "Game" for f, _ in frozen)
    assert not any(sheet.field_category[f] == "Ancillary" for f, _ in frozen)


def test_the_ancillary_block_is_last_and_holds_the_keys(built):
    """The seven columns earn their quietness from POSITION, not from a muted colour — a
    neutral grey was measured and fails dichromacy against the palette's two low-chroma bands.
    So being last is load-bearing, not incidental.
    """
    sheet = _scores()
    tail = sheet.columns[-7:]
    assert all(sheet.field_category[f] == "Ancillary" for f, _ in tail), tail
    assert [f for f, _ in tail] == [
        "team_id", "conference", "classification", "opponent_team_id",
        "opponent_conference", "is_neutral_site", "is_completed"]

    # And they are painted their own colour in the file, not the Game navy Marc first sketched
    # — which would have been one category in two runs.
    _, book, _, _ = built
    tab = book["Scores"]
    header = workbook.header_row(1)
    for index in range(len(sheet.columns) - 6, len(sheet.columns) + 1):
        assert tab.cell(header, index).fill.fgColor.rgb.endswith(
            workbook.CATEGORY_FILLS["Ancillary"])
    assert workbook.CATEGORY_FILLS["Ancillary"] != workbook.CATEGORY_FILLS["Game"]
